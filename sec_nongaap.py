from __future__ import annotations

import html as html_lib
import io
import math
import re
import threading
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from pathlib import PurePosixPath
from typing import Any, Callable, Iterable, Optional
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup, Tag
from pypdf import PdfReader


SEC_ARCHIVES = "https://www.sec.gov/Archives/edgar/data"
SEC_SUBMISSIONS = "https://data.sec.gov/submissions"
SEC_COMPANY_TICKERS = "https://www.sec.gov/files/company_tickers.json"
SEC_SIC_LIST = "https://www.sec.gov/search-filings/standard-industrial-classification-sic-code-list"

MAX_DOCUMENT_BYTES = 35 * 1024 * 1024
APP_VERSION = "6.0.0"

QUARTER_ORDER = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
QUARTER_NAMES = {
    "Q1": "first quarter",
    "Q2": "second quarter",
    "Q3": "third quarter",
    "Q4": "fourth quarter",
}
MONTH_PATTERN = (
    r"January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan\.?|Feb\.?|Mar\.?|Apr\.?|Jun\.?|Jul\.?|Aug\.?|Sep\.?|Sept\.?|Oct\.?|Nov\.?|Dec\."
)


@dataclass(frozen=True)
class FetchedResource:
    content: bytes
    content_type: str
    url: str


class SecClient:
    """Small SEC client with identification, rate limiting, retries, and in-process caching."""

    def __init__(
        self,
        contact_email: str,
        app_name: str = "SEC Non-GAAP Reconciliation Explorer",
        minimum_interval_seconds: float = 0.12,
    ) -> None:
        contact_email = (contact_email or "").strip()
        self.contact_email = contact_email
        self.user_agent = f"{app_name}/{APP_VERSION} {contact_email or 'contact-not-configured'}"
        self.minimum_interval_seconds = max(0.11, float(minimum_interval_seconds))
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": self.user_agent,
                "Accept-Encoding": "gzip, deflate",
                "Accept": "*/*",
            }
        )
        self._last_request_at = 0.0
        self._lock = threading.Lock()
        self._cache: dict[str, FetchedResource] = {}
        self._json_cache: dict[str, Any] = {}

    @staticmethod
    def valid_contact(contact_email: str) -> bool:
        return bool(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", (contact_email or "").strip()))

    def _wait_for_slot(self) -> None:
        with self._lock:
            elapsed = time.monotonic() - self._last_request_at
            if elapsed < self.minimum_interval_seconds:
                time.sleep(self.minimum_interval_seconds - elapsed)
            self._last_request_at = time.monotonic()

    def get_bytes(self, url: str, timeout: int = 60) -> FetchedResource:
        if url in self._cache:
            return self._cache[url]

        last_error: Optional[Exception] = None
        for attempt in range(4):
            try:
                self._wait_for_slot()
                response = self.session.get(url, timeout=timeout, allow_redirects=True)
                if response.status_code in {429, 500, 502, 503, 504}:
                    wait = min(8.0, 0.8 * (2**attempt))
                    time.sleep(wait)
                    continue
                response.raise_for_status()
                content = response.content
                if len(content) > MAX_DOCUMENT_BYTES:
                    raise ValueError(
                        f"SEC document is larger than the {MAX_DOCUMENT_BYTES // (1024 * 1024)} MB app limit."
                    )
                resource = FetchedResource(
                    content=content,
                    content_type=(response.headers.get("Content-Type") or "").split(";")[0].lower(),
                    url=response.url,
                )
                self._cache[url] = resource
                if response.url != url:
                    self._cache[response.url] = resource
                return resource
            except Exception as exc:  # pragma: no cover - network behavior
                last_error = exc
                if attempt < 3:
                    time.sleep(min(8.0, 0.8 * (2**attempt)))

        if last_error:
            raise last_error
        raise RuntimeError(f"Unable to retrieve {url}")

    def get_text(self, url: str, timeout: int = 60) -> str:
        resource = self.get_bytes(url, timeout=timeout)
        for encoding in ("utf-8", "windows-1252", "latin-1"):
            try:
                return resource.content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return resource.content.decode("utf-8", errors="replace")

    def get_json(self, url: str, timeout: int = 60) -> Any:
        if url in self._json_cache:
            return self._json_cache[url]
        resource = self.get_bytes(url, timeout=timeout)
        try:
            data = requests.models.complexjson.loads(resource.content.decode("utf-8"))
        except Exception:
            data = requests.models.complexjson.loads(resource.content.decode("utf-8", errors="replace"))
        self._json_cache[url] = data
        return data


# ---------------------------------------------------------------------------
# SEC issuer and filing metadata
# ---------------------------------------------------------------------------


def clean_space(value: Any) -> str:
    return re.sub(r"\s+", " ", html_lib.unescape(str(value or ""))).strip()


def safe_int(value: Any) -> Optional[int]:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return None
        return int(float(str(value).replace(",", "").strip()))
    except Exception:
        return None


def safe_date(value: Any) -> Optional[date]:
    try:
        if value is None or pd.isna(value):
            return None
        return pd.Timestamp(value).date()
    except Exception:
        return None


def accession_folder(accession: str) -> str:
    return re.sub(r"[^0-9]", "", str(accession or ""))


def filing_document_url(cik: int, accession: str, document: str) -> str:
    return f"{SEC_ARCHIVES}/{int(cik)}/{accession_folder(accession)}/{document}"


def filing_index_url(cik: int, accession: str) -> str:
    return f"{SEC_ARCHIVES}/{int(cik)}/{accession_folder(accession)}/{accession}-index.htm"


def filing_submission_url(cik: int, accession: str) -> str:
    return f"{SEC_ARCHIVES}/{int(cik)}/{accession_folder(accession)}/{accession}.txt"


def columnar_to_frame(payload: Any) -> pd.DataFrame:
    """Convert SEC columnar filing JSON into a stable DataFrame."""
    if not isinstance(payload, dict) or not payload:
        return pd.DataFrame()

    arrays = {key: value for key, value in payload.items() if isinstance(value, list)}
    if not arrays:
        return pd.DataFrame()

    max_len = max((len(value) for value in arrays.values()), default=0)
    padded: dict[str, list[Any]] = {}
    for key, values in arrays.items():
        padded[key] = list(values) + [None] * (max_len - len(values))
    frame = pd.DataFrame(padded)

    expected = [
        "accessionNumber",
        "filingDate",
        "reportDate",
        "acceptanceDateTime",
        "act",
        "form",
        "fileNumber",
        "filmNumber",
        "items",
        "coreg",
        "size",
        "isXBRL",
        "isInlineXBRL",
        "primaryDocument",
        "primaryDocDescription",
    ]
    for column in expected:
        if column not in frame.columns:
            frame[column] = pd.NA

    for column in ("filingDate", "reportDate"):
        frame[column] = pd.to_datetime(frame[column], errors="coerce")
    return frame


def load_company_submissions(client: SecClient, cik: int) -> dict[str, Any]:
    return client.get_json(f"{SEC_SUBMISSIONS}/CIK{int(cik):010d}.json")


def load_all_filings(
    client: SecClient,
    submissions: dict[str, Any],
    years_back: int = 4,
    max_history_files: int = 4,
) -> pd.DataFrame:
    recent_payload = (submissions.get("filings") or {}).get("recent") or {}
    frames = [columnar_to_frame(recent_payload)]

    cutoff = date.today() - timedelta(days=366 * max(2, years_back))
    history_refs = (submissions.get("filings") or {}).get("files") or []
    loaded = 0
    for ref in history_refs:
        if loaded >= max_history_files:
            break
        if not isinstance(ref, dict) or not ref.get("name"):
            continue
        filing_to = safe_date(ref.get("filingTo"))
        filing_from = safe_date(ref.get("filingFrom"))
        if filing_to and filing_to < cutoff and frames[0] is not None and not frames[0].empty:
            continue
        if filing_from and filing_from > date.today() + timedelta(days=7):
            continue
        history = client.get_json(f"{SEC_SUBMISSIONS}/{ref['name']}")
        payload = (history.get("filings") or {}).get("recent") if isinstance(history, dict) else None
        if not payload:
            payload = history
        frames.append(columnar_to_frame(payload))
        loaded += 1

    valid_frames = [frame for frame in frames if frame is not None and not frame.empty]
    if not valid_frames:
        return columnar_to_frame({})
    combined = pd.concat(valid_frames, ignore_index=True, sort=False)
    combined = combined.drop_duplicates(subset=["accessionNumber"], keep="first")
    combined = combined.sort_values("filingDate", ascending=False, na_position="last").reset_index(drop=True)
    return combined


def company_record(submissions: dict[str, Any]) -> dict[str, Any]:
    return {
        "cik": int(submissions.get("cik") or 0),
        "name": submissions.get("name") or "",
        "ticker": ", ".join(submissions.get("tickers") or []),
        "exchange": ", ".join(submissions.get("exchanges") or []),
        "sic": str(submissions.get("sic") or ""),
        "sic_description": submissions.get("sicDescription") or "",
        "fiscal_year_end": str(submissions.get("fiscalYearEnd") or ""),
    }


def load_company_tickers(client: SecClient) -> pd.DataFrame:
    data = client.get_json(SEC_COMPANY_TICKERS)
    rows: list[dict[str, Any]] = []
    for item in (data or {}).values():
        rows.append(
            {
                "cik": int(item.get("cik_str") or 0),
                "ticker": clean_space(item.get("ticker")).upper(),
                "name": clean_space(item.get("title")),
            }
        )
    return pd.DataFrame(rows).drop_duplicates("cik")


def load_sic_codes(client: SecClient) -> pd.DataFrame:
    html = client.get_text(SEC_SIC_LIST)
    try:
        tables = pd.read_html(io.StringIO(html))
    except ValueError:
        return pd.DataFrame(columns=["sic", "industry_title"])
    for table in tables:
        normalized = [clean_space(col).lower().replace(" ", "_") for col in table.columns]
        table.columns = normalized
        sic_col = next((col for col in normalized if "sic" in col), None)
        title_col = next((col for col in normalized if "industry" in col and "title" in col), None)
        if sic_col and title_col:
            result = table[[sic_col, title_col]].rename(columns={sic_col: "sic", title_col: "industry_title"})
            result["sic"] = result["sic"].astype(str).str.extract(r"(\d{4})", expand=False)
            result["industry_title"] = result["industry_title"].map(clean_space)
            return result.dropna().drop_duplicates().reset_index(drop=True)
    return pd.DataFrame(columns=["sic", "industry_title"])


def browse_edgar_sic(client: SecClient, sic_code: str, count: int = 100) -> pd.DataFrame:
    url = (
        "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany"
        f"&sic={re.sub(r'[^0-9]', '', sic_code)}&owner=exclude&count={int(count)}"
    )
    soup = BeautifulSoup(client.get_text(url), "html.parser")
    rows: list[dict[str, Any]] = []
    for table_row in soup.find_all("tr"):
        link = table_row.find("a", href=re.compile(r"CIK=", re.I))
        if not link:
            continue
        match = re.search(r"CIK=(\d+)", link.get("href", ""), re.I)
        cells = [clean_space(cell.get_text(" ", strip=True)) for cell in table_row.find_all("td")]
        if match:
            name = cells[1] if len(cells) > 1 else clean_space(link.get_text(" ", strip=True))
            rows.append({"cik": int(match.group(1)), "name": name})
    return pd.DataFrame(rows).drop_duplicates("cik") if rows else pd.DataFrame(columns=["cik", "name"])


def search_companies(
    client: SecClient,
    query: str = "",
    sic: str = "",
    industry: str = "",
    limit: int = 75,
) -> pd.DataFrame:
    universe = load_company_tickers(client).copy()
    query_norm = clean_space(query).upper()

    if query_norm:
        ticker = universe["ticker"].fillna("").str.upper()
        name = universe["name"].fillna("").str.upper()
        matched = universe[ticker.str.contains(re.escape(query_norm), na=False) | name.str.contains(re.escape(query_norm), na=False)].copy()
        matched["search_rank"] = 50
        matched.loc[ticker.loc[matched.index] == query_norm, "search_rank"] = 0
        matched.loc[ticker.loc[matched.index].str.startswith(query_norm), "search_rank"] = 10
        matched.loc[name.loc[matched.index] == query_norm, "search_rank"] = 5
        matched.loc[name.loc[matched.index].str.startswith(query_norm), "search_rank"] = 20
        universe = matched
    else:
        universe["search_rank"] = 50

    sic_norm = re.sub(r"[^0-9]", "", sic or "")
    if sic_norm:
        sic_companies = browse_edgar_sic(client, sic_norm)
        universe = universe[universe["cik"].isin(sic_companies["cik"])]

    industry_norm = clean_space(industry)
    if industry_norm:
        sic_codes = load_sic_codes(client)
        matching_codes = sic_codes[
            sic_codes["industry_title"].str.contains(re.escape(industry_norm), case=False, na=False)
        ].head(15)
        company_frames: list[pd.DataFrame] = []
        for code in matching_codes["sic"].tolist():
            try:
                company_frames.append(browse_edgar_sic(client, code))
            except requests.RequestException:
                continue
        if company_frames:
            allowed = pd.concat(company_frames, ignore_index=True)["cik"].drop_duplicates()
            universe = universe[universe["cik"].isin(allowed)]
        else:
            universe = universe.iloc[0:0]

    return (
        universe.sort_values(["search_rank", "ticker", "name"])
        .drop(columns=["search_rank"], errors="ignore")
        .head(limit)
        .reset_index(drop=True)
    )


# ---------------------------------------------------------------------------
# Fiscal-period anchors from 10-Q / 10-K metadata only
# ---------------------------------------------------------------------------


def _xbrl_named_value(soup: BeautifulSoup, suffix: str) -> str:
    pattern = re.compile(rf"(?:^|:){re.escape(suffix)}$", re.I)
    for tag in soup.find_all(attrs={"name": pattern}):
        value = clean_space(tag.get_text(" ", strip=True))
        if value:
            return value
    return ""


def parse_dei_focus(document_html: str) -> tuple[Optional[int], str, Optional[date]]:
    soup = BeautifulSoup(document_html, "html.parser")
    fy_text = _xbrl_named_value(soup, "DocumentFiscalYearFocus")
    fp_text = _xbrl_named_value(soup, "DocumentFiscalPeriodFocus")
    end_text = _xbrl_named_value(soup, "DocumentPeriodEndDate")

    if not fy_text:
        match = re.search(
            r"name=[\"'][^\"']*DocumentFiscalYearFocus[\"'][^>]*>(?:<[^>]+>)*\s*(\d{4})",
            document_html,
            re.I,
        )
        fy_text = match.group(1) if match else ""
    if not fp_text:
        match = re.search(
            r"name=[\"'][^\"']*DocumentFiscalPeriodFocus[\"'][^>]*>(?:<[^>]+>)*\s*(Q[1-4]|FY)",
            document_html,
            re.I,
        )
        fp_text = match.group(1) if match else ""

    fiscal_year = safe_int(re.search(r"\d{4}", fy_text).group(0)) if re.search(r"\d{4}", fy_text) else None
    fiscal_period = clean_space(fp_text).upper()
    if fiscal_period not in {"Q1", "Q2", "Q3", "Q4", "FY"}:
        fiscal_period = ""
    period_end = safe_date(end_text)
    return fiscal_year, fiscal_period, period_end


def _parse_fye(fiscal_year_end: str) -> tuple[int, int]:
    value = re.sub(r"[^0-9]", "", fiscal_year_end or "")
    if len(value) != 4:
        return 12, 31
    month, day = int(value[:2]), int(value[2:])
    if not 1 <= month <= 12:
        month = 12
    if not 1 <= day <= 31:
        day = 28
    return month, day


def _valid_target_date(year: int, month: int, day: int) -> date:
    while day >= 28:
        try:
            return date(year, month, day)
        except ValueError:
            day -= 1
    return date(year, month, day)


def fiscal_year_end_date_for_report(report_date: date, fiscal_year_end: str) -> date:
    month, day = _parse_fye(fiscal_year_end)
    candidates = [
        _valid_target_date(report_date.year - 1, month, day),
        _valid_target_date(report_date.year, month, day),
        _valid_target_date(report_date.year + 1, month, day),
    ]
    nearest = min(candidates, key=lambda item: abs((item - report_date).days))
    if abs((nearest - report_date).days) <= 21:
        return nearest
    future = [item for item in candidates if item >= report_date]
    return min(future) if future else candidates[-1]


def guess_fiscal_year(report_date: date, fiscal_year_end: str) -> int:
    target = fiscal_year_end_date_for_report(report_date, fiscal_year_end)
    fye_month, _ = _parse_fye(fiscal_year_end)
    return target.year - 1 if fye_month <= 2 else target.year


def guess_fiscal_quarter(report_date: date, fiscal_year_end: str) -> str:
    target = fiscal_year_end_date_for_report(report_date, fiscal_year_end)
    target_ts = pd.Timestamp(target)
    candidates = {
        "Q1": (target_ts - pd.DateOffset(months=9)).date(),
        "Q2": (target_ts - pd.DateOffset(months=6)).date(),
        "Q3": (target_ts - pd.DateOffset(months=3)).date(),
        "Q4": target,
    }
    return min(candidates, key=lambda key: abs((candidates[key] - report_date).days))


def build_period_anchors(
    client: SecClient,
    cik: int,
    filings: pd.DataFrame,
    fiscal_year_end: str,
    max_periodic_filings: int = 16,
    progress: Optional[Callable[[str], None]] = None,
) -> pd.DataFrame:
    periodic = filings[filings["form"].isin(["10-Q", "10-K"])].copy()
    periodic = periodic.dropna(subset=["accessionNumber", "primaryDocument"])
    periodic = periodic.sort_values("filingDate", ascending=False).head(max_periodic_filings)

    anchors: list[dict[str, Any]] = []
    for _, filing in periodic.iterrows():
        accession = clean_space(filing.get("accessionNumber"))
        document = clean_space(filing.get("primaryDocument"))
        form = clean_space(filing.get("form"))
        report_date = safe_date(filing.get("reportDate"))
        filing_date = safe_date(filing.get("filingDate"))
        document_url = filing_document_url(cik, accession, document)
        fiscal_year: Optional[int] = None
        fiscal_period = ""
        xbrl_period_end: Optional[date] = None

        if progress:
            progress(f"Reading fiscal metadata from {form} filed {filing_date or ''}...")
        try:
            filing_html = client.get_text(document_url)
            fiscal_year, fiscal_period, xbrl_period_end = parse_dei_focus(filing_html)
        except Exception:
            pass

        period_end = xbrl_period_end or report_date
        if not period_end:
            continue
        if not fiscal_year:
            fiscal_year = guess_fiscal_year(period_end, fiscal_year_end)
        if not fiscal_period:
            fiscal_period = "FY" if form == "10-K" else guess_fiscal_quarter(period_end, fiscal_year_end)

        quarter = "Q4" if fiscal_period == "FY" else fiscal_period
        if quarter not in QUARTER_ORDER:
            quarter = guess_fiscal_quarter(period_end, fiscal_year_end)

        anchors.append(
            {
                "fiscal_year": int(fiscal_year),
                "fiscal_quarter": quarter,
                "period_end": period_end,
                "periodic_form": form,
                "periodic_filing_date": filing_date,
                "periodic_accession": accession,
                "periodic_document": document,
                "periodic_url": document_url,
                "periodic_index_url": filing_index_url(cik, accession),
                "metadata_source": "Inline XBRL" if xbrl_period_end or fiscal_period else "Date inference",
            }
        )

    if not anchors:
        return pd.DataFrame(
            columns=[
                "fiscal_year",
                "fiscal_quarter",
                "period_end",
                "periodic_form",
                "periodic_filing_date",
                "periodic_accession",
                "periodic_document",
                "periodic_url",
                "periodic_index_url",
                "metadata_source",
            ]
        )

    frame = pd.DataFrame(anchors)
    frame["quarter_order"] = frame["fiscal_quarter"].map(QUARTER_ORDER)
    frame = frame.sort_values(
        ["fiscal_year", "quarter_order", "periodic_filing_date"],
        ascending=[False, False, False],
    )
    frame = frame.drop_duplicates(["fiscal_year", "fiscal_quarter"], keep="first")
    return frame.drop(columns=["quarter_order"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# 8-K matching and exhibit discovery
# ---------------------------------------------------------------------------


def classify_document_role(document: str, description: str, doc_type: str) -> str:
    text = " ".join([document, description, doc_type]).lower()
    extension = PurePosixPath(urlparse(document).path).suffix.lower()
    if re.search(r"press release|news release|earnings release|results release", text):
        return "Press release"
    if re.search(r"investor presentation|earnings presentation|financial results presentation|slide|presentation|deck", text):
        return "Investor presentation"
    if re.search(r"financial supplement|earnings supplement|supplemental data|data tables|financial tables", text):
        return "Financial supplement"
    if extension == ".pdf" and re.search(r"99(?:\.|$)|exhibit", text):
        return "Investor presentation"
    if re.match(r"EX-99(?:\.|$)", doc_type.upper()):
        return "Other EX-99 exhibit"
    return "Other"


def parse_filing_index(index_html: str, base_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(index_html, "html.parser")
    documents: list[dict[str, Any]] = []

    for table in soup.find_all("table"):
        header_cells = [clean_space(cell.get_text(" ", strip=True)).lower() for cell in table.find_all("th")]
        header_text = " | ".join(header_cells)
        if "document" not in header_text or "type" not in header_text:
            continue
        for row in table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) < 3:
                continue
            values = [clean_space(cell.get_text(" ", strip=True)) for cell in cells]
            link = row.find("a", href=True)
            if not link:
                continue
            document = clean_space(link.get_text(" ", strip=True))
            href = link.get("href", "")
            if href.startswith("/ixviewer/doc/action") or href.startswith("/ix?doc="):
                match = re.search(r"(?:doc=)([^&]+)", href)
                href = match.group(1) if match else href
            url = urljoin(base_url, href)

            # SEC document tables normally use Seq, Description, Document, Type, Size.
            description = values[1] if len(values) >= 2 else ""
            doc_type = values[3] if len(values) >= 4 else ""
            size = values[4] if len(values) >= 5 else ""
            documents.append(
                {
                    "sequence": values[0] if values else "",
                    "description": description,
                    "document": document,
                    "doc_type": doc_type,
                    "size": size,
                    "url": url,
                    "role": classify_document_role(document, description, doc_type),
                }
            )

    if documents:
        unique: list[dict[str, Any]] = []
        seen: set[str] = set()
        for document in documents:
            if document["url"] not in seen:
                unique.append(document)
                seen.add(document["url"])
        return unique

    # Fallback for older or unusual filing-index markup.
    for row in soup.find_all("tr"):
        link = row.find("a", href=True)
        cells = row.find_all(["td", "th"])
        if not link or len(cells) < 2:
            continue
        values = [clean_space(cell.get_text(" ", strip=True)) for cell in cells]
        description = " | ".join(values)
        document = clean_space(link.get_text(" ", strip=True))
        url = urljoin(base_url, link.get("href", ""))
        type_match = re.search(r"EX-99(?:\.\d+)?", description, re.I)
        doc_type = type_match.group(0).upper() if type_match else ""
        documents.append(
            {
                "sequence": values[0] if values else "",
                "description": description,
                "document": document,
                "doc_type": doc_type,
                "size": "",
                "url": url,
                "role": classify_document_role(document, description, doc_type),
            }
        )
    return documents


def relevant_exhibits(documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for document in documents:
        doc_type = clean_space(document.get("doc_type")).upper()
        role = document.get("role") or "Other"
        extension = PurePosixPath(urlparse(document.get("url") or "").path).suffix.lower()
        if doc_type.startswith("EX-99") or role in {
            "Press release",
            "Investor presentation",
            "Financial supplement",
            "Other EX-99 exhibit",
        }:
            if extension in {".htm", ".html", ".txt", ".pdf", ""}:
                candidates.append(document)

    role_rank = {
        "Press release": 0,
        "Financial supplement": 1,
        "Investor presentation": 2,
        "Other EX-99 exhibit": 3,
        "Other": 4,
    }
    return sorted(candidates, key=lambda item: (role_rank.get(item.get("role") or "Other", 9), item.get("sequence") or ""))


def html_plain_text(document_html: str) -> str:
    soup = BeautifulSoup(document_html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    return re.sub(r"[ \t]+", " ", soup.get_text("\n", strip=True))


def period_reference_score(text: str, anchor: dict[str, Any]) -> int:
    lower = clean_space(text).lower()
    score = 0
    period_end = safe_date(anchor.get("period_end"))
    fiscal_year = safe_int(anchor.get("fiscal_year"))
    quarter = clean_space(anchor.get("fiscal_quarter")).upper()

    if period_end:
        full_month = period_end.strftime("%B %d, %Y").replace(" 0", " ").lower()
        short_month = period_end.strftime("%b %d, %Y").replace(" 0", " ").lower()
        if full_month in lower or short_month in lower or period_end.isoformat() in lower:
            score += 80
    if fiscal_year and quarter in QUARTER_NAMES:
        if f"{QUARTER_NAMES[quarter]} {fiscal_year}" in lower:
            score += 55
        if re.search(rf"\b{re.escape(quarter.lower())}\s*['’]?(?:{str(fiscal_year)[-2:]}|{fiscal_year})\b", lower):
            score += 45
        if quarter == "Q4" and f"full year {fiscal_year}" in lower:
            score += 25
    return score


def _metadata_8k_candidates(filings: pd.DataFrame, anchor: dict[str, Any]) -> pd.DataFrame:
    forms = filings[filings["form"].isin(["8-K", "8-K/A"])].copy()
    if forms.empty:
        return forms
    period_end = pd.Timestamp(anchor["period_end"])
    periodic_filed = pd.Timestamp(anchor["periodic_filing_date"])

    primary = forms[
        (forms["filingDate"] >= period_end - pd.Timedelta(days=3))
        & (forms["filingDate"] <= periodic_filed + pd.Timedelta(days=5))
    ].copy()
    if primary.empty:
        primary = forms[
            (forms["filingDate"] >= periodic_filed - pd.Timedelta(days=45))
            & (forms["filingDate"] <= periodic_filed + pd.Timedelta(days=14))
        ].copy()
    return primary


def match_earnings_8k(
    client: SecClient,
    cik: int,
    filings: pd.DataFrame,
    anchor: dict[str, Any],
    progress: Optional[Callable[[str], None]] = None,
) -> Optional[dict[str, Any]]:
    candidates = _metadata_8k_candidates(filings, anchor)
    if candidates.empty:
        return None

    scored: list[dict[str, Any]] = []
    periodic_filing_date = safe_date(anchor.get("periodic_filing_date"))
    for _, filing in candidates.iterrows():
        accession = clean_space(filing.get("accessionNumber"))
        primary_document = clean_space(filing.get("primaryDocument"))
        if not accession or not primary_document:
            continue
        filing_date = safe_date(filing.get("filingDate"))
        items = clean_space(filing.get("items"))
        score = 0
        reasons: list[str] = []

        if re.search(r"(?:^|[,;\s])2\.02(?:$|[,;\s])", items):
            score += 120
            reasons.append("Item 2.02")
        if re.search(r"(?:^|[,;\s])7\.01(?:$|[,;\s])", items):
            score += 20
            reasons.append("Item 7.01")
        if re.search(r"(?:^|[,;\s])9\.01(?:$|[,;\s])", items):
            score += 10
        if filing_date and periodic_filing_date:
            gap = abs((periodic_filing_date - filing_date).days)
            score += max(0, 35 - gap)
            if filing_date <= periodic_filing_date:
                score += 10

        index = filing_index_url(cik, accession)
        primary_url = filing_document_url(cik, accession, primary_document)
        if progress:
            progress(f"Checking 8-K filed {filing_date or ''} for the earnings release...")
        try:
            primary_text = html_plain_text(client.get_text(primary_url))
        except Exception:
            primary_text = ""
        primary_lower = primary_text.lower()
        if "item 2.02" in primary_lower or "results of operations and financial condition" in primary_lower:
            score += 55
            reasons.append("earnings-results item text")
        if re.search(r"financial results|earnings release|press release announcing", primary_lower):
            score += 30
            reasons.append("financial-results language")
        score += period_reference_score(primary_text, anchor)

        try:
            documents = parse_filing_index(client.get_text(index), index)
        except Exception:
            documents = []
        exhibits = relevant_exhibits(documents)
        if any(item.get("role") == "Press release" for item in exhibits):
            score += 35
            reasons.append("press-release exhibit")
        if any(item.get("role") == "Investor presentation" for item in exhibits):
            score += 12
            reasons.append("presentation exhibit")

        # Inspect the most likely release to distinguish multiple 8-Ks filed near the same date.
        preview_score = 0
        for exhibit in exhibits[:3]:
            try:
                resource = client.get_bytes(exhibit["url"])
                if resource.content.startswith(b"%PDF") or resource.content_type == "application/pdf":
                    preview = extract_pdf_text(resource.content, max_pages=4)[0]
                else:
                    preview = html_plain_text(decode_document(resource.content))
                local_score = period_reference_score(preview[:50000], anchor)
                if re.search(r"reconciliation.{0,80}(?:gaap|non-gaap)|gaap.{0,80}non-gaap", preview, re.I | re.S):
                    local_score += 20
                if re.search(r"financial results|reports .{0,40}quarter|earnings", preview, re.I):
                    local_score += 20
                preview_score = max(preview_score, local_score)
            except Exception:
                continue
        score += preview_score

        scored.append(
            {
                "score": score,
                "reasons": ", ".join(reasons),
                "form": clean_space(filing.get("form")),
                "filing_date": filing_date,
                "report_date": safe_date(filing.get("reportDate")),
                "items": items,
                "accession": accession,
                "primary_document": primary_document,
                "primary_url": primary_url,
                "index_url": index,
                "documents": documents,
                "exhibits": exhibits,
            }
        )

    if not scored:
        return None
    best = max(scored, key=lambda item: item["score"])
    return best if best["score"] >= 70 else None


# ---------------------------------------------------------------------------
# Structured reconciliation extraction
# ---------------------------------------------------------------------------


METRIC_PATTERNS: list[tuple[str, str]] = [
    ("Adjusted EBITDA margin", r"\badjusted\s+ebitda\s+margin\b"),
    ("Adjusted EBITDA", r"\badjusted\s+ebitda\b"),
    ("Adjusted EBIT margin", r"\badjusted\s+ebit\s+margin\b"),
    ("Adjusted EBIT", r"\badjusted\s+ebit\b"),
    ("Adjusted operating income margin", r"\badjusted\s+operating\s+(?:income|profit)\s+margin\b"),
    ("Adjusted operating income", r"\badjusted\s+operating\s+(?:income|profit)\b"),
    ("Non-GAAP operating margin", r"\bnon[- ]gaap\s+operating\s+margin\b"),
    ("Non-GAAP operating expenses", r"\bnon[- ]gaap\s+(?:total\s+)?operating\s+expenses?\b"),
    ("Adjusted net income margin", r"\badjusted\s+net\s+income\s+margin\b"),
    ("Adjusted net income", r"\badjusted\s+net\s+income\b"),
    ("Adjusted diluted EPS", r"\badjusted\s+(?:diluted\s+)?(?:eps|earnings\s+per\s+share)\b"),
    ("Non-GAAP diluted EPS", r"\bnon[- ]gaap\s+(?:diluted\s+)?(?:eps|earnings\s+per\s+share|net\s+income\s+per\s+diluted\s+share)\b"),
    ("Non-GAAP gross margin %", r"\bnon[- ]gaap\s+gross\s+margin(?:\s+(?:percentage|percent))?\s*%"),
    ("Non-GAAP gross margin", r"\bnon[- ]gaap\s+gross\s+margin(?:\s+(?:percentage|percent))?\b(?!\s*%)"),
    ("Free cash flow margin", r"\bfree\s+cash\s+flow\s+margin\b"),
    ("Free cash flow", r"\bfree\s+cash\s+flow\b"),
    ("Organic revenue growth", r"\borganic\s+(?:revenue|sales)\s+(?:growth|change)\b"),
    ("Organic revenue", r"\borganic\s+(?:revenue|sales)\b"),
    ("Constant-currency growth", r"\bconstant[- ]currency\s+(?:revenue|sales|growth|change)\b"),
    ("Adjusted revenue", r"\badjusted\s+(?:net\s+)?revenue\b"),
    ("Core earnings", r"\bcore\s+(?:net\s+)?earnings\b"),
    ("Core EPS", r"\bcore\s+(?:diluted\s+)?eps\b"),
    ("Adjusted funds from operations", r"\badjusted\s+funds\s+from\s+operations\b|\baffo\b"),
    ("Funds from operations", r"\bfunds\s+from\s+operations\b|\bffo\b"),
    ("Normalized FFO", r"\bnormalized\s+ffo\b"),
    ("EBITDAre", r"\bebitdare\b"),
    ("EBITDAX", r"\bebitdax\b"),
    ("Distributable earnings", r"\bdistributable\s+earnings\b"),
    ("Cash available for distribution", r"\bcash\s+available\s+for\s+distribution\b"),
    ("Same-store NOI", r"\bsame[- ]store\s+(?:net\s+operating\s+income|noi)\b"),
    ("Tangible book value", r"\btangible\s+(?:common\s+)?book\s+value\b"),
    ("Tangible common equity ratio", r"\btangible\s+common\s+equity\s+(?:ratio|to\s+tangible\s+assets)\b|\btce\s+ratio\b"),
    ("Pre-provision net revenue", r"\bpre[- ]provision\s+net\s+revenue\b|\bppnr\b"),
    ("Net interest margin FTE", r"\bnet\s+interest\s+margin\s+(?:on\s+a\s+)?(?:fully\s+taxable\s+equivalent|fte)\b"),
    ("Net debt", r"\bnet\s+debt\b"),
    ("Net leverage ratio", r"\bnet\s+leverage\s+ratio\b"),
    ("Adjusted ROE", r"\badjusted\s+(?:return\s+on\s+(?:common\s+)?equity|roe)\b"),
    ("Adjusted ROA", r"\badjusted\s+(?:return\s+on\s+assets|roa)\b"),
]

ADJUSTMENT_TERMS = re.compile(
    r"stock[- ]based compensation|share[- ]based compensation|amortization|depreciation|restructur|"
    r"impairment|acquisition|transaction costs?|tax effect|legal expenses?|litigation|severance|"
    r"capital expenditures?|gain on|loss on|integration|transformation|one[- ]time|non[- ]cash",
    re.I,
)

# Ordered from the most specific labels to broad catch-alls.  The issuer's exact
# wording is always retained; this taxonomy is only a comparison aid.
ADJUSTMENT_CATEGORY_PATTERNS: list[tuple[str, str]] = [
    (
        "Stock-based and equity compensation",
        r"stock[- ]based|share[- ]based|equity[- ]based|incentive compensation.{0,40}(?:equity|shares?)|"
        r"founder shares?|restricted stock|option expense",
    ),
    (
        "Amortization of acquired intangibles",
        r"amortization.{0,45}(?:acquired|acquisition[- ]related|purchase accounting|intangible)|"
        r"(?:acquired|acquisition[- ]related).{0,45}amortization",
    ),
    (
        "Acquisition and transaction costs",
        r"acquisition[- ]related|business combination|transaction costs?|deal costs?|due diligence|"
        r"contingent consideration|purchase accounting",
    ),
    (
        "Separation and spin-off costs",
        r"separation[- ]related|spin[- ]?off|stand[- ]alone company|disentanglement",
    ),
    (
        "Restructuring and severance",
        r"restructur|severance|termination benefits?|workforce reduction|facility closure|exit costs?",
    ),
    (
        "Integration and transformation costs",
        r"integration|transformation|business optimization|strategic initiatives?|systems? conversion",
    ),
    (
        "Impairments and write-downs",
        r"impairment|write[- ]?down|write[- ]?off|abandonment",
    ),
    (
        "Litigation, legal, and regulatory",
        r"litigation|legal (?:expense|cost|settlement|reserve)|regulatory|investigation|compliance matter",
    ),
    (
        "Gains/losses on assets, investments, or divestitures",
        r"(?:gain|loss).{0,55}(?:sale|disposition|divestiture|asset|investment|equity method)|"
        r"(?:sale|disposition|divestiture).{0,55}(?:gain|loss)",
    ),
    (
        "Fair-value and mark-to-market changes",
        r"fair value|mark[- ]to[- ]market|remeasurement|valuation adjustment",
    ),
    (
        "Debt, refinancing, and extinguishment",
        r"debt extinguishment|loss on extinguishment|refinancing|financing costs?|debt issuance|"
        r"convertible note|early repayment",
    ),
    (
        "Income-tax effects and discrete tax items",
        r"tax effect|income tax|tax adjustment|tax benefit|tax expense|effective tax rate|discrete tax",
    ),
    (
        "Foreign-exchange effects",
        r"foreign exchange|foreign currency|currency translation|fx impact|constant currency",
    ),
    (
        "Pension and postretirement items",
        r"pension|postretirement|actuarial|settlement accounting",
    ),
    (
        "Depreciation and amortization",
        r"depreciation|amortization",
    ),
    (
        "Capital expenditures",
        r"capital expenditures?|purchases? of (?:property|plant|equipment)|capex",
    ),
    (
        "Insurance and settlement items",
        r"insurance (?:recovery|proceeds|settlement)|settlement (?:gain|loss|cost)",
    ),
    (
        "Non-cash and other accounting items",
        r"non[- ]cash|accounting change|adoption of|inventory step[- ]?up|lifo|fifo",
    ),
]

ADJUSTMENT_CATEGORY_ORDER = {
    category: index for index, (category, _pattern) in enumerate(ADJUSTMENT_CATEGORY_PATTERNS, start=1)
}
ADJUSTMENT_CATEGORY_ORDER["Other issuer-specific adjustment"] = len(ADJUSTMENT_CATEGORY_ORDER) + 1

NON_GAAP_START = re.compile(
    r"^(?:non[- ]gaap|adjusted|normalized|core|free cash flow|funds from operations|affo|ffo|"
    r"ebitda|ebit|ebitdare|ebitdax|distributable earnings|same[- ]store)",
    re.I,
)

HEADER_LABEL = re.compile(
    rf"(?:months? ended|years? ended|quarter ended|unaudited|in thousands|in millions|in billions|"
    rf"^(?:{MONTH_PATTERN})\s+\d{{1,2}}|^q[1-4]\s+\d{{4}})",
    re.I,
)

VALUE_PATTERN = re.compile(
    r"(?P<currency>\$)?\s*(?P<open>\()?\s*(?P<sign>-)?\s*"
    r"(?P<number>\d[\d,]*(?:\.\d+)?)\s*(?P<close>\))?\s*(?P<suffix>%|bps)?",
    re.I,
)


def decode_document(content: bytes) -> str:
    for encoding in ("utf-8", "windows-1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def source_extension(url: str) -> str:
    return PurePosixPath(urlparse(url).path).suffix.lower()


def is_pdf_resource(resource: FetchedResource, url: str = "") -> bool:
    return (
        resource.content.startswith(b"%PDF")
        or resource.content_type == "application/pdf"
        or source_extension(url or resource.url) == ".pdf"
    )


def table_context(table: Tag, max_items: int = 5) -> str:
    contexts: list[str] = []
    node: Any = table
    for _ in range(max_items * 3):
        node = node.find_previous(["h1", "h2", "h3", "h4", "h5", "p", "div", "strong", "b"])
        if not node:
            break
        if node.find_parent("table") is not None:
            continue
        text = clean_space(node.get_text(" ", strip=True))
        if 3 <= len(text) <= 240 and text not in contexts:
            contexts.append(text)
        if len(contexts) >= max_items:
            break
    return " | ".join(reversed(contexts))


def html_table_rows(table: Tag) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.find_all("tr"):
        cells = row.find_all(["th", "td"], recursive=False)
        if not cells:
            cells = row.find_all(["th", "td"])
        values = [clean_space(cell.get_text(" ", strip=True)) for cell in cells]
        if any(values):
            rows.append(values)
    return rows


def table_reconciliation_score(context: str, rows: list[list[str]]) -> int:
    text = clean_space(context + " " + " ".join(" ".join(row) for row in rows))
    lower = text.lower()
    score = 0
    if "reconciliation" in lower:
        score += 8
    gaap_rows = len(re.findall(r"(?:^|\s)gaap\s+[a-z]", lower))
    non_gaap_rows = len(re.findall(r"non[- ]gaap\s+[a-z]", lower))
    if gaap_rows and non_gaap_rows:
        score += 4
    if ADJUSTMENT_TERMS.search(text):
        score += 3
    if re.search(r"most directly comparable|u\.s\. gaap to non[- ]gaap", lower):
        score += 2
    return score


def extract_scale(text: str) -> str:
    lower = text.lower()
    if re.search(r"\bin billions\b|\$\s*in\s*billions", lower):
        return "billions"
    if re.search(r"\bin millions\b|\$\s*in\s*millions", lower):
        return "millions"
    if re.search(r"\bin thousands\b|\$\s*in\s*thousands", lower):
        return "thousands"
    return "units"


def _strip_leading_footnotes(tokens: list[str]) -> list[str]:
    """Remove unmistakable standalone footnote markers without discarding real values.

    A plain token such as ``62`` may be the current-period value in an SEC table, so
    only parenthesized/bracketed numbers or asterisks are treated as footnotes.
    """
    result = list(tokens)
    while result and re.fullmatch(r"(?:\(\d{1,2}\)|\[\d{1,2}\]|\*+)", clean_space(result[0])):
        result.pop(0)
    return result


def parse_first_value(tokens: Iterable[str]) -> Optional[dict[str, Any]]:
    clean_tokens = [clean_space(token) for token in tokens if clean_space(token)]
    clean_tokens = _strip_leading_footnotes(clean_tokens)
    if not clean_tokens:
        return None

    joined = " ".join(clean_tokens)
    if re.fullmatch(r"[-—–\s]+", joined):
        return None

    for match in VALUE_PATTERN.finditer(joined):
        raw = clean_space(match.group(0))
        number_text = match.group("number")
        if not number_text:
            continue
        # Parenthesized/bracketed footnotes were removed above. A plain small
        # number is kept because it can be a legitimate current-period value.
        try:
            value = float(number_text.replace(",", ""))
        except ValueError:
            continue
        if match.group("sign") == "-" or (match.group("open") and match.group("close")):
            value = -value
        return {
            "value": value,
            "raw": raw,
            "currency": bool(match.group("currency")) or "$" in joined[: match.end() + 1],
            "suffix": (match.group("suffix") or "").lower(),
        }
    return None


def row_label_and_value(cells: list[str]) -> tuple[str, Optional[dict[str, Any]]]:
    nonempty = [(index, clean_space(value)) for index, value in enumerate(cells) if clean_space(value)]
    if not nonempty:
        return "", None
    label_pos: Optional[int] = None
    for index, value in nonempty:
        if re.search(r"[A-Za-z]", value) and value not in {"$", "%"}:
            label_pos = index
            break
    if label_pos is None:
        return "", None
    label = clean_space(cells[label_pos])
    if HEADER_LABEL.search(label) and not re.match(r"^(?:GAAP|Non[- ]GAAP|Adjusted)", label, re.I):
        return label, None
    value = parse_first_value(cells[label_pos + 1 :])
    return label, value


def row_kind(label: str) -> str:
    """Classify a reconciliation row without assuming issuers use a GAAP prefix.

    Many issuers label the comparable GAAP row as ``as reported`` rather than
    ``GAAP``.  Subtotal rows such as ``Total special items`` are identified
    separately so they are not double counted with the detailed adjustments.
    """
    normalized = clean_space(label).replace("–", "-").replace("—", "-")
    lower = normalized.lower()
    if re.match(r"^(?:total\s+)?(?:special items|non[- ]gaap adjustments?|adjustments?)\b", lower):
        return "subtotal"
    if re.match(r"^gaap\b", lower) or re.search(r"\b(?:as reported|reported results?)\b", lower):
        return "gaap"
    if NON_GAAP_START.match(lower) or re.search(r"\bnon[- ]gaap\b", lower):
        # "Adjusted for..." can be an adjustment row, not a final metric.
        if re.match(r"^adjusted\s+for\b", lower):
            return "adjustment"
        return "non_gaap"
    return "adjustment"


def base_metric_key(label: str) -> str:
    value = clean_space(label).lower().replace("–", "-")
    value = re.sub(r"\([^)]*\)$", "", value)
    value = re.sub(r"\b(?:u\.s\.)?\s*gaap\b", "", value)
    value = re.sub(r"\bnon[- ]gaap\b", "", value)
    value = re.sub(r"\badjusted\b|\bnormalized\b|\bcore\b", "", value)
    value = re.sub(r"\breconciliation\b", "", value)
    value = re.sub(r"\bpercentage\b", "%", value)
    value = re.sub(r"\bearnings per share\b", "eps", value)
    value = re.sub(r"\bnet income per share\b", "eps", value)
    value = re.sub(r"\bincome from operations\b", "operating income", value)
    value = re.sub(r"[^a-z0-9%]+", " ", value)
    return clean_space(value)


def canonical_metric_name(label: str, section: str = "") -> str:
    text = clean_space(label)
    for canonical, pattern in METRIC_PATTERNS:
        if re.search(pattern, text, re.I):
            return canonical
    cleaned = re.sub(r"^(?:non[- ]gaap|adjusted|normalized|core)\s+", "", text, flags=re.I)
    cleaned = re.sub(r"\s+reconciliation$", "", cleaned, flags=re.I)
    if cleaned and cleaned.lower() not in {"total", "subtotal"}:
        prefix = "Non-GAAP " if re.match(r"^non[- ]gaap", text, re.I) else "Adjusted " if re.match(r"^adjusted", text, re.I) else ""
        return clean_space(prefix + cleaned)
    section_clean = re.sub(r"\s+reconciliation$", "", clean_space(section), flags=re.I)
    return section_clean or text or "Non-GAAP metric"


def infer_unit(label: str, value_meta: dict[str, Any], scale: str) -> str:
    """Infer the displayed unit without mistaking dollar gross margin for a rate.

    Currency markers and explicit percent markers take precedence over words such
    as ``margin`` because issuers often label both a dollar amount and a percentage
    as "gross margin" in the same reconciliation table.
    """
    lower = label.lower()
    suffix = value_meta.get("suffix") or ""
    if suffix == "%" or "%" in label or "percentage" in lower or " percent" in lower:
        return "percent"
    if suffix == "bps":
        return "bps"
    if re.search(r"per share|\beps\b", lower):
        return "usd_per_share"
    if value_meta.get("currency"):
        return "usd"
    if "margin" in lower or re.search(r"\brate\b|\bratio\b", lower):
        return "percent"
    if scale in {"thousands", "millions", "billions"}:
        return "usd"
    return "number"


def scale_multiplier(scale: str, unit: str) -> float:
    if unit != "usd":
        return 1.0
    return {"thousands": 1_000.0, "millions": 1_000_000.0, "billions": 1_000_000_000.0}.get(scale, 1.0)


def normalize_value(value: Optional[float], scale: str, unit: str) -> Optional[float]:
    if value is None:
        return None
    return float(value) * scale_multiplier(scale, unit)


def format_value(value: Optional[float], unit: str, scale: str = "units") -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "—"
    if unit == "percent":
        return f"{value:,.1f}%"
    if unit == "bps":
        return f"{value:,.0f} bps"
    if unit == "usd_per_share":
        sign = "-" if value < 0 else ""
        return f"{sign}${abs(value):,.2f}/share"
    if unit == "usd":
        normalized = normalize_value(value, scale, unit) or 0.0
        sign = "-" if normalized < 0 else ""
        absolute = abs(normalized)
        if absolute >= 1_000_000_000:
            return f"{sign}${absolute / 1_000_000_000:,.2f}B"
        if absolute >= 1_000_000:
            return f"{sign}${absolute / 1_000_000:,.1f}M"
        if absolute >= 1_000:
            return f"{sign}${absolute / 1_000:,.1f}K"
        return f"{sign}${absolute:,.2f}"
    if abs(value - round(value)) < 1e-9:
        return f"{value:,.0f}"
    return f"{value:,.2f}"



def normalize_adjustment_label(label: str) -> str:
    """Normalize superficial footnote markers while preserving issuer wording."""
    value = clean_space(label).replace("–", "-").replace("—", "-")
    previous = None
    while previous != value:
        previous = value
        value = re.sub(r"\s*(?:\(\s*[a-z0-9]{1,3}\s*\)|\[\s*[a-z0-9]{1,3}\s*\])\s*$", "", value, flags=re.I)
    return clean_space(value)


def classify_adjustment_label(label: str) -> str:
    """Map an issuer label to a comparison category without replacing the raw label."""
    value = normalize_adjustment_label(label)
    for category, pattern in ADJUSTMENT_CATEGORY_PATTERNS:
        if re.search(pattern, value, re.I):
            return category
    return "Other issuer-specific adjustment"


def fiscal_period_rank(fiscal_year: Any, fiscal_quarter: Any) -> int:
    try:
        year = int(float(fiscal_year))
    except Exception:
        year = 0
    quarter = QUARTER_ORDER.get(clean_space(fiscal_quarter).upper(), 0)
    return year * 10 + quarter


def ordered_fiscal_periods(frame: pd.DataFrame) -> list[str]:
    if frame is None or frame.empty or "period" not in frame.columns:
        return []
    data = frame.copy()
    if {"fiscal_year", "fiscal_quarter"}.issubset(data.columns):
        data["_period_rank"] = data.apply(
            lambda row: fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
        )
    else:
        def _parse_period(value: Any) -> int:
            match = re.search(r"FY(\d{4})\s+Q([1-4])", clean_space(value), re.I)
            return int(match.group(1)) * 10 + int(match.group(2)) if match else 0
        data["_period_rank"] = data["period"].map(_parse_period)
    return (
        data[["period", "_period_rank"]]
        .dropna(subset=["period"])
        .drop_duplicates()
        .sort_values(["_period_rank", "period"])["period"]
        .astype(str)
        .tolist()
    )


def _normalized_adjustment_amount(row: pd.Series) -> Optional[float]:
    try:
        value = float(row.get("adjustment_value"))
    except Exception:
        return None
    unit = clean_space(row.get("unit"))
    scale = clean_space(row.get("scale"))
    return normalize_value(value, scale, unit) if unit == "usd" else value


def _format_normalized_amount(value: Optional[float], unit: str) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "—"
    return format_value(float(value), unit, "units")


def enrich_adjustments(
    adjustments: pd.DataFrame,
    reconciliations: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Add normalized categories, period history, and recurrence indicators.

    The exact issuer label remains in ``adjustment_label``.  Categories and
    recurrence fields are analytical aids, not conclusions about whether an
    adjustment is permissible or truly non-recurring.
    """
    data = adjustments.copy() if isinstance(adjustments, pd.DataFrame) else pd.DataFrame()
    added_columns = {
        "normalized_adjustment_label": "object",
        "adjustment_category": "object",
        "adjustment_category_order": "int64",
        "normalized_adjustment_value": "float64",
        "absolute_adjustment_value": "float64",
        "effect_on_non_gaap": "object",
        "period_rank": "int64",
        "observed_period_count": "int64",
        "available_period_count": "int64",
        "observed_frequency": "object",
        "recurrence_indicator": "object",
        "period_lifecycle": "object",
        "first_observed_period": "object",
        "last_observed_period": "object",
    }
    if data.empty:
        for column, dtype in added_columns.items():
            if column not in data.columns:
                data[column] = pd.Series(dtype=dtype)
        return data

    for column in ["adjustment_label", "metric", "period", "fiscal_quarter", "unit", "scale"]:
        if column not in data.columns:
            data[column] = ""
    if "fiscal_year" not in data.columns:
        data["fiscal_year"] = 0

    data["normalized_adjustment_label"] = data["adjustment_label"].map(normalize_adjustment_label)
    data["adjustment_category"] = data["adjustment_label"].map(classify_adjustment_label)
    data["adjustment_category_order"] = data["adjustment_category"].map(ADJUSTMENT_CATEGORY_ORDER).fillna(999).astype(int)
    data["normalized_adjustment_value"] = data.apply(_normalized_adjustment_amount, axis=1)
    data["absolute_adjustment_value"] = data["normalized_adjustment_value"].abs()
    data["effect_on_non_gaap"] = data["normalized_adjustment_value"].map(
        lambda value: "No numeric effect parsed"
        if value is None or pd.isna(value) or float(value) == 0
        else "Increases the reported non-GAAP measure"
        if float(value) > 0
        else "Decreases the reported non-GAAP measure"
    )
    data["period_rank"] = data.apply(
        lambda row: fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
    )

    period_source = reconciliations if isinstance(reconciliations, pd.DataFrame) and not reconciliations.empty else data
    metric_periods: dict[str, list[int]] = defaultdict(list)
    metric_period_labels: dict[tuple[str, int], str] = {}
    if isinstance(period_source, pd.DataFrame) and not period_source.empty:
        source = period_source.copy()
        for column in ["metric", "period", "fiscal_year", "fiscal_quarter"]:
            if column not in source.columns:
                source[column] = ""
        source["_period_rank"] = source.apply(
            lambda row: fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
        )
        for metric, group in source.groupby("metric", dropna=False):
            metric_name = clean_space(metric)
            ranks = sorted({int(value) for value in group["_period_rank"].dropna().tolist() if int(value) > 0})
            metric_periods[metric_name] = ranks
            for _, row in group.iterrows():
                rank = int(row.get("_period_rank") or 0)
                if rank > 0:
                    metric_period_labels[(metric_name, rank)] = clean_space(row.get("period"))

    for _, row in data.iterrows():
        metric_name = clean_space(row.get("metric"))
        rank = int(row.get("period_rank") or 0)
        if rank > 0:
            metric_period_labels[(metric_name, rank)] = clean_space(row.get("period"))
            if rank not in metric_periods[metric_name]:
                metric_periods[metric_name].append(rank)
                metric_periods[metric_name] = sorted(set(metric_periods[metric_name]))

    presence: dict[tuple[str, str], list[int]] = {}
    for (metric, category), group in data.groupby(["metric", "adjustment_category"], dropna=False):
        key = (clean_space(metric), clean_space(category))
        presence[key] = sorted({int(value) for value in group["period_rank"].dropna().tolist() if int(value) > 0})

    observed_counts: list[int] = []
    available_counts: list[int] = []
    frequencies: list[str] = []
    recurrence: list[str] = []
    lifecycle: list[str] = []
    first_periods: list[str] = []
    last_periods: list[str] = []

    for _, row in data.iterrows():
        metric_name = clean_space(row.get("metric"))
        category = clean_space(row.get("adjustment_category"))
        rank = int(row.get("period_rank") or 0)
        observed = presence.get((metric_name, category), [])
        available = metric_periods.get(metric_name, []) or observed
        observed_count = len(observed)
        available_count = len(available)
        observed_counts.append(observed_count)
        available_counts.append(available_count)
        frequencies.append(f"{observed_count} of {available_count} analyzed periods" if available_count else "")
        recurrence.append("Repeated in selected history" if observed_count >= 2 else "Single-period in selected history")

        if not observed or rank <= 0:
            lifecycle.append("")
            first_periods.append("")
            last_periods.append("")
            continue
        first_periods.append(metric_period_labels.get((metric_name, observed[0]), str(observed[0])))
        last_periods.append(metric_period_labels.get((metric_name, observed[-1]), str(observed[-1])))
        if rank == observed[0]:
            lifecycle.append("First observed in selected history")
        else:
            try:
                position = available.index(rank)
            except ValueError:
                position = -1
            previous_available = available[position - 1] if position > 0 else None
            lifecycle.append(
                "Continued from prior analyzed period"
                if previous_available in observed
                else "Returned after a gap"
            )

    data["observed_period_count"] = observed_counts
    data["available_period_count"] = available_counts
    data["observed_frequency"] = frequencies
    data["recurrence_indicator"] = recurrence
    data["period_lifecycle"] = lifecycle
    data["first_observed_period"] = first_periods
    data["last_observed_period"] = last_periods

    sort_columns = [column for column in ["fiscal_year", "period_rank", "metric", "adjustment_category_order", "adjustment_label"] if column in data.columns]
    return data.sort_values(sort_columns).reset_index(drop=True)


def _tie_out_tolerance(expected: float, unit: str) -> float:
    if unit == "percent":
        return 0.15
    if unit == "bps":
        return 1.0
    if unit == "usd_per_share":
        return 0.015
    if unit == "usd":
        return max(1.0, abs(expected) * 0.0025)
    return max(0.01, abs(expected) * 0.005)


def build_adjustment_tieouts(
    reconciliations: pd.DataFrame,
    adjustments: pd.DataFrame,
) -> pd.DataFrame:
    """Compare parsed adjustment lines with non-GAAP minus GAAP for each bridge."""
    columns = [
        "pair_id", "fiscal_year", "fiscal_quarter", "period", "metric", "gaap_display",
        "expected_adjustment_display", "parsed_adjustment_display", "variance_display",
        "non_gaap_display", "detail_line_count", "category_count", "tie_out_status",
        "tie_out_note", "source_role", "source_page", "source_url",
    ]
    if reconciliations is None or reconciliations.empty:
        return pd.DataFrame(columns=columns)

    detail = adjustments.copy() if isinstance(adjustments, pd.DataFrame) else pd.DataFrame()
    rows: list[dict[str, Any]] = []
    for _, pair in reconciliations.iterrows():
        pair_id = clean_space(pair.get("pair_id"))
        pair_detail = detail[detail.get("pair_id", pd.Series(index=detail.index, dtype=str)).astype(str).eq(pair_id)].copy() if not detail.empty else pd.DataFrame()
        if not pair_detail.empty:
            pair_detail = pair_detail.drop_duplicates(
                [column for column in ["pair_id", "adjustment_label", "adjustment_value", "unit", "scale", "source_url", "source_page"] if column in pair_detail.columns],
                keep="first",
            )
        numeric = pd.to_numeric(pair_detail.get("adjustment_value", pd.Series(dtype=float)), errors="coerce").dropna()
        parsed_sum = float(numeric.sum()) if not numeric.empty else None
        try:
            expected = float(pair.get("adjustment_value"))
        except Exception:
            expected = None
        unit = clean_space(pair.get("unit"))
        scale = clean_space(pair.get("scale")) or "units"
        variance = parsed_sum - expected if parsed_sum is not None and expected is not None else None
        if pair_detail.empty:
            status = "No line-item detail"
            note = "The GAAP and non-GAAP endpoints were parsed, but individual reconciling rows were not."
        elif expected is None:
            status = "No total available"
            note = "Individual rows were parsed, but the endpoint difference was unavailable."
        elif variance is not None and abs(variance) <= _tie_out_tolerance(expected, unit):
            status = "Ties within rounding"
            note = "The parsed line items agree with non-GAAP minus GAAP within the rounding threshold."
        else:
            status = "Review difference"
            note = "Parsed line items do not fully agree with the endpoint difference; review for subtotals, omitted rows, or table parsing issues."
        categories = (
            pair_detail["adjustment_category"].nunique()
            if not pair_detail.empty and "adjustment_category" in pair_detail.columns
            else pair_detail.get("adjustment_label", pd.Series(dtype=str)).map(classify_adjustment_label).nunique()
            if not pair_detail.empty
            else 0
        )
        rows.append(
            {
                "pair_id": pair_id,
                "fiscal_year": pair.get("fiscal_year"),
                "fiscal_quarter": pair.get("fiscal_quarter"),
                "period": pair.get("period"),
                "metric": pair.get("metric"),
                "gaap_display": pair.get("gaap_display"),
                "expected_adjustment_display": pair.get("adjustment_display"),
                "parsed_adjustment_display": format_value(parsed_sum, unit, scale) if parsed_sum is not None else "—",
                "variance_display": format_value(variance, unit, scale) if variance is not None else "—",
                "non_gaap_display": pair.get("non_gaap_display"),
                "detail_line_count": int(len(pair_detail)),
                "category_count": int(categories),
                "tie_out_status": status,
                "tie_out_note": note,
                "source_role": pair.get("source_role"),
                "source_page": pair.get("source_page"),
                "source_url": pair.get("source_url"),
            }
        )
    result = pd.DataFrame(rows, columns=columns)
    if not result.empty:
        result["_rank"] = result.apply(
            lambda row: fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
        )
        result = result.sort_values(["_rank", "metric"]).drop(columns=["_rank"]).reset_index(drop=True)
    return result


def _aggregate_adjustment_values(group: pd.DataFrame) -> tuple[str, dict[str, float]]:
    displays: list[str] = []
    numeric_by_unit: dict[str, float] = {}
    for unit, unit_group in group.groupby("unit", dropna=False):
        unit_name = clean_space(unit) or "number"
        if unit_name == "usd":
            values = pd.to_numeric(unit_group.get("normalized_adjustment_value"), errors="coerce").dropna()
        else:
            values = pd.to_numeric(unit_group.get("adjustment_value"), errors="coerce").dropna()
        if values.empty:
            continue
        total = float(values.sum())
        numeric_by_unit[unit_name] = total
        displays.append(_format_normalized_amount(total, unit_name))
    return " / ".join(displays) if displays else "—", numeric_by_unit


def make_adjustment_value_matrix(adjustment_history: pd.DataFrame, metric: str) -> pd.DataFrame:
    """Create an adjustment-category-by-period matrix for one non-GAAP metric."""
    if adjustment_history is None or adjustment_history.empty or not clean_space(metric):
        return pd.DataFrame()
    data = adjustment_history[adjustment_history["metric"].astype(str).eq(str(metric))].copy()
    if data.empty:
        return pd.DataFrame()
    records: list[dict[str, Any]] = []
    for (category, period), group in data.groupby(["adjustment_category", "period"], dropna=False):
        display, _numeric = _aggregate_adjustment_values(group)
        records.append(
            {
                "adjustment_category": category,
                "period": period,
                "matrix_value": display,
                "category_order": int(group["adjustment_category_order"].min()) if "adjustment_category_order" in group else 999,
            }
        )
    values = pd.DataFrame(records)
    matrix = values.pivot_table(index="adjustment_category", columns="period", values="matrix_value", aggfunc="first")
    ordered_periods = ordered_fiscal_periods(data)
    category_order = (
        values[["adjustment_category", "category_order"]]
        .drop_duplicates()
        .sort_values(["category_order", "adjustment_category"])["adjustment_category"]
        .tolist()
    )
    matrix = matrix.reindex(index=category_order, columns=ordered_periods).reset_index()
    return matrix.rename(columns={"adjustment_category": "Adjustment category"})


def make_adjustment_presence_matrix(adjustment_history: pd.DataFrame) -> pd.DataFrame:
    """Show where each category appears without summing the same item across metrics."""
    if adjustment_history is None or adjustment_history.empty:
        return pd.DataFrame()
    data = adjustment_history.copy()
    records: list[dict[str, Any]] = []
    for (category, period), group in data.groupby(["adjustment_category", "period"], dropna=False):
        metric_count = int(group["metric"].nunique())
        line_count = int(len(group))
        metric_word = "metric" if metric_count == 1 else "metrics"
        line_word = "line" if line_count == 1 else "lines"
        records.append(
            {
                "adjustment_category": category,
                "period": period,
                "matrix_value": f"{metric_count} {metric_word} · {line_count} {line_word}",
                "category_order": int(group["adjustment_category_order"].min()) if "adjustment_category_order" in group else 999,
            }
        )
    values = pd.DataFrame(records)
    matrix = values.pivot_table(index="adjustment_category", columns="period", values="matrix_value", aggfunc="first")
    ordered_periods = ordered_fiscal_periods(data)
    category_order = (
        values[["adjustment_category", "category_order"]]
        .drop_duplicates()
        .sort_values(["category_order", "adjustment_category"])["adjustment_category"]
        .tolist()
    )
    matrix = matrix.reindex(index=category_order, columns=ordered_periods).reset_index()
    return matrix.rename(columns={"adjustment_category": "Adjustment category"})


def make_adjustment_metric_matrix(adjustment_history: pd.DataFrame) -> pd.DataFrame:
    """Create an export-friendly matrix keyed by metric and adjustment category."""
    if adjustment_history is None or adjustment_history.empty:
        return pd.DataFrame()
    frames: list[pd.DataFrame] = []
    for metric in sorted(adjustment_history["metric"].dropna().astype(str).unique().tolist()):
        matrix = make_adjustment_value_matrix(adjustment_history, metric)
        if matrix.empty:
            continue
        matrix.insert(0, "Non-GAAP metric", metric)
        frames.append(matrix)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def adjustment_category_summary(adjustment_history: pd.DataFrame) -> pd.DataFrame:
    if adjustment_history is None or adjustment_history.empty:
        return pd.DataFrame()
    records: list[dict[str, Any]] = []
    for category, group in adjustment_history.groupby("adjustment_category", dropna=False):
        periods = ordered_fiscal_periods(group)
        labels = sorted({clean_space(value) for value in group["adjustment_label"].tolist() if clean_space(value)})
        records.append(
            {
                "adjustment_category": category,
                "periods_observed": len(periods),
                "first_period": periods[0] if periods else "",
                "latest_period": periods[-1] if periods else "",
                "metrics_affected": int(group["metric"].nunique()),
                "issuer_labels": "; ".join(labels),
                "recurrence_indicator": "Repeated in selected history" if len(periods) >= 2 else "Single-period in selected history",
                "category_order": int(group["adjustment_category_order"].min()) if "adjustment_category_order" in group else 999,
            }
        )
    result = pd.DataFrame(records).sort_values(["category_order", "adjustment_category"]).drop(columns=["category_order"])
    return result.reset_index(drop=True)


def compare_adjustment_periods(
    adjustment_history: pd.DataFrame,
    metric: str,
    current_period: str,
    prior_period: str,
) -> pd.DataFrame:
    """Compare adjustment categories for one metric across two selected fiscal periods."""
    columns = [
        "status", "adjustment_category", "prior_value", "current_value", "change",
        "prior_issuer_labels", "current_issuer_labels", "observed_periods", "source_url",
    ]
    if adjustment_history is None or adjustment_history.empty or not metric or not current_period or not prior_period:
        return pd.DataFrame(columns=columns)
    data = adjustment_history[adjustment_history["metric"].astype(str).eq(str(metric))].copy()
    current = data[data["period"].astype(str).eq(str(current_period))]
    prior = data[data["period"].astype(str).eq(str(prior_period))]
    categories = sorted(set(current["adjustment_category"].tolist()) | set(prior["adjustment_category"].tolist()), key=lambda value: (ADJUSTMENT_CATEGORY_ORDER.get(value, 999), value))
    rows: list[dict[str, Any]] = []
    for category in categories:
        current_group = current[current["adjustment_category"].eq(category)]
        prior_group = prior[prior["adjustment_category"].eq(category)]
        if not current_group.empty and prior_group.empty:
            status = "New in current period"
        elif not current_group.empty and not prior_group.empty:
            status = "Continued"
        else:
            status = "No longer reported"
        current_display, current_numeric = _aggregate_adjustment_values(current_group) if not current_group.empty else ("—", {})
        prior_display, prior_numeric = _aggregate_adjustment_values(prior_group) if not prior_group.empty else ("—", {})
        change_display = ""
        common_units = set(current_numeric) & set(prior_numeric)
        if len(common_units) == 1:
            unit = next(iter(common_units))
            change_display = _format_normalized_amount(current_numeric[unit] - prior_numeric[unit], unit)
        rows.append(
            {
                "status": status,
                "adjustment_category": category,
                "prior_value": prior_display,
                "current_value": current_display,
                "change": change_display,
                "prior_issuer_labels": "; ".join(sorted(set(prior_group.get("adjustment_label", pd.Series(dtype=str)).astype(str).tolist()))),
                "current_issuer_labels": "; ".join(sorted(set(current_group.get("adjustment_label", pd.Series(dtype=str)).astype(str).tolist()))),
                "observed_periods": int(data[data["adjustment_category"].eq(category)]["period"].nunique()),
                "source_url": clean_space(current_group.iloc[0].get("source_url")) if not current_group.empty else clean_space(prior_group.iloc[0].get("source_url")) if not prior_group.empty else "",
            }
        )
    order = {"New in current period": 0, "Continued": 1, "No longer reported": 2}
    result = pd.DataFrame(rows, columns=columns)
    if not result.empty:
        result["_status_order"] = result["status"].map(order).fillna(9)
        result["_category_order"] = result["adjustment_category"].map(ADJUSTMENT_CATEGORY_ORDER).fillna(999)
        result = result.sort_values(["_status_order", "_category_order", "adjustment_category"]).drop(columns=["_status_order", "_category_order"])
    return result.reset_index(drop=True)


def _find_gaap_match(rows: list[dict[str, Any]], non_gaap_index: int) -> Optional[int]:
    """Find the comparable GAAP/base row for a non-GAAP result.

    The first pass uses explicit ``GAAP`` or ``as reported`` labels.  A second
    pass handles common reconciliations whose baseline is plainly comparable
    but not prefixed with GAAP, such as ``Net cash provided by operating
    activities`` in a free-cash-flow bridge.
    """
    target = rows[non_gaap_index]
    explicit = [
        index
        for index, row in enumerate(rows[:non_gaap_index])
        if row.get("kind") == "gaap"
    ]
    if not explicit:
        explicit = [index for index, row in enumerate(rows) if row.get("kind") == "gaap"]

    target_key = base_metric_key(target.get("label", ""))
    section_text = clean_space(target.get("section", "")).lower()

    candidates = explicit
    if not candidates:
        candidates = []
        for index, row in enumerate(rows[:non_gaap_index]):
            label = clean_space(row.get("label", ""))
            lower = label.lower()
            if not label or row.get("kind") == "subtotal":
                continue
            if re.search(r"^(?:special items|adjustments?|add:|less:)", lower):
                continue
            if ADJUSTMENT_TERMS.search(label):
                continue
            if re.search(
                r"\b(?:tax effect|interest expense|interest income|income tax|equity income|"
                r"capital expenditures?|purchases? of property|depreciation|amortization|"
                r"stock[- ]based|share[- ]based|restructur|impairment|acquisition)\b",
                lower,
            ):
                continue
            candidates.append(index)

    if not candidates:
        return None

    best_index: Optional[int] = None
    best_score = -999.0
    for index in candidates:
        label = clean_space(rows[index].get("label", ""))
        lower = label.lower()
        key = base_metric_key(label)
        similarity = SequenceMatcher(None, target_key, key).ratio() if target_key and key else 0.0
        distance_bonus = max(0.0, 0.25 - abs(non_gaap_index - index) * 0.012)
        score = similarity + distance_bonus
        if rows[index].get("kind") == "gaap":
            score += 2.0
        if target_key == key and target_key:
            score += 1.2
        if index == 0:
            score += 0.25

        target_lower = clean_space(target.get("label", "")).lower()
        if "free cash flow" in target_lower and re.search(r"cash (?:provided by|flow from) operating activities", lower):
            score += 3.0
        if "ebitda" in target_lower and re.search(r"net income|operating income", lower):
            score += 2.5
        if re.search(r"gross (?:profit|margin)", target_lower) and re.search(r"gross (?:profit|margin)", lower):
            score += 2.0
        if "operating expense" in target_lower and "operating expense" in lower:
            score += 2.0
        if re.search(r"net income|earnings", target_lower) and re.search(r"net income|earnings", lower):
            score += 1.8
        if re.search(r"per share|\beps\b", target_lower) and re.search(r"per share|\beps\b", lower):
            score += 1.8
        if label.lower() in section_text:
            score += 0.3

        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def build_pairs_from_rows(
    rows: list[dict[str, Any]],
    table_meta: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    pairs: list[dict[str, Any]] = []
    adjustments: list[dict[str, Any]] = []
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["section"]].append(row)

    for section, section_rows in grouped.items():
        for ng_index, ng_row in enumerate(section_rows):
            if ng_row["kind"] != "non_gaap":
                continue
            gaap_index = _find_gaap_match(section_rows, ng_index)
            if gaap_index is None:
                continue
            gaap_row = section_rows[gaap_index]
            # Each endpoint's unit is inferred from ITS OWN label/value metadata, not
            # borrowed from the other side. A reconciliation to a margin/rate metric
            # (e.g. "Free cash flow margin", "Adjusted EBITDA margin") has a dollar
            # GAAP anchor (net income, operating cash flow) but a percent non-GAAP
            # endpoint -- treating them as one unit mislabels the GAAP dollar value
            # as a percentage.
            gaap_unit = infer_unit(gaap_row["label"], gaap_row["value_meta"], table_meta["scale"])
            non_gaap_unit = infer_unit(ng_row["label"], ng_row["value_meta"], table_meta["scale"])
            gaap_value = gaap_row["value"]
            non_gaap_value = ng_row["value"]
            mixed_units = gaap_unit != non_gaap_unit
            adjustment_value = (
                non_gaap_value - gaap_value
                if gaap_value is not None and non_gaap_value is not None and not mixed_units
                else None
            )
            metric_name = canonical_metric_name(ng_row["label"], section)
            pair_id = f"{table_meta['source_url']}|{table_meta.get('source_page') or ''}|{section}|{ng_index}|{metric_name}"

            pair = {
                "pair_id": pair_id,
                "metric": metric_name,
                "section": section,
                "gaap_label": gaap_row["label"],
                "gaap_value": gaap_value,
                "gaap_display": format_value(gaap_value, gaap_unit, table_meta["scale"]),
                "non_gaap_label": ng_row["label"],
                "non_gaap_value": non_gaap_value,
                "non_gaap_display": format_value(non_gaap_value, non_gaap_unit, table_meta["scale"]),
                "adjustment_value": adjustment_value,
                "adjustment_display": (
                    "n/m" if mixed_units else format_value(adjustment_value, gaap_unit, table_meta["scale"])
                ),
                "unit": non_gaap_unit,
                "gaap_unit": gaap_unit,
                "scale": table_meta["scale"],
                "source_role": table_meta["source_role"],
                "source_document": table_meta["source_document"],
                "source_description": table_meta["source_description"],
                "source_url": table_meta["source_url"],
                "source_page": table_meta.get("source_page"),
                "table_title": table_meta["title"],
                "confidence": "High" if any(
                    row.get("kind") in {"adjustment", "subtotal"}
                    for row in section_rows[min(gaap_index, ng_index) + 1 : max(gaap_index, ng_index)]
                ) else "Medium",
            }
            pairs.append(pair)

            start, end = sorted((gaap_index, ng_index))
            bridge_rows = section_rows[start + 1 : end]
            detail_rows = [row for row in bridge_rows if row.get("kind") == "adjustment"]
            subtotal_rows = [row for row in bridge_rows if row.get("kind") == "subtotal"]
            rows_to_emit = detail_rows if detail_rows else subtotal_rows
            for adjustment_order, adjustment_row in enumerate(rows_to_emit, start=1):
                adjustment_meta = adjustment_row["value_meta"]
                adjustment_label_lower = clean_space(adjustment_row.get("label")).lower()
                # Reconciliation tables usually print the currency or percent sign only
                # on the GAAP/non-GAAP endpoint rows. Individual adjustments therefore
                # inherit the bridge unit unless their own cell or label explicitly
                # states another unit. This avoids treating dollar gross-margin
                # adjustments as percentages merely because the label contains "margin".
                if adjustment_meta.get("currency"):
                    adjustment_unit = "usd"
                elif adjustment_meta.get("suffix") == "%" or "%" in adjustment_label_lower or "percentage" in adjustment_label_lower:
                    adjustment_unit = "percent"
                elif adjustment_meta.get("suffix") == "bps" or re.search(r"\bbps\b|basis points?", adjustment_label_lower):
                    adjustment_unit = "bps"
                elif re.search(r"per share|\beps\b", adjustment_label_lower):
                    adjustment_unit = "usd_per_share"
                else:
                    # Individual bridge lines (SBC, D&A, restructuring, capex, etc.)
                    # are conventionally denominated in the GAAP anchor's unit even
                    # when the final non-GAAP endpoint is a margin/rate/per-share
                    # figure -- the ratio is typically a separate derived row with
                    # its own explicit "%" marker (handled above), not the bridge
                    # itself. Falling back to the GAAP unit (not the endpoint unit)
                    # avoids mislabeling dollar reconciling items as percentages.
                    adjustment_unit = gaap_unit
                adjustments.append(
                    {
                        "pair_id": pair_id,
                        "metric": metric_name,
                        "section": section,
                        "adjustment_order": adjustment_order,
                        "is_subtotal": adjustment_row.get("kind") == "subtotal",
                        "adjustment_label": adjustment_row["label"],
                        "adjustment_value": adjustment_row["value"],
                        "adjustment_display": format_value(
                            adjustment_row["value"], adjustment_unit, table_meta["scale"]
                        ),
                        "unit": adjustment_unit,
                        "scale": table_meta["scale"],
                        "source_role": table_meta["source_role"],
                        "source_document": table_meta["source_document"],
                        "source_url": table_meta["source_url"],
                        "source_page": table_meta.get("source_page"),
                    }
                )

    # Remove exact duplicates caused by nested or repeated HTML tables.
    unique_pairs: list[dict[str, Any]] = []
    seen_pairs: set[tuple[Any, ...]] = set()
    for pair in pairs:
        key = (
            pair["metric"].lower(),
            pair["gaap_value"],
            pair["non_gaap_value"],
            pair["source_url"],
            pair.get("source_page"),
        )
        if key not in seen_pairs:
            unique_pairs.append(pair)
            seen_pairs.add(key)
    valid_pair_ids = {pair["pair_id"] for pair in unique_pairs}
    unique_adjustments = [item for item in adjustments if item["pair_id"] in valid_pair_ids]
    return unique_pairs, unique_adjustments


def extract_html_reconciliations(
    document_html: str,
    source: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    soup = BeautifulSoup(document_html, "html.parser")
    pairs: list[dict[str, Any]] = []
    adjustments: list[dict[str, Any]] = []
    evidence: list[dict[str, Any]] = []

    for table_number, table in enumerate(soup.find_all("table"), start=1):
        rows = html_table_rows(table)
        context = table_context(table)
        score = table_reconciliation_score(context, rows)
        if score < 8:
            continue
        table_text = clean_space(" ".join(" ".join(row) for row in rows))
        scale = extract_scale(context + " " + table_text[:1200])
        title = context or next((row[0] for row in rows if row and "reconciliation" in row[0].lower()), f"Table {table_number}")
        current_section = re.sub(r"\s+", " ", title).strip()
        parsed_rows: list[dict[str, Any]] = []

        for cells in rows:
            label, value_meta = row_label_and_value(cells)
            if not label:
                continue
            if value_meta is None:
                if "reconciliation" in label.lower() and len(label) <= 220:
                    current_section = label
                continue
            kind = row_kind(label)
            parsed_rows.append(
                {
                    "section": current_section,
                    "label": label,
                    "value": value_meta["value"],
                    "value_meta": value_meta,
                    "kind": kind,
                }
            )

        table_meta = {
            "title": title,
            "scale": scale,
            "source_role": source.get("role") or "EX-99 exhibit",
            "source_document": source.get("document") or "",
            "source_description": source.get("description") or "",
            "source_url": source.get("url") or "",
            "source_page": None,
        }
        table_pairs, table_adjustments = build_pairs_from_rows(parsed_rows, table_meta)
        pairs.extend(table_pairs)
        adjustments.extend(table_adjustments)
        evidence.append(
            {
                "source_url": source.get("url") or "",
                "source_document": source.get("document") or "",
                "source_role": source.get("role") or "",
                "table_number": table_number,
                "table_title": title,
                "reconciliation_score": score,
                "parsed_rows": len(parsed_rows),
                "parsed_pairs": len(table_pairs),
                "text_preview": table_text[:700],
            }
        )
    return pairs, adjustments, evidence



TEXT_ROW_START = re.compile(
    r"\b(?:GAAP\s|Non[- ]GAAP\s|Adjusted\s|Normalized\s|Core\s|"
    r"Stock[- ]based\s|Share[- ]based\s|Amortization\s|Depreciation\s|"
    r"Acquisition[- ]related\s|Restructuring\s|Transformation\s|Impairment\s|"
    r"Legal\s|Litigation\s|Severance\s|Tax effect\s|Estimated tax\s|Income tax\s|"
    r"Purchases of\s|Capital expenditures?\s|Free cash flow\s|Loss contingency\s|"
    r"Equity income\s|Release of reserves\s|Cumulative effect\s|Gains?\s|Loss(?:es)?\s)",
    re.I,
)


def _split_dense_reconciliation_line(line: str) -> list[str]:
    """Split slide/PDF text where several reconciliation rows share one text line."""
    line = clean_space(line)
    if not line:
        return []

    candidates: list[int] = []
    for match in TEXT_ROW_START.finditer(line):
        position = match.start()
        token = match.group(0).lower()
        before = line[max(0, position - 12) : position].lower()
        # Do not split the GAAP portion inside "Non-GAAP".
        if token.startswith("gaap") and re.search(r"non[- ]$", before):
            continue
        # Do not split a compound adjustment label such as
        # "Amortization of acquisition-related intangibles".
        if token.startswith("acquisition-related") and before.endswith("of "):
            continue
        candidates.append(position)

    if len(candidates) <= 1:
        return [line]

    raw_segments: list[tuple[int, int, str]] = []
    for index, position in enumerate(candidates):
        end = candidates[index + 1] if index + 1 < len(candidates) else len(line)
        raw_segments.append((position, end, clean_space(line[position:end])))

    valid_positions: list[int] = []
    valid_segments: dict[int, str] = {}
    for position, _, segment in raw_segments:
        label, value_meta = _pdf_line_to_row(segment)
        if not label or value_meta is None:
            continue
        lower_label = label.lower()
        if "reconciliation" in lower_label or lower_label.endswith(" to"):
            continue
        # Period headers such as Q2'26 can look numeric to the generic value parser.
        if abs(float(value_meta["value"])) <= 4 and re.search(r"\bq[1-4]\s*['’]?\d{2,4}\b", segment, re.I):
            continue
        valid_positions.append(position)
        valid_segments[position] = segment

    if not valid_positions:
        return [line]

    first = min(valid_positions)
    parts: list[str] = []
    prefix = clean_space(line[:first])
    if prefix:
        parts.append(prefix)
    for position in sorted(valid_positions):
        parts.append(valid_segments[position])
    return parts


def extract_text_reconciliations(
    text: str,
    source: dict[str, Any],
    source_page: Optional[int] = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Fallback parser for slide-style HTML or PDF text that is not an HTML table.

    The routine activates only around reconciliation headings and requires both a
    GAAP and a non-GAAP/adjusted row before it emits a metric pair. It is therefore
    intentionally conservative and is used to supplement, not replace, table parsing.
    """
    raw_lines = [clean_space(line) for line in (text or "").splitlines() if clean_space(line)]
    lines: list[str] = []
    for line in raw_lines:
        lines.extend(_split_dense_reconciliation_line(line))

    parsed_rows: list[dict[str, Any]] = []
    evidence: list[dict[str, Any]] = []
    current_section = ""
    active_budget = 0
    section_started_at = 0

    for line_number, line in enumerate(lines, start=1):
        lower = line.lower()
        if "reconciliation" in lower and len(line) <= 420:
            current_section = line
            active_budget = 90
            section_started_at = line_number
            continue
        if active_budget <= 0:
            continue
        active_budget -= 1

        label, value_meta = _pdf_line_to_row(line)
        if not label or value_meta is None:
            # A new short reconciliation heading can appear after a prior bridge.
            if "reconciliation" in lower and len(line) <= 420:
                current_section = line
                active_budget = 90
                section_started_at = line_number
            continue
        parsed_rows.append(
            {
                "section": current_section or "Text reconciliation",
                "label": label,
                "value": value_meta["value"],
                "value_meta": value_meta,
                "kind": row_kind(label),
            }
        )

    if not parsed_rows:
        return [], [], []

    scale = extract_scale(text[:100000])
    table_meta = {
        "title": current_section or "Text reconciliation",
        "scale": scale,
        "source_role": source.get("role") or "EX-99 exhibit",
        "source_document": source.get("document") or "",
        "source_description": source.get("description") or "",
        "source_url": source.get("url") or "",
        "source_page": source_page,
    }
    pairs, adjustments = build_pairs_from_rows(parsed_rows, table_meta)
    if pairs:
        evidence.append(
            {
                "source_url": source.get("url") or "",
                "source_document": source.get("document") or "",
                "source_role": source.get("role") or "",
                "table_number": None,
                "source_page": source_page,
                "table_title": current_section or "Text reconciliation",
                "reconciliation_score": 7,
                "parsed_rows": len(parsed_rows),
                "parsed_pairs": len(pairs),
                "text_preview": clean_space(text)[:700],
                "parser": "text fallback",
                "section_start_line": section_started_at,
            }
        )
    return pairs, adjustments, evidence


def _merge_extractions(
    primary_pairs: list[dict[str, Any]],
    primary_adjustments: list[dict[str, Any]],
    extra_pairs: list[dict[str, Any]],
    extra_adjustments: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    pairs: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    accepted_ids: set[str] = set()
    for pair in [*primary_pairs, *extra_pairs]:
        key = (
            clean_space(pair.get("metric")).lower(),
            pair.get("gaap_value"),
            pair.get("non_gaap_value"),
            pair.get("unit"),
            pair.get("source_url"),
            pair.get("source_page"),
        )
        if key in seen:
            continue
        seen.add(key)
        pairs.append(pair)
        accepted_ids.add(str(pair.get("pair_id") or ""))
    adjustments = [
        item
        for item in [*primary_adjustments, *extra_adjustments]
        if str(item.get("pair_id") or "") in accepted_ids
    ]
    return pairs, adjustments


def extract_pdf_text(content: bytes, max_pages: Optional[int] = None) -> tuple[str, list[str]]:
    reader = PdfReader(io.BytesIO(content))
    pages: list[str] = []
    page_count = len(reader.pages) if max_pages is None else min(len(reader.pages), max_pages)
    for page_index in range(page_count):
        try:
            pages.append(reader.pages[page_index].extract_text() or "")
        except Exception:
            pages.append("")
    return "\n\n".join(pages), pages


def _pdf_line_to_row(line: str) -> tuple[str, Optional[dict[str, Any]]]:
    line = clean_space(line)
    if not line or not re.search(r"[A-Za-z]", line):
        return "", None
    if HEADER_LABEL.search(line) and not re.match(r"^(?:GAAP|Non[- ]GAAP|Adjusted)", line, re.I):
        return line, None
    match = VALUE_PATTERN.search(line)
    if not match:
        return line, None
    # Prefer a value after the descriptive label. If the first match is a footnote, find the next.
    matches = list(VALUE_PATTERN.finditer(line))
    selected = None
    for candidate in matches:
        prefix = line[: candidate.start()].strip()
        if re.search(r"[A-Za-z]", prefix) and not re.fullmatch(r".*\(\d{1,2}\)\s*", prefix):
            selected = candidate
            break
    if selected is None:
        return line, None
    label = clean_space(line[: selected.start()])
    value_meta = parse_first_value([line[selected.start() :]])
    return label, value_meta


def extract_pdf_reconciliations(
    content: bytes,
    source: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    full_text, page_texts = extract_pdf_text(content)
    pairs: list[dict[str, Any]] = []
    adjustments: list[dict[str, Any]] = []
    evidence: list[dict[str, Any]] = []
    warnings: list[str] = []

    if not clean_space(full_text):
        warnings.append(f"{source.get('document') or 'PDF'} contains no extractable text; it may be image-only.")
        return pairs, adjustments, evidence, warnings

    for page_number, page_text in enumerate(page_texts, start=1):
        if not re.search(r"reconciliation", page_text, re.I):
            continue
        if not (re.search(r"\bgaap\b", page_text, re.I) and re.search(r"non[- ]gaap|adjusted|free cash flow|ebitda", page_text, re.I)):
            continue
        scale = extract_scale(page_text)
        current_section = f"PDF reconciliation page {page_number}"
        parsed_rows: list[dict[str, Any]] = []
        for raw_line in page_text.splitlines():
            line = clean_space(raw_line)
            if not line:
                continue
            if "reconciliation" in line.lower() and len(line) <= 220:
                current_section = line
                continue
            label, value_meta = _pdf_line_to_row(line)
            if not label or value_meta is None:
                continue
            parsed_rows.append(
                {
                    "section": current_section,
                    "label": label,
                    "value": value_meta["value"],
                    "value_meta": value_meta,
                    "kind": row_kind(label),
                }
            )

        table_meta = {
            "title": current_section,
            "scale": scale,
            "source_role": source.get("role") or "Investor presentation",
            "source_document": source.get("document") or "",
            "source_description": source.get("description") or "",
            "source_url": source.get("url") or "",
            "source_page": page_number,
        }
        page_pairs, page_adjustments = build_pairs_from_rows(parsed_rows, table_meta)
        for pair in page_pairs:
            pair["confidence"] = "Medium"
        pairs.extend(page_pairs)
        adjustments.extend(page_adjustments)
        evidence.append(
            {
                "source_url": source.get("url") or "",
                "source_document": source.get("document") or "",
                "source_role": source.get("role") or "",
                "table_number": None,
                "source_page": page_number,
                "table_title": current_section,
                "reconciliation_score": 8,
                "parsed_rows": len(parsed_rows),
                "parsed_pairs": len(page_pairs),
                "text_preview": clean_space(page_text)[:700],
            }
        )
    return pairs, adjustments, evidence, warnings


def _metric_name_match(metric: str, reconciled_metrics: set[str]) -> bool:
    key = base_metric_key(metric)
    for reconciled in reconciled_metrics:
        other = base_metric_key(reconciled)
        if key == other or SequenceMatcher(None, key, other).ratio() >= 0.68:
            return True
    return False


def _reported_number(text: str, match: re.Match[str]) -> str:
    """Return the closest reported value to a metric mention.

    Values on either side of the metric name are ranked by character distance.
    Scale words are retained so that ``$42.2 million`` is more useful than a bare
    ``$42.2`` in the callout table.
    """

    candidates: list[tuple[int, str]] = []
    windows = [
        (text[match.end() : min(len(text), match.end() + 150)], match.end(), "after"),
        (text[max(0, match.start() - 110) : match.start()], max(0, match.start() - 110), "before"),
    ]
    for window, offset, direction in windows:
        for number in VALUE_PATTERN.finditer(window):
            raw = clean_space(number.group(0))
            numeric = number.group("number") or ""
            if not raw or (numeric.isdigit() and 1900 <= int(numeric) <= 2100):
                continue
            suffix_text = window[number.end() : number.end() + 18]
            scale_match = re.match(r"\s*(thousand|million|billion)s?\b", suffix_text, re.I)
            if scale_match:
                raw = f"{raw} {scale_match.group(1).lower()}"
            absolute_start = offset + number.start()
            absolute_end = offset + number.end()
            distance = absolute_start - match.end() if direction == "after" else match.start() - absolute_end
            candidates.append((max(0, distance), raw))
    if not candidates:
        return ""
    candidates.sort(key=lambda item: item[0])
    return candidates[0][1]


def extract_metric_mentions(
    text: str,
    source: dict[str, Any],
    reconciled_metrics: set[str],
    source_page: Optional[int] = None,
) -> list[dict[str, Any]]:
    mentions: list[dict[str, Any]] = []
    compact = re.sub(r"[ \t]+", " ", text or "")
    for canonical, pattern in METRIC_PATTERNS:
        candidates: list[dict[str, Any]] = []
        for match in re.finditer(pattern, compact, re.I):
            start = max(0, match.start() - 180)
            end = min(len(compact), match.end() + 260)
            snippet = clean_space(compact[start:end])
            numbers = [clean_space(item.group(0)) for item in VALUE_PATTERN.finditer(snippet)]
            primary_value = _reported_number(compact, match)
            boilerplate = bool(re.search(r"management believes|not a substitute|definitions?|reconciliation", snippet, re.I))
            rank = min(4, len(numbers)) * 3 + (5 if primary_value else 0) + (0 if boilerplate else 4)
            candidates.append(
                {
                    "metric": canonical,
                    "context": snippet,
                    "primary_value": primary_value,
                    "reported_values": ", ".join(numbers[:4]),
                    "rank": rank,
                }
            )
        if not candidates:
            continue
        best = max(candidates, key=lambda item: item["rank"])
        mentions.append(
            {
                "metric": canonical,
                "status": "Reconciled in an 8-K exhibit" if _metric_name_match(canonical, reconciled_metrics) else "Additional non-GAAP measure",
                "primary_value": best["primary_value"],
                "context": best["context"],
                "reported_values": best["reported_values"],
                "source_role": source.get("role") or "EX-99 exhibit",
                "source_document": source.get("document") or "",
                "source_description": source.get("description") or "",
                "source_url": source.get("url") or "",
                "source_page": source_page,
            }
        )
    return mentions




BENCHMARK_MEASURE_ORDER = [
    "Non-GAAP income tax rate",
    "Non-GAAP other income / expense",
    "Non-GAAP operating expenses",
    "Adjusted EBITDA",
    "Non-GAAP income from operations",
    "Non-GAAP income from operations %",
    "Non-GAAP gross margin / gross profit",
    "Non-GAAP gross margin %",
    "Non-GAAP net income",
    "Non-GAAP EPS - basic",
    "Non-GAAP EPS - diluted",
    "Free cash flow",
]

KPI_PATTERNS: list[tuple[str, str]] = [
    ("Bookings / book-to-bill / backlog", r"\bbookings?\b|\bbook[- ]to[- ]bill\b|\bbacklog\b"),
    ("Capital expenditures", r"\bcapital expenditures?\b|\bcapex\b|purchases? of (?:property|plant|equipment)"),
    ("Data center revenue growth", r"\bdata cent(?:er|re)\b.{0,100}\b(?:revenue|sales)\b.{0,80}\b(?:growth|grew|increase|up|%)\b"),
    ("Free cash flow / FCF margin", r"\b(?:adjusted )?free cash flow\b|\bfcf margin\b"),
    ("Gross margin", r"\bgross (?:margin|profit)(?: percentage| %)\b|\bgross margin\b"),
    ("Inventory days / months on hand", r"\binventory\b.{0,80}\b(?:days?|months?)\b.{0,30}\b(?:on hand|on-hand|supply|inventory)\b|\bdays inventory\b"),
    ("Operating cash flow", r"\boperating cash flow\b|\bnet cash (?:provided by|from) operating activities\b"),
    ("Operating expenses", r"\boperating expenses?\b|\bopex\b"),
    ("Operating income / margin", r"\boperating (?:income|profit|margin)\b|\bincome from operations\b"),
    ("Revenue growth", r"\brevenue\b.{0,90}\b(?:growth|grew|increase(?:d)?|decrease(?:d)?|up|down|year[- ]over[- ]year|yoy)\b|\b(?:growth|grew|increase(?:d)?|decrease(?:d)?)\b.{0,90}\brevenue\b"),
    ("Revenue", r"\b(?:net )?(?:revenue|sales)\b"),
]

KPI_ORDER = {name: index for index, (name, _pattern) in enumerate(KPI_PATTERNS, start=1)}


def benchmark_metric_family(metric: str, gaap_label: str = "", non_gaap_label: str = "") -> str:
    """Map issuer-specific metric wording to a presentation-ready peer family."""
    text = clean_space(" ".join([metric, gaap_label, non_gaap_label])).lower().replace("–", "-")
    if re.search(r"income tax.*(?:rate|percentage)|tax rate", text):
        return "Non-GAAP income tax rate"
    if re.search(r"other (?:income|expense|loss)|interest and other", text):
        return "Non-GAAP other income / expense"
    if re.search(r"operating expenses?|opex", text):
        return "Non-GAAP operating expenses"
    if "ebitda" in text:
        return "Adjusted EBITDA"
    if re.search(r"operating (?:income|profit|margin)|income from operations", text):
        if re.search(r"margin|percentage|%", text):
            return "Non-GAAP income from operations %"
        return "Non-GAAP income from operations"
    if re.search(r"gross (?:margin|profit)", text):
        if re.search(r"percentage|%", text):
            return "Non-GAAP gross margin %"
        return "Non-GAAP gross margin / gross profit"
    if re.search(r"net income|earnings", text):
        if re.search(r"per share|\beps\b", text):
            if "basic" in text and "diluted" not in text:
                return "Non-GAAP EPS - basic"
            return "Non-GAAP EPS - diluted"
        return "Non-GAAP net income"
    if "free cash flow" in text:
        return "Free cash flow"
    return clean_space(metric) or "Other non-GAAP measure"


def extract_kpi_mentions(
    text: str,
    source: dict[str, Any],
    source_page: Optional[int] = None,
) -> list[dict[str, Any]]:
    """Identify operating KPIs discussed in the same 8-K exhibit package.

    The result is a disclosure-presence aid rather than an accounting conclusion.
    Exact context and the closest reported value are retained for auditability.
    """
    compact = re.sub(r"[ \t]+", " ", text or "")
    rows: list[dict[str, Any]] = []
    for kpi, pattern in KPI_PATTERNS:
        candidates: list[dict[str, Any]] = []
        for match in re.finditer(pattern, compact, re.I | re.S):
            start = max(0, match.start() - 160)
            end = min(len(compact), match.end() + 220)
            context = clean_space(compact[start:end])
            primary_value = _reported_number(compact, match)
            numeric_count = len(list(VALUE_PATTERN.finditer(context)))
            boilerplate = bool(re.search(r"definition|management believes|not a substitute", context, re.I))
            score = numeric_count * 2 + (5 if primary_value else 0) + (0 if boilerplate else 3)
            candidates.append({"context": context, "primary_value": primary_value, "score": score})
        if not candidates:
            continue
        best = max(candidates, key=lambda item: item["score"])
        rows.append(
            {
                "kpi": kpi,
                "primary_value": best["primary_value"],
                "context": best["context"],
                "source_role": source.get("role") or "EX-99 exhibit",
                "source_document": source.get("document") or "",
                "source_description": source.get("description") or "",
                "source_url": source.get("url") or "",
                "source_page": source_page,
            }
        )
    return rows


def make_peer_presence_matrix(
    frame: pd.DataFrame,
    row_field: str,
    company_field: str = "company",
    ordered_rows: Optional[list[str]] = None,
    minimum_companies: int = 1,
) -> pd.DataFrame:
    """Create a dot matrix like the peer-benchmarking slides."""
    if frame is None or frame.empty or row_field not in frame.columns or company_field not in frame.columns:
        return pd.DataFrame()
    data = frame[[row_field, company_field]].dropna().copy()
    data[row_field] = data[row_field].map(clean_space)
    data[company_field] = data[company_field].map(clean_space)
    data = data[(data[row_field] != "") & (data[company_field] != "")].drop_duplicates()
    if data.empty:
        return pd.DataFrame()
    matrix = pd.crosstab(data[row_field], data[company_field]).clip(upper=1)
    matrix["Total"] = matrix.sum(axis=1)
    matrix = matrix[matrix["Total"] >= max(1, int(minimum_companies))]
    if matrix.empty:
        return pd.DataFrame()
    order_map = {value: index for index, value in enumerate(ordered_rows or [], start=0)}
    matrix["_order"] = [order_map.get(index, 9999) for index in matrix.index]
    matrix["_name"] = matrix.index.astype(str)
    matrix = matrix.sort_values(["_order", "Total", "_name"], ascending=[True, False, True]).drop(columns=["_order", "_name"])
    company_columns = sorted([column for column in matrix.columns if column != "Total"])
    matrix = matrix[company_columns + ["Total"]]
    matrix = matrix.reset_index().rename(columns={row_field: "Disclosure"})
    for column in company_columns:
        matrix[column] = matrix[column].map(lambda value: "●" if int(value) else "")
    return matrix


def make_reconciliation_bridge_table(
    reconciliations: pd.DataFrame,
    adjustments: pd.DataFrame,
    metric: str,
    periods: Optional[list[str]] = None,
) -> pd.DataFrame:
    """Return GAAP -> individual adjustments -> non-GAAP in period columns.

    Exact issuer labels are retained.  The function is intentionally suitable for
    direct display, Excel export, or a presentation-style HTML renderer.
    """
    columns = ["Line item", "Row type"] + list(periods or [])
    if reconciliations is None or reconciliations.empty:
        return pd.DataFrame(columns=columns)
    pairs = reconciliations[reconciliations["metric"].astype(str).eq(str(metric))].copy()
    if pairs.empty:
        return pd.DataFrame(columns=columns)
    available_periods = ordered_fiscal_periods(pairs)
    selected_periods = [period for period in (periods or available_periods) if period in available_periods]
    if not selected_periods:
        selected_periods = available_periods[-2:]
    pairs = pairs[pairs["period"].astype(str).isin(selected_periods)].copy()
    if pairs.empty:
        return pd.DataFrame(columns=["Line item", "Row type"] + selected_periods)

    pair_by_period = {
        str(row["period"]): row for _, row in pairs.sort_values(["fiscal_year", "fiscal_quarter"]).iterrows()
    }
    adjustment_frame = adjustments.copy() if isinstance(adjustments, pd.DataFrame) else pd.DataFrame()
    if not adjustment_frame.empty:
        adjustment_frame = adjustment_frame[adjustment_frame["metric"].astype(str).eq(str(metric))].copy()

    line_order: list[str] = []
    line_meta: dict[str, dict[str, str]] = {}
    for period in selected_periods:
        pair = pair_by_period.get(period)
        if pair is None:
            continue
        gaap_key = "__GAAP__"
        if gaap_key not in line_order:
            line_order.append(gaap_key)
        line_meta[gaap_key] = {"label": clean_space(pair.get("gaap_label")) or "Comparable GAAP measure", "type": "GAAP"}
        pair_id = clean_space(pair.get("pair_id"))
        period_adjustments = adjustment_frame[
            adjustment_frame.get("pair_id", pd.Series(dtype=str)).astype(str).eq(pair_id)
        ] if not adjustment_frame.empty and "pair_id" in adjustment_frame.columns else pd.DataFrame()
        if not period_adjustments.empty:
            sort_columns = [column for column in ["adjustment_order", "adjustment_label"] if column in period_adjustments.columns]
            if sort_columns:
                period_adjustments = period_adjustments.sort_values(sort_columns)
            for _, adjustment in period_adjustments.iterrows():
                label = clean_space(adjustment.get("adjustment_label")) or "Adjustment"
                key = f"ADJ::{label.lower()}"
                if key not in line_order:
                    line_order.append(key)
                line_meta[key] = {"label": label, "type": "Adjustment"}
        else:
            # Keep the endpoint difference visible even when the source table's
            # individual lines could not be parsed. This makes the review gap
            # explicit rather than showing a silent GAAP-to-non-GAAP jump.
            total_key = "__TOTAL_ADJUSTMENTS__"
            if total_key not in line_order:
                line_order.append(total_key)
            line_meta[total_key] = {
                "label": "Total adjustments (individual items not parsed)",
                "type": "Adjustment",
            }
        ng_key = "__NON_GAAP__"
        if ng_key not in line_order:
            line_order.append(ng_key)
        line_meta[ng_key] = {"label": clean_space(pair.get("non_gaap_label")) or metric, "type": "Non-GAAP"}

    rows: list[dict[str, Any]] = []
    for key in line_order:
        row: dict[str, Any] = {"Line item": line_meta[key]["label"], "Row type": line_meta[key]["type"]}
        for period in selected_periods:
            pair = pair_by_period.get(period)
            display = "—"
            if pair is not None:
                if key == "__GAAP__":
                    display = clean_space(pair.get("gaap_display")) or "—"
                elif key == "__NON_GAAP__":
                    display = clean_space(pair.get("non_gaap_display")) or "—"
                elif key == "__TOTAL_ADJUSTMENTS__":
                    pair_id = clean_space(pair.get("pair_id"))
                    has_detail = (
                        not adjustment_frame.empty
                        and "pair_id" in adjustment_frame.columns
                        and adjustment_frame["pair_id"].astype(str).eq(pair_id).any()
                    )
                    display = "—" if has_detail else (clean_space(pair.get("adjustment_display")) or "—")
                else:
                    pair_id = clean_space(pair.get("pair_id"))
                    label_key = key.removeprefix("ADJ::")
                    candidates = adjustment_frame[
                        adjustment_frame.get("pair_id", pd.Series(dtype=str)).astype(str).eq(pair_id)
                        & adjustment_frame.get("adjustment_label", pd.Series(dtype=str)).astype(str).str.lower().eq(label_key)
                    ] if not adjustment_frame.empty and {"pair_id", "adjustment_label"}.issubset(adjustment_frame.columns) else pd.DataFrame()
                    if not candidates.empty:
                        display = clean_space(candidates.iloc[0].get("adjustment_display")) or "—"
            row[period] = display
        rows.append(row)
    return pd.DataFrame(rows, columns=["Line item", "Row type"] + selected_periods)

def parse_exhibit(
    client: SecClient,
    exhibit: dict[str, Any],
) -> dict[str, Any]:
    resource = client.get_bytes(exhibit["url"], timeout=90)
    warnings: list[str] = []
    if is_pdf_resource(resource, exhibit.get("url") or ""):
        full_text, pages = extract_pdf_text(resource.content)
        pairs, adjustments, evidence, pdf_warnings = extract_pdf_reconciliations(resource.content, exhibit)
        warnings.extend(pdf_warnings)
        if not pairs:
            fallback_pairs: list[dict[str, Any]] = []
            fallback_adjustments: list[dict[str, Any]] = []
            fallback_evidence: list[dict[str, Any]] = []
            for page_number, page_text in enumerate(pages, start=1):
                page_pairs, page_adjustments, page_evidence = extract_text_reconciliations(
                    page_text, exhibit, source_page=page_number
                )
                fallback_pairs.extend(page_pairs)
                fallback_adjustments.extend(page_adjustments)
                fallback_evidence.extend(page_evidence)
            pairs, adjustments = _merge_extractions(
                pairs, adjustments, fallback_pairs, fallback_adjustments
            )
            evidence.extend(fallback_evidence)
        reconciled_metrics = {pair["metric"] for pair in pairs}
        mentions: list[dict[str, Any]] = []
        kpis: list[dict[str, Any]] = []
        for page_number, page_text in enumerate(pages, start=1):
            mentions.extend(extract_metric_mentions(page_text, exhibit, reconciled_metrics, source_page=page_number))
            kpis.extend(extract_kpi_mentions(page_text, exhibit, source_page=page_number))
        document_type = "PDF"
    else:
        document_html = decode_document(resource.content)
        full_text = html_plain_text(document_html)
        pairs, adjustments, evidence = extract_html_reconciliations(document_html, exhibit)
        fallback_pairs, fallback_adjustments, fallback_evidence = extract_text_reconciliations(
            full_text, exhibit
        )
        pairs, adjustments = _merge_extractions(
            pairs, adjustments, fallback_pairs, fallback_adjustments
        )
        evidence.extend(fallback_evidence)
        soup = BeautifulSoup(document_html, "html.parser")
        if len(clean_space(full_text)) < 1000 and len(soup.find_all("img")) >= 3:
            warnings.append(
                f"{exhibit.get('document') or 'HTML exhibit'} is image-heavy and contains limited extractable text."
            )
        reconciled_metrics = {pair["metric"] for pair in pairs}
        mentions = extract_metric_mentions(full_text, exhibit, reconciled_metrics)
        kpis = extract_kpi_mentions(full_text, exhibit)
        document_type = "HTML"

    return {
        "document": exhibit.get("document") or "",
        "description": exhibit.get("description") or "",
        "doc_type": exhibit.get("doc_type") or "",
        "role": exhibit.get("role") or "Other EX-99 exhibit",
        "url": exhibit.get("url") or "",
        "content_type": document_type,
        "has_reconciliation": bool(pairs),
        "reconciliations": pairs,
        "adjustments": adjustments,
        "mentions": mentions,
        "kpis": kpis,
        "evidence": evidence,
        "warnings": warnings,
        "text_preview": clean_space(full_text)[:500],
    }


# ---------------------------------------------------------------------------
# End-to-end company analysis
# ---------------------------------------------------------------------------


def analyze_company_quarters(
    client: SecClient,
    cik: int,
    filings: pd.DataFrame,
    anchors: pd.DataFrame,
    selected_years: list[int],
    progress: Optional[Callable[[str], None]] = None,
    max_exhibits_per_8k: int = 8,
) -> dict[str, pd.DataFrame]:
    selected = anchors[anchors["fiscal_year"].isin(selected_years)].copy()
    selected["quarter_order"] = selected["fiscal_quarter"].map(QUARTER_ORDER)
    selected = selected.sort_values(["fiscal_year", "quarter_order"], ascending=[True, True])

    coverage_rows: list[dict[str, Any]] = []
    reconciliation_rows: list[dict[str, Any]] = []
    adjustment_rows: list[dict[str, Any]] = []
    mention_rows: list[dict[str, Any]] = []
    kpi_rows: list[dict[str, Any]] = []
    source_rows: list[dict[str, Any]] = []
    evidence_rows: list[dict[str, Any]] = []
    warning_rows: list[dict[str, Any]] = []

    for _, anchor_row in selected.iterrows():
        anchor = anchor_row.to_dict()
        fy = int(anchor["fiscal_year"])
        quarter = anchor["fiscal_quarter"]
        period_label = f"FY{fy} {quarter}"
        if progress:
            progress(f"Matching the earnings 8-K for {period_label}...")
        matched = match_earnings_8k(client, cik, filings, anchor, progress=progress)

        coverage = {
            "fiscal_year": fy,
            "fiscal_quarter": quarter,
            "period": period_label,
            "period_end": anchor.get("period_end"),
            "periodic_form": anchor.get("periodic_form"),
            "periodic_filing_date": anchor.get("periodic_filing_date"),
            "periodic_url": anchor.get("periodic_url"),
            "earnings_8k_found": bool(matched),
            "earnings_8k_filing_date": matched.get("filing_date") if matched else None,
            "earnings_8k_url": matched.get("primary_url") if matched else "",
            "earnings_8k_index_url": matched.get("index_url") if matched else "",
            "matching_score": matched.get("score") if matched else None,
            "matching_reasons": matched.get("reasons") if matched else "",
            "source_documents": 0,
            "reconciled_metrics": 0,
            "additional_measures": 0,
            "status": "No matching earnings 8-K found",
        }

        if not matched:
            coverage_rows.append(coverage)
            continue

        parsed_documents: list[dict[str, Any]] = []
        exhibits = matched.get("exhibits") or []
        for exhibit in exhibits[:max_exhibits_per_8k]:
            if progress:
                progress(f"Reading {period_label}: {exhibit.get('role')} ({exhibit.get('document')})...")
            try:
                parsed = parse_exhibit(client, exhibit)
                parsed_documents.append(parsed)
            except Exception as exc:
                warning_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "source_document": exhibit.get("document") or "",
                        "source_url": exhibit.get("url") or "",
                        "warning": f"Could not parse exhibit: {exc}",
                    }
                )

        # The reconciliation may be in the release while the presentation discusses more measures.
        all_reconciled_metrics = {
            pair["metric"]
            for parsed in parsed_documents
            for pair in parsed.get("reconciliations", [])
        }

        for parsed in parsed_documents:
            source_rows.append(
                {
                    "fiscal_year": fy,
                    "fiscal_quarter": quarter,
                    "period": period_label,
                    "8k_filing_date": matched.get("filing_date"),
                    "8k_form": matched.get("form"),
                    "8k_items": matched.get("items"),
                    "8k_url": matched.get("primary_url"),
                    "8k_index_url": matched.get("index_url"),
                    "document_role": parsed.get("role"),
                    "document": parsed.get("document"),
                    "description": parsed.get("description"),
                    "doc_type": parsed.get("doc_type"),
                    "content_type": parsed.get("content_type"),
                    "has_reconciliation": parsed.get("has_reconciliation"),
                    "document_url": parsed.get("url"),
                }
            )

            for pair in parsed.get("reconciliations", []):
                reconciliation_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "period_end": anchor.get("period_end"),
                        "8k_filing_date": matched.get("filing_date"),
                        "8k_url": matched.get("primary_url"),
                        **pair,
                    }
                )
            for item in parsed.get("adjustments", []):
                adjustment_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "period_end": anchor.get("period_end"),
                        **item,
                    }
                )
            for item in parsed.get("mentions", []):
                item = dict(item)
                item["status"] = (
                    "Reconciled in an 8-K exhibit"
                    if _metric_name_match(item["metric"], all_reconciled_metrics)
                    else "Additional non-GAAP measure"
                )
                item["metric_family"] = benchmark_metric_family(item.get("metric", ""))
                item["source_has_reconciliation"] = bool(parsed.get("has_reconciliation"))
                item["source_content_type"] = parsed.get("content_type") or ""
                mention_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "period_end": anchor.get("period_end"),
                        **item,
                    }
                )
            for item in parsed.get("kpis", []):
                kpi_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "period_end": anchor.get("period_end"),
                        **dict(item),
                    }
                )
            for evidence in parsed.get("evidence", []):
                evidence_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        **evidence,
                    }
                )
            for warning in parsed.get("warnings", []):
                warning_rows.append(
                    {
                        "fiscal_year": fy,
                        "fiscal_quarter": quarter,
                        "period": period_label,
                        "source_document": parsed.get("document"),
                        "source_url": parsed.get("url"),
                        "warning": warning,
                    }
                )

        quarter_reconciliations = [row for row in reconciliation_rows if row["period"] == period_label]
        quarter_additional = [
            row for row in mention_rows if row["period"] == period_label and row["status"] == "Additional non-GAAP measure"
        ]
        coverage["source_documents"] = len(parsed_documents)
        coverage["reconciled_metrics"] = len(quarter_reconciliations)
        coverage["additional_measures"] = len({row["metric"] for row in quarter_additional})
        if quarter_reconciliations:
            coverage["status"] = "Reconciliation metrics extracted"
        elif parsed_documents:
            coverage["status"] = "8-K found; no structured reconciliation parsed"
        else:
            coverage["status"] = "8-K found; exhibits could not be parsed"
        coverage_rows.append(coverage)

    coverage_df = pd.DataFrame(coverage_rows)
    reconciliations_df = pd.DataFrame(reconciliation_rows)
    adjustments_df = pd.DataFrame(adjustment_rows)
    mentions_df = pd.DataFrame(mention_rows)
    kpis_df = pd.DataFrame(kpi_rows)
    sources_df = pd.DataFrame(source_rows)
    evidence_df = pd.DataFrame(evidence_rows)
    warnings_df = pd.DataFrame(warning_rows)

    if not reconciliations_df.empty:
        reconciliations_df["metric_family"] = reconciliations_df.apply(
            lambda row: benchmark_metric_family(
                row.get("metric", ""), row.get("gaap_label", ""), row.get("non_gaap_label", "")
            ),
            axis=1,
        )
        role_rank = {"Press release": 0, "Financial supplement": 1, "Investor presentation": 2, "Other EX-99 exhibit": 3}
        reconciliations_df["role_rank"] = reconciliations_df["source_role"].map(role_rank).fillna(9)
        reconciliations_df = reconciliations_df.sort_values(
            ["fiscal_year", "fiscal_quarter", "metric", "role_rank", "confidence"],
            ascending=[True, True, True, True, True],
        )
        reconciliations_df = reconciliations_df.drop_duplicates(
            ["fiscal_year", "fiscal_quarter", "metric", "non_gaap_value", "unit"], keep="first"
        ).drop(columns=["role_rank"])

    if not adjustments_df.empty:
        if not reconciliations_df.empty and "pair_id" in reconciliations_df.columns:
            valid_pair_ids = set(reconciliations_df["pair_id"].dropna().astype(str).tolist())
            adjustments_df = adjustments_df[adjustments_df["pair_id"].astype(str).isin(valid_pair_ids)].copy()
            family_map = reconciliations_df.drop_duplicates("pair_id").set_index("pair_id")["metric_family"].to_dict()
            adjustments_df["metric_family"] = adjustments_df["pair_id"].map(family_map).fillna(
                adjustments_df["metric"].map(benchmark_metric_family)
            )
        elif "metric" in adjustments_df.columns:
            adjustments_df["metric_family"] = adjustments_df["metric"].map(benchmark_metric_family)
        duplicate_columns = [
            column
            for column in [
                "fiscal_year",
                "fiscal_quarter",
                "pair_id",
                "adjustment_label",
                "adjustment_value",
                "unit",
                "scale",
                "source_url",
                "source_page",
            ]
            if column in adjustments_df.columns
        ]
        if duplicate_columns:
            adjustments_df = adjustments_df.drop_duplicates(duplicate_columns, keep="first")
        adjustments_df["_period_rank"] = adjustments_df.apply(
            lambda row: fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
        )
        adjustment_sort = ["_period_rank", "metric"]
        if "adjustment_order" in adjustments_df.columns:
            adjustment_sort.append("adjustment_order")
        adjustment_sort.append("adjustment_label")
        adjustments_df = adjustments_df.sort_values(adjustment_sort).drop(
            columns=["_period_rank"]
        ).reset_index(drop=True)

    if not mentions_df.empty:
        mentions_df = mentions_df.drop_duplicates(
            ["fiscal_year", "fiscal_quarter", "metric", "source_url", "source_page"], keep="first"
        )

    if not kpis_df.empty:
        kpis_df = kpis_df.drop_duplicates(
            ["fiscal_year", "fiscal_quarter", "kpi", "source_url", "source_page"], keep="first"
        )
        kpis_df["kpi_order"] = kpis_df["kpi"].map(KPI_ORDER).fillna(999).astype(int)
        kpis_df = kpis_df.sort_values(["fiscal_year", "fiscal_quarter", "kpi_order", "kpi"]).reset_index(drop=True)

    adjustment_history_df = enrich_adjustments(adjustments_df, reconciliations_df)
    adjustment_tieouts_df = build_adjustment_tieouts(reconciliations_df, adjustment_history_df)

    return {
        "coverage": coverage_df,
        "reconciliations": reconciliations_df,
        "adjustments": adjustments_df,
        "adjustment_history": adjustment_history_df,
        "adjustment_tieouts": adjustment_tieouts_df,
        "mentions": mentions_df,
        "kpis": kpis_df,
        "sources": sources_df,
        "evidence": evidence_df,
        "warnings": warnings_df,
    }


def make_metric_matrix(reconciliations: pd.DataFrame, include_gaap: bool = False) -> pd.DataFrame:
    if reconciliations.empty:
        return pd.DataFrame()
    data = reconciliations.copy()
    data["quarter_order"] = data["fiscal_quarter"].map(QUARTER_ORDER)
    data = data.sort_values(["fiscal_year", "quarter_order", "metric"])
    if include_gaap:
        data["matrix_value"] = data.apply(
            lambda row: f"{row['non_gaap_display']} (GAAP {row['gaap_display']})", axis=1
        )
    else:
        data["matrix_value"] = data["non_gaap_display"]
    matrix = data.pivot_table(index="metric", columns="period", values="matrix_value", aggfunc="first")
    ordered_periods = (
        data[["period", "fiscal_year", "quarter_order"]]
        .drop_duplicates()
        .sort_values(["fiscal_year", "quarter_order"])["period"]
        .tolist()
    )
    return matrix.reindex(columns=ordered_periods).reset_index()


def build_export_zip(analysis: dict[str, pd.DataFrame]) -> bytes:
    import zipfile

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, frame in analysis.items():
            if isinstance(frame, pd.DataFrame):
                archive.writestr(f"{name}.csv", frame.to_csv(index=False))
    return output.getvalue()


def build_excel_report(company: dict[str, Any], analysis: dict[str, pd.DataFrame]) -> bytes:
    """Create a formatted, source-linked Excel workbook for a completed analysis."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dark_blue = "17365D"
    medium_blue = "2F5597"
    light_blue = "D9EAF7"
    light_green = "E2F0D9"
    light_orange = "FCE4D6"
    white = "FFFFFF"
    gray = "666666"
    imported_green = "008000"
    thin_gray = Side(style="thin", color="D9E1F2")

    workbook = Workbook()
    workbook.remove(workbook.active)

    def safe_sheet_title(value: str) -> str:
        value = re.sub(r"[\\/*?:\[\]]", " ", value).strip() or "Sheet"
        return value[:31]

    def write_frame(sheet_name: str, frame: pd.DataFrame, title: str) -> None:
        sheet = workbook.create_sheet(safe_sheet_title(sheet_name))
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, color=white, size=14)
        sheet["A1"].fill = PatternFill("solid", fgColor=dark_blue)
        width = max(1, len(frame.columns) if not frame.empty else 1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)

        if frame.empty:
            sheet["A3"] = "No records extracted."
            sheet["A3"].font = Font(color=gray, italic=True)
            sheet.column_dimensions["A"].width = 34
            return

        clean_frame = frame.copy()
        for column in clean_frame.columns:
            if pd.api.types.is_datetime64_any_dtype(clean_frame[column]):
                clean_frame[column] = clean_frame[column].dt.strftime("%Y-%m-%d")

        header_row = 3
        for column_index, column_name in enumerate(clean_frame.columns, start=1):
            cell = sheet.cell(row=header_row, column=column_index, value=str(column_name).replace("_", " ").title())
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=medium_blue)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[header_row].height = 30

        for row_index, record in enumerate(clean_frame.itertuples(index=False, name=None), start=header_row + 1):
            for column_index, value in enumerate(record, start=1):
                if value is None or (isinstance(value, float) and math.isnan(value)):
                    value = None
                if isinstance(value, (date, datetime, pd.Timestamp)):
                    value = pd.Timestamp(value).date().isoformat()
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin_gray)
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = '#,##0.0;[Red](#,##0.0);-'
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                    cell.font = Font(color=imported_green)
                elif isinstance(value, str) and value.startswith("https://"):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                elif value not in (None, ""):
                    cell.font = Font(color=imported_green)

                column_name = str(clean_frame.columns[column_index - 1]).lower()
                if column_name in {"warning", "status", "matching_reasons", "confidence"} and value:
                    text = str(value).lower()
                    if any(flag in text for flag in ["no matching", "could not", "image-only", "failed", "medium", "low"]):
                        cell.fill = PatternFill("solid", fgColor=light_orange)

        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(width)}{sheet.max_row}"
        for column_index, column_name in enumerate(clean_frame.columns, start=1):
            samples = [str(column_name)] + [str(value) for value in clean_frame.iloc[:100, column_index - 1].fillna("").tolist()]
            max_length = max(len(value) for value in samples)
            if any(token in str(column_name).lower() for token in ["context", "description", "reason", "warning", "preview"]):
                max_length = max(max_length, 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(58, max(11, max_length + 2))

    summary = workbook.create_sheet("Summary")
    summary.sheet_view.showGridLines = False
    summary["A1"] = "SEC Non-GAAP Reconciliation Analysis"
    summary["A1"].font = Font(bold=True, color=white, size=17)
    summary["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    summary.merge_cells("A1:F1")

    metadata = [
        ("Company", company.get("name", "")),
        ("Ticker", company.get("ticker", "")),
        ("CIK", company.get("cik", "")),
        ("SIC", company.get("sic", "")),
        ("Industry", company.get("sic_description", "")),
        ("Fiscal year end", company.get("fiscal_year_end", "")),
        (
            "Source policy",
            "10-Q and 10-K filings are used only as fiscal-period anchors. Metric values are extracted only from EX-99 exhibits attached to matched earnings 8-K/8-K-A filings.",
        ),
    ]
    for row_index, (label, value) in enumerate(metadata, start=3):
        summary.cell(row=row_index, column=1, value=label).font = Font(bold=True, color=gray)
        value_cell = summary.cell(row=row_index, column=2, value=value)
        value_cell.font = Font(color=imported_green)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
    summary.column_dimensions["A"].width = 23
    summary.column_dimensions["B"].width = 100

    coverage = analysis.get("coverage", pd.DataFrame())
    reconciliations = analysis.get("reconciliations", pd.DataFrame())
    mentions = analysis.get("mentions", pd.DataFrame())
    sources = analysis.get("sources", pd.DataFrame())
    metrics_count = int(reconciliations["metric"].nunique()) if not reconciliations.empty and "metric" in reconciliations else 0
    complete_periods = int((coverage.get("status", pd.Series(dtype=str)) == "Reconciliation metrics extracted").sum()) if not coverage.empty else 0
    additional_count = int(
        mentions.loc[mentions.get("status", pd.Series(index=mentions.index, dtype=str)) == "Additional non-GAAP measure", "metric"].nunique()
    ) if not mentions.empty and "metric" in mentions else 0
    presentation_count = int((sources.get("document_role", pd.Series(dtype=str)) == "Investor presentation").sum()) if not sources.empty else 0

    summary["A11"] = "Coverage and output"
    summary["A11"].font = Font(bold=True, color=white)
    summary["A11"].fill = PatternFill("solid", fgColor=medium_blue)
    summary.merge_cells("A11:B11")
    kpis = [
        ("Periods with reconciliations", complete_periods),
        ("Distinct reconciled metrics", metrics_count),
        ("Additional non-GAAP measures", additional_count),
        ("Investor presentations checked", presentation_count),
    ]
    for row_index, (label, value) in enumerate(kpis, start=12):
        summary.cell(row=row_index, column=1, value=label).font = Font(bold=True, color=gray)
        summary.cell(row=row_index, column=2, value=value).font = Font(bold=True, color=imported_green)

    matrix = make_metric_matrix(reconciliations)
    matrix_start = 18
    summary.cell(row=matrix_start, column=1, value="Reconciled metric matrix").font = Font(bold=True, color=white)
    summary.cell(row=matrix_start, column=1).fill = PatternFill("solid", fgColor=medium_blue)
    matrix_width = max(1, len(matrix.columns))
    summary.merge_cells(start_row=matrix_start, start_column=1, end_row=matrix_start, end_column=matrix_width)
    if not matrix.empty:
        for column_index, column_name in enumerate(matrix.columns, start=1):
            cell = summary.cell(row=matrix_start + 1, column=column_index, value=str(column_name))
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=medium_blue)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_index, record in enumerate(matrix.itertuples(index=False, name=None), start=matrix_start + 2):
            for column_index, value in enumerate(record, start=1):
                cell = summary.cell(row=row_index, column=column_index, value=None if pd.isna(value) else value)
                cell.border = Border(bottom=thin_gray)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column_index == 1:
                    cell.font = Font(bold=True, color=gray)
                    cell.fill = PatternFill("solid", fgColor=light_blue)
                else:
                    cell.font = Font(color=imported_green)
                    cell.fill = PatternFill("solid", fgColor=light_green)
        for column_index, column_name in enumerate(matrix.columns, start=1):
            summary.column_dimensions[get_column_letter(column_index)].width = 30 if column_index == 1 else 19
    summary.freeze_panes = "A19"

    additional_measures = mentions.copy()
    if not additional_measures.empty and "status" in additional_measures:
        additional_measures = additional_measures[additional_measures["status"] == "Additional non-GAAP measure"]

    raw_adjustments = analysis.get("adjustments", pd.DataFrame())
    adjustment_history = analysis.get("adjustment_history", pd.DataFrame())
    if adjustment_history.empty and not raw_adjustments.empty:
        adjustment_history = enrich_adjustments(raw_adjustments, reconciliations)
    adjustment_tieouts = analysis.get("adjustment_tieouts", pd.DataFrame())
    if adjustment_tieouts.empty and not reconciliations.empty:
        adjustment_tieouts = build_adjustment_tieouts(reconciliations, adjustment_history)
    adjustment_matrix = make_adjustment_metric_matrix(adjustment_history)
    adjustment_summary = adjustment_category_summary(adjustment_history)

    write_frame("Metrics", reconciliations, "Reconciled non-GAAP metrics extracted from earnings 8-K exhibits")
    write_frame("Adjustment History", adjustment_history, "Issuer adjustment lines normalized by category and fiscal period")
    write_frame("Adjustment Matrix", adjustment_matrix, "Adjustment categories by non-GAAP metric and fiscal period")
    write_frame("Adjustment Summary", adjustment_summary, "Observed adjustment categories and recurrence across selected periods")
    write_frame("Adjustment Tie-Outs", adjustment_tieouts, "Parsed adjustment lines compared with non-GAAP minus GAAP")
    write_frame("Raw Adjustments", raw_adjustments, "Raw GAAP-to-non-GAAP adjustment bridge rows")
    write_frame("Other Measures", additional_measures, "Other non-GAAP measures discussed in matched 8-K exhibits")
    write_frame("Quarter Coverage", coverage, "Periodic filings paired to earnings 8-K filings")
    write_frame("Source Documents", sources, "EX-99 source documents scanned")
    write_frame("Evidence", analysis.get("evidence", pd.DataFrame()), "Detected reconciliation-table evidence")
    write_frame("Warnings", analysis.get("warnings", pd.DataFrame()), "Parsing and source warnings")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
