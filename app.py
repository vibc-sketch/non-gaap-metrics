from __future__ import annotations

import html
import importlib
import io
import os
import re
from datetime import date, datetime, timezone
from typing import Any, Optional

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sec_nongaap as ng


APP_NAME = "SEC Non-GAAP Reconciliation & Peer Benchmarking"
REQUIRED_ENGINE_API = (
    "enrich_adjustments",
    "build_adjustment_tieouts",
    "make_adjustment_metric_matrix",
    "adjustment_category_summary",
    "make_adjustment_value_matrix",
    "make_adjustment_presence_matrix",
    "compare_adjustment_periods",
    "benchmark_metric_family",
    "make_peer_presence_matrix",
    "make_reconciliation_bridge_table",
    "extract_kpi_mentions",
)


def _reload_engine_if_needed() -> tuple[object, list[str]]:
    """Reload the local engine when Streamlit retained an older imported module."""
    module = ng
    missing = [name for name in REQUIRED_ENGINE_API if not callable(getattr(module, name, None))]
    if missing:
        importlib.invalidate_caches()
        module = importlib.reload(module)
        missing = [name for name in REQUIRED_ENGINE_API if not callable(getattr(module, name, None))]
    return module, missing


ng, ENGINE_API_MISSING = _reload_engine_if_needed()
APP_VERSION = getattr(ng, "APP_VERSION", "unknown")

st.set_page_config(
    page_title=APP_NAME,
    page_icon="\U0001F4CA",
    layout="wide",
    initial_sidebar_state="expanded",
)

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@300;400;600;700&display=swap');

:root {
  --dl-green: #86BC25;
  --dl-green-dark: #6B991E;
  --dl-neon-green: #86EB22;
  --dl-blue: #00A3E0;
  --dl-blue-dark: #005587;
  --dl-dark-gray: #282728;
  --dl-text-secondary: #5A5A5A;
  --dl-border: #E2E2E2;
  --dl-bg-page: #FAFAFA;
  --dl-bg-card: #FFFFFF;
  --dl-amber: #E8A317;
  --dl-red: #DA291C;
  --app-border: var(--dl-border);
}

html, body, [class*="css"] {
  font-family: 'Open Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
}

.stApp {
  background: var(--dl-bg-page);
}

.block-container {
  max-width: 1500px;
  padding-top: 0.9rem;
  padding-bottom: 3rem;
}

/* Signature Deloitte green top rule across the whole app */
.block-container::before {
  content: "";
  display: block;
  height: 4px;
  border-radius: 4px;
  margin-bottom: 1.1rem;
  background: linear-gradient(90deg, var(--dl-green), var(--dl-neon-green) 55%, var(--dl-blue));
}

[data-testid="stSidebar"] {
  min-width: 310px;
  background: var(--dl-bg-card);
  border-right: 1px solid var(--dl-border);
}
[data-testid="stSidebar"] > div:first-child {
  border-top: 4px solid var(--dl-green);
}

h1, h2, h3, h4 {
  color: var(--dl-dark-gray);
  font-weight: 700;
}

.app-header {
  display: flex;
  align-items: center;
  gap: 0.9rem;
  margin-bottom: 0.15rem;
}
.app-header .app-mark {
  width: 34px;
  height: 34px;
  border-radius: 50%;
  flex-shrink: 0;
  background: conic-gradient(from 210deg, var(--dl-green), var(--dl-neon-green), var(--dl-blue), var(--dl-green));
  position: relative;
}
.app-header .app-mark::after {
  content: "";
  position: absolute;
  inset: 6px;
  border-radius: 50%;
  background: var(--dl-bg-page);
}
.app-kicker {
  color: var(--dl-green-dark);
  font-size: 0.78rem;
  font-weight: 700;
  letter-spacing: 0.09em;
  text-transform: uppercase;
  margin-bottom: 0.15rem;
}
.app-title {
  font-size: 1.65rem;
  font-weight: 700;
  color: var(--dl-dark-gray);
  line-height: 1.2;
  margin: 0;
}
.app-title .accent {
  font-style: italic;
  color: var(--dl-green-dark);
}

.source-rule {
  border: 1px solid rgba(0, 163, 224, 0.28);
  border-left: 5px solid var(--dl-blue);
  border-radius: 10px;
  padding: 0.9rem 1rem;
  background: rgba(0, 163, 224, 0.06);
  margin: 0.6rem 0 1.15rem 0;
  color: var(--dl-dark-gray);
  font-size: 0.92rem;
}

.info-card {
  border: 1px solid var(--dl-border);
  border-top: 3px solid var(--dl-green);
  border-radius: 10px;
  padding: 0.9rem 1rem;
  min-height: 108px;
  background: var(--dl-bg-card);
  box-shadow: 0 1px 2px rgba(40, 39, 40, 0.04);
}
.info-label {
  font-size: 0.76rem;
  color: var(--dl-text-secondary);
  text-transform: uppercase;
  letter-spacing: 0.045em;
  font-weight: 600;
}
.info-value {
  font-size: 1.25rem;
  font-weight: 700;
  line-height: 1.2;
  margin-top: 0.25rem;
  overflow-wrap: anywhere;
  color: var(--dl-dark-gray);
}
.info-note {
  font-size: 0.78rem;
  color: var(--dl-text-secondary);
  margin-top: 0.35rem;
}

.recon-card {
  border: 1px solid var(--dl-border);
  border-left: 3px solid var(--dl-blue);
  border-radius: 10px;
  padding: 0.8rem 0.9rem;
  min-height: 102px;
  background: var(--dl-bg-card);
  box-shadow: 0 1px 2px rgba(40, 39, 40, 0.04);
}
.recon-card .label {
  font-size: 0.76rem;
  color: var(--dl-text-secondary);
  text-transform: uppercase;
  font-weight: 600;
}
.recon-card .value {
  font-size: 1.14rem;
  font-weight: 700;
  margin-top: 0.25rem;
  color: var(--dl-dark-gray);
}

.small-note {
  font-size: 0.83rem;
  color: var(--dl-text-secondary);
}
.success-note {
  border-left: 4px solid var(--dl-green);
  padding: 0.55rem 0.8rem;
  background: rgba(134, 188, 37, 0.08);
  border-radius: 8px;
  color: var(--dl-dark-gray);
}
.warning-note {
  border-left: 4px solid var(--dl-amber);
  padding: 0.55rem 0.8rem;
  background: rgba(232, 163, 23, 0.08);
  border-radius: 8px;
  color: var(--dl-dark-gray);
}

.bridge-wrap {
  overflow-x: auto;
  margin: 0.6rem 0 1.2rem 0;
  border: 1px solid var(--dl-border);
  border-radius: 10px;
}
.bridge-table {
  width: 100%;
  min-width: 620px;
  border-collapse: collapse;
  font-size: 0.94rem;
  background: var(--dl-bg-card);
}
.bridge-table th, .bridge-table td {
  padding: 0.58rem 0.72rem;
  border-bottom: 1px solid var(--dl-border);
  text-align: right;
  white-space: nowrap;
}
.bridge-table th:first-child, .bridge-table td:first-child {
  text-align: left;
  white-space: normal;
  min-width: 300px;
}
.bridge-table thead tr.company-head th {
  background: var(--dl-dark-gray);
  color: #ffffff;
  font-size: 1.02rem;
  border-bottom: none;
}
.bridge-table thead tr.period-head th {
  background: var(--dl-green);
  color: #ffffff;
  font-weight: 700;
}
.bridge-table tr:nth-child(even) td {
  background: #F5F5F5;
}
.bridge-table tr.row-gaap td, .bridge-table tr.row-non-gaap td {
  background: #EAEAEA;
  font-weight: 700;
  border-top: 2px solid var(--dl-dark-gray);
}
.bridge-table tr.row-adjustment td:first-child {
  padding-left: 1.4rem;
}
.matrix-dot {
  font-size: 1.1rem;
  line-height: 1;
  color: var(--dl-green-dark);
}
.peer-note {
  border-left: 4px solid var(--dl-green);
  background: rgba(134, 188, 37, 0.09);
  border-radius: 8px;
  padding: 0.65rem 0.85rem;
  margin: 0.55rem 0 1rem 0;
  color: var(--dl-dark-gray);
}

/* Tabs: green underline on the active tab instead of Streamlit's red default */
.stTabs [data-baseweb="tab-list"] {
  gap: 4px;
  border-bottom: 1px solid var(--dl-border);
}
.stTabs [aria-selected="true"] {
  color: var(--dl-green-dark) !important;
  font-weight: 700;
}
.stTabs [data-baseweb="tab-highlight"] {
  background-color: var(--dl-green) !important;
}

/* Dataframes / tables */
[data-testid="stDataFrame"] {
  border: 1px solid var(--dl-border);
  border-radius: 8px;
}

hr, [data-testid="stDivider"] {
  border-color: var(--dl-border) !important;
}

@media (max-width: 760px) {
  .block-container {
    padding-left: 0.8rem;
    padding-right: 0.8rem;
    padding-top: 0.8rem;
  }
  h1 { font-size: 1.75rem !important; }
  h2 { font-size: 1.35rem !important; }
  .info-card { min-height: 90px; }
  .app-title { font-size: 1.3rem; }
}
</style>
"""
st.html(CUSTOM_CSS)

if ENGINE_API_MISSING:
    loaded_path = getattr(ng, "__file__", "unknown module path")
    st.error(
        "The deployed app loaded an incompatible copy of sec_nongaap.py. "
        "Upload app.py and sec_nongaap.py from the same release, then reboot the Streamlit app."
    )
    st.code(
        f"Loaded engine: {loaded_path}\n"
        f"Engine version: {APP_VERSION}\n"
        f"Missing functions: {', '.join(ENGINE_API_MISSING)}"
    )
    st.stop()


STATE_DEFAULTS: dict[str, Any] = {
    "issuer_matches": pd.DataFrame(),
    "company": None,
    "submissions": None,
    "filings": pd.DataFrame(),
    "anchors": pd.DataFrame(),
    "analysis": None,
    "analysis_years": [],
    "loaded_cik": None,
    "loaded_contact": "",
    "peer_analysis": None,
    "peer_companies": pd.DataFrame(),
    "peer_errors": [],
    "engine_version": None,
}
for state_key, default_value in STATE_DEFAULTS.items():
    if state_key not in st.session_state:
        st.session_state[state_key] = default_value

# A code upgrade can leave a prior session's analysis object in memory. Clear only
# derived analysis state when the engine version changes; issuer search inputs stay intact.
if st.session_state.get("engine_version") != APP_VERSION:
    st.session_state.analysis = None
    st.session_state.analysis_years = []
    st.session_state.peer_analysis = None
    st.session_state.peer_companies = pd.DataFrame()
    st.session_state.peer_errors = []
    st.session_state.engine_version = APP_VERSION


@st.cache_resource(show_spinner=False)
def get_client(contact_email: str) -> ng.SecClient:
    return ng.SecClient(contact_email=contact_email, app_name=APP_NAME)


def esc(value: Any) -> str:
    return html.escape(str(value if value is not None else ""))


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def format_date(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    try:
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    except Exception:
        return clean_text(value)


def fye_display(value: str) -> str:
    digits = re.sub(r"[^0-9]", "", value or "")
    if len(digits) == 4:
        return f"{digits[:2]}/{digits[2:]}"
    return value or "Not reported"


def period_sort_key(period: str) -> tuple[int, int]:
    match = re.search(r"FY(\d{4})\s+Q([1-4])", period or "")
    if not match:
        return (0, 0)
    return int(match.group(1)), int(match.group(2))


def render_info_card(label: str, value: Any, note: str = "") -> None:
    st.markdown(
        f"""
        <div class="info-card">
          <div class="info-label">{esc(label)}</div>
          <div class="info-value">{esc(value)}</div>
          <div class="info-note">{esc(note)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_recon_card(label: str, value: Any, note: str = "") -> None:
    st.markdown(
        f"""
        <div class="recon-card">
          <div class="label">{esc(label)}</div>
          <div class="value">{esc(value)}</div>
          <div class="info-note">{esc(note)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def reset_loaded_issuer() -> None:
    st.session_state.company = None
    st.session_state.submissions = None
    st.session_state.filings = pd.DataFrame()
    st.session_state.anchors = pd.DataFrame()
    st.session_state.analysis = None
    st.session_state.analysis_years = []
    st.session_state.loaded_cik = None
    st.session_state.loaded_contact = ""
    st.session_state.peer_analysis = None
    st.session_state.peer_companies = pd.DataFrame()
    st.session_state.peer_errors = []


def normalized_value(row: pd.Series) -> Optional[float]:
    try:
        value = float(row.get("non_gaap_value"))
    except Exception:
        return None
    return ng.normalize_value(value, clean_text(row.get("scale")), clean_text(row.get("unit")))


def format_change(current: Optional[float], prior: Optional[float], unit: str) -> str:
    if current is None or prior is None or pd.isna(current) or pd.isna(prior):
        return ""
    delta = float(current) - float(prior)
    if unit == "percent":
        return f"{delta:+,.1f} pp"
    if unit == "bps":
        return f"{delta:+,.0f} bps"
    if float(prior) == 0:
        return "n/m"
    return f"{((float(current) / float(prior)) - 1.0) * 100:+,.1f}%"


def add_change_columns(reconciliations: pd.DataFrame) -> pd.DataFrame:
    if reconciliations.empty:
        return reconciliations.copy()

    data = reconciliations.copy()
    data["quarter_order"] = data["fiscal_quarter"].map(ng.QUARTER_ORDER)
    data["normalized_non_gaap"] = data.apply(normalized_value, axis=1)
    data = data.sort_values(["metric", "fiscal_year", "quarter_order", "confidence"])
    data = data.drop_duplicates(["metric", "fiscal_year", "fiscal_quarter"], keep="first")

    lookup: dict[tuple[str, int, str], pd.Series] = {}
    for _, row in data.iterrows():
        lookup[(str(row["metric"]), int(row["fiscal_year"]), str(row["fiscal_quarter"]))] = row

    qoq_changes: list[str] = []
    yoy_changes: list[str] = []
    for _, row in data.iterrows():
        metric = str(row["metric"])
        fiscal_year = int(row["fiscal_year"])
        quarter = str(row["fiscal_quarter"])
        quarter_number = ng.QUARTER_ORDER.get(quarter, 0)
        if quarter_number > 1:
            previous_q_key = (metric, fiscal_year, f"Q{quarter_number - 1}")
        else:
            previous_q_key = (metric, fiscal_year - 1, "Q4")
        previous_y_key = (metric, fiscal_year - 1, quarter)
        previous_q = lookup.get(previous_q_key)
        previous_y = lookup.get(previous_y_key)
        current_value = row.get("normalized_non_gaap")
        unit = clean_text(row.get("unit"))
        qoq_changes.append(
            format_change(current_value, previous_q.get("normalized_non_gaap") if previous_q is not None else None, unit)
        )
        yoy_changes.append(
            format_change(current_value, previous_y.get("normalized_non_gaap") if previous_y is not None else None, unit)
        )

    data["qoq_change"] = qoq_changes
    data["yoy_change"] = yoy_changes
    return data.drop(columns=["quarter_order"], errors="ignore")


def chart_frame(trends: pd.DataFrame, metric: str) -> tuple[pd.DataFrame, str]:
    selected = trends[trends["metric"].eq(metric)].copy()
    if selected.empty:
        return pd.DataFrame(), ""
    selected["quarter_order"] = selected["fiscal_quarter"].map(ng.QUARTER_ORDER)
    selected = selected.sort_values(["fiscal_year", "quarter_order"])
    unit = clean_text(selected.iloc[0].get("unit"))
    if unit == "usd":
        selected["chart_value"] = selected["normalized_non_gaap"] / 1_000_000.0
        unit_label = "USD millions"
    elif unit == "percent":
        selected["chart_value"] = selected["non_gaap_value"]
        unit_label = "Percent"
    elif unit == "usd_per_share":
        selected["chart_value"] = selected["non_gaap_value"]
        unit_label = "USD per share"
    elif unit == "bps":
        selected["chart_value"] = selected["non_gaap_value"]
        unit_label = "Basis points"
    else:
        selected["chart_value"] = selected["non_gaap_value"]
        unit_label = "Reported value"
    return selected[["period", "chart_value"]].set_index("period"), unit_label


DELOITTE_CHART_SEQUENCE = [
    "#86BC25",  # Deloitte Green (primary, always first)
    "#00A3E0",  # Blue
    "#282728",  # Dark Gray
    "#86EB22",  # Neon Green
    "#A0DCFF",  # Light Blue
    "#005587",  # Blue Dark
    "#B7E320",  # Bright Lime
    "#63C631",  # Green
]


def chart_series_colors(column_count: int) -> list[str]:
    """Deloitte's brand chart color sequence, one color per series, cycled if needed."""
    if column_count <= 0:
        return []
    return [DELOITTE_CHART_SEQUENCE[index % len(DELOITTE_CHART_SEQUENCE)] for index in range(column_count)]


def display_dataframe(
    frame: pd.DataFrame,
    column_config: Optional[dict[str, Any]] = None,
    height: Optional[int] = None,
) -> None:
    if frame.empty:
        st.info("No records are available for this section.")
        return
    dataframe_kwargs: dict[str, Any] = {
        "data": frame,
        "use_container_width": True,
        "hide_index": True,
        "column_config": column_config or {},
    }
    if height is not None:
        dataframe_kwargs["height"] = height
    st.dataframe(**dataframe_kwargs)


def excel_ready(frame: pd.DataFrame) -> pd.DataFrame:
    data = frame.copy()
    for column in data.columns:
        if pd.api.types.is_datetime64_any_dtype(data[column]):
            data[column] = data[column].dt.strftime("%Y-%m-%d")
        elif data[column].dtype == "object":
            data[column] = data[column].map(
                lambda value: format_date(value)
                if isinstance(value, (date, datetime, pd.Timestamp))
                else clean_text(value)
            )
    return data


def style_data_sheet(worksheet: Any, header_row: int = 1) -> None:
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    thin_border = Border(bottom=Side(style="thin", color="9EADBA"))

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = f"A{header_row + 1}"
    max_column = worksheet.max_column
    max_row = worksheet.max_row
    if max_row >= header_row:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(max_column)}{max_row}"

    for cell in worksheet[header_row]:
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    worksheet.row_dimensions[header_row].height = 30

    for column_index in range(1, max_column + 1):
        header = clean_text(worksheet.cell(header_row, column_index).value).lower()
        max_length = len(clean_text(worksheet.cell(header_row, column_index).value))
        for row_index in range(header_row + 1, max_row + 1):
            cell = worksheet.cell(row_index, column_index)
            value = clean_text(cell.value)
            max_length = max(max_length, min(len(value), 80))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = light_fill
            if "url" in header and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 55)


def build_bridge_export_frame(
    reconciliations: pd.DataFrame,
    adjustment_history: pd.DataFrame,
) -> pd.DataFrame:
    """Flatten every reconciliation into presentation-order rows for Excel/CSV review."""
    columns = [
        "Fiscal period",
        "Non-GAAP metric",
        "Row order",
        "Row type",
        "Line item",
        "Reported value",
        "Normalized adjustment category",
        "Source type",
        "SEC source",
    ]
    if reconciliations is None or reconciliations.empty:
        return pd.DataFrame(columns=columns)
    rows: list[dict[str, Any]] = []
    recs = reconciliations.copy()
    recs["_period_rank"] = recs.apply(
        lambda row: ng.fiscal_period_rank(row.get("fiscal_year"), row.get("fiscal_quarter")), axis=1
    )
    recs = recs.sort_values(["_period_rank", "metric"])
    for _, pair in recs.iterrows():
        base = {
            "Fiscal period": clean_text(pair.get("period")),
            "Non-GAAP metric": clean_text(pair.get("metric")),
            "Source type": clean_text(pair.get("source_role")),
            "SEC source": clean_text(pair.get("source_url")),
        }
        rows.append(
            {
                **base,
                "Row order": 0,
                "Row type": "GAAP",
                "Line item": clean_text(pair.get("gaap_label")),
                "Reported value": clean_text(pair.get("gaap_display")),
                "Normalized adjustment category": "",
            }
        )
        pair_id = clean_text(pair.get("pair_id"))
        details = (
            adjustment_history[adjustment_history["pair_id"].astype(str).eq(pair_id)].copy()
            if isinstance(adjustment_history, pd.DataFrame)
            and not adjustment_history.empty
            and "pair_id" in adjustment_history.columns
            else pd.DataFrame()
        )
        if not details.empty:
            sort_columns = [column for column in ["adjustment_order", "adjustment_label"] if column in details.columns]
            if sort_columns:
                details = details.sort_values(sort_columns)
            for offset, (_, detail) in enumerate(details.iterrows(), start=1):
                rows.append(
                    {
                        **base,
                        "Row order": offset,
                        "Row type": "Adjustment",
                        "Line item": clean_text(detail.get("adjustment_label")),
                        "Reported value": clean_text(detail.get("adjustment_display")),
                        "Normalized adjustment category": clean_text(detail.get("adjustment_category")),
                    }
                )
        else:
            rows.append(
                {
                    **base,
                    "Row order": 1,
                    "Row type": "Adjustment total — detail not parsed",
                    "Line item": "Total adjustments",
                    "Reported value": clean_text(pair.get("adjustment_display")),
                    "Normalized adjustment category": "",
                }
            )
        rows.append(
            {
                **base,
                "Row order": 999,
                "Row type": "Non-GAAP",
                "Line item": clean_text(pair.get("non_gaap_label")),
                "Reported value": clean_text(pair.get("non_gaap_display")),
                "Normalized adjustment category": "",
            }
        )
    return pd.DataFrame(rows, columns=columns)


def build_excel_export(
    company: dict[str, Any],
    selected_years: list[int],
    analysis: dict[str, pd.DataFrame],
    matrix: pd.DataFrame,
    trends: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    additional = analysis.get("mentions", pd.DataFrame())
    if not additional.empty and "status" in additional.columns:
        additional = additional[additional["status"].eq("Additional non-GAAP measure")].copy()
    raw_adjustments = analysis.get("adjustments", pd.DataFrame())
    reconciliations = analysis.get("reconciliations", pd.DataFrame())
    adjustment_history = analysis.get("adjustment_history", pd.DataFrame())
    if adjustment_history.empty and not raw_adjustments.empty:
        adjustment_history = ng.enrich_adjustments(raw_adjustments, reconciliations)
    adjustment_tieouts = analysis.get("adjustment_tieouts", pd.DataFrame())
    if adjustment_tieouts.empty and not reconciliations.empty:
        adjustment_tieouts = ng.build_adjustment_tieouts(reconciliations, adjustment_history)
    adjustment_matrix = ng.make_adjustment_metric_matrix(adjustment_history)
    adjustment_summary = ng.adjustment_category_summary(adjustment_history)
    presentation_bridges = build_bridge_export_frame(reconciliations, adjustment_history)

    summary = pd.DataFrame(
        {
            "Field": [
                "Company",
                "Ticker",
                "CIK",
                "SIC",
                "Industry",
                "Fiscal year end",
                "Fiscal years analyzed",
                "Source rule",
                "Generated at UTC",
                "App version",
            ],
            "Value": [
                company.get("name", ""),
                company.get("ticker", ""),
                company.get("cik", ""),
                company.get("sic", ""),
                company.get("sic_description", ""),
                company.get("fiscal_year_end", ""),
                ", ".join(str(year) for year in sorted(selected_years)),
                "Metrics come only from matched earnings 8-K EX-99 exhibits; 10-Q/10-K filings supply fiscal-period metadata.",
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
                APP_VERSION,
            ],
        }
    )

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("Summary", summary),
        ("Metric matrix", matrix),
        ("Trend analysis", trends),
        ("Reconciliations", reconciliations),
        ("Presentation bridges", presentation_bridges),
        ("Adjustment history", adjustment_history),
        ("Adjustment matrix", adjustment_matrix),
        ("Adjustment summary", adjustment_summary),
        ("Adjustment tie-outs", adjustment_tieouts),
        ("Raw adjustments", raw_adjustments),
        ("Additional measures", additional),
        ("KPIs in 8-K package", analysis.get("kpis", pd.DataFrame())),
        ("Coverage", analysis.get("coverage", pd.DataFrame())),
        ("Source audit", analysis.get("sources", pd.DataFrame())),
        ("Evidence", analysis.get("evidence", pd.DataFrame())),
        ("Warnings", analysis.get("warnings", pd.DataFrame())),
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            safe_frame = excel_ready(frame) if isinstance(frame, pd.DataFrame) else pd.DataFrame()
            if safe_frame.empty:
                safe_frame = pd.DataFrame({"Message": ["No records extracted for this section."]})
            safe_frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        workbook = writer.book
        for worksheet in workbook.worksheets:
            style_data_sheet(worksheet)
        summary_sheet = workbook["Summary"]
        summary_sheet["A1"].font = Font(bold=True)
        summary_sheet["B1"].font = Font(bold=True)

    return output.getvalue()


def compact_reconciliation_view(trends: pd.DataFrame) -> pd.DataFrame:
    if trends.empty:
        return pd.DataFrame()
    columns = [
        "period",
        "metric",
        "gaap_display",
        "adjustment_display",
        "non_gaap_display",
        "qoq_change",
        "yoy_change",
        "source_role",
        "confidence",
        "source_url",
    ]
    available = [column for column in columns if column in trends.columns]
    view = trends[available].copy()
    return view.rename(
        columns={
            "period": "Fiscal period",
            "metric": "Non-GAAP metric",
            "gaap_display": "GAAP",
            "adjustment_display": "Total adjustments",
            "non_gaap_display": "Non-GAAP",
            "qoq_change": "QoQ change",
            "yoy_change": "YoY change",
            "source_role": "Source type",
            "confidence": "Parse confidence",
            "source_url": "SEC source",
        }
    )



def adjustment_chart_frame(adjustment_history: pd.DataFrame, metric: str) -> tuple[pd.DataFrame, str]:
    if adjustment_history.empty:
        return pd.DataFrame(), ""
    data = adjustment_history[adjustment_history["metric"].astype(str).eq(str(metric))].copy()
    if data.empty:
        return pd.DataFrame(), ""
    unit = clean_text(data["unit"].mode().iloc[0]) if "unit" in data.columns and not data["unit"].mode().empty else "number"
    data = data[data["unit"].astype(str).eq(unit)].copy()
    if unit == "usd":
        data["chart_value"] = pd.to_numeric(data["normalized_adjustment_value"], errors="coerce") / 1_000_000.0
        unit_label = "USD millions"
    elif unit == "percent":
        data["chart_value"] = pd.to_numeric(data["adjustment_value"], errors="coerce")
        unit_label = "Percentage points"
    elif unit == "usd_per_share":
        data["chart_value"] = pd.to_numeric(data["adjustment_value"], errors="coerce")
        unit_label = "USD per share"
    elif unit == "bps":
        data["chart_value"] = pd.to_numeric(data["adjustment_value"], errors="coerce")
        unit_label = "Basis points"
    else:
        data["chart_value"] = pd.to_numeric(data["adjustment_value"], errors="coerce")
        unit_label = "Reported value"
    data = data.dropna(subset=["chart_value"])
    if data.empty:
        return pd.DataFrame(), unit_label
    top_categories = (
        data.groupby("adjustment_category")["chart_value"]
        .apply(lambda values: values.abs().sum())
        .sort_values(ascending=False)
        .head(8)
        .index
        .tolist()
    )
    data = data[data["adjustment_category"].isin(top_categories)]
    chart = data.pivot_table(
        index="period",
        columns="adjustment_category",
        values="chart_value",
        aggfunc="sum",
        fill_value=0,
    )
    ordered = [period for period in ng.ordered_fiscal_periods(data) if period in chart.index]
    return chart.reindex(ordered), unit_label


def compact_adjustment_history_view(adjustment_history: pd.DataFrame) -> pd.DataFrame:
    if adjustment_history.empty:
        return pd.DataFrame()
    columns = [
        "period",
        "metric",
        "adjustment_category",
        "adjustment_label",
        "adjustment_display",
        "effect_on_non_gaap",
        "observed_frequency",
        "period_lifecycle",
        "source_role",
        "source_page",
        "source_url",
    ]
    available = [column for column in columns if column in adjustment_history.columns]
    return adjustment_history[available].copy().rename(
        columns={
            "period": "Fiscal period",
            "metric": "Non-GAAP metric",
            "adjustment_category": "Normalized category",
            "adjustment_label": "Issuer-reported adjustment",
            "adjustment_display": "Reported value",
            "effect_on_non_gaap": "Effect on non-GAAP",
            "observed_frequency": "Observed frequency",
            "period_lifecycle": "Period status",
            "source_role": "Source type",
            "source_page": "PDF page",
            "source_url": "SEC source",
        }
    )


def tieout_view(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    columns = [
        "period",
        "metric",
        "gaap_display",
        "parsed_adjustment_display",
        "expected_adjustment_display",
        "variance_display",
        "non_gaap_display",
        "detail_line_count",
        "category_count",
        "tie_out_status",
        "tie_out_note",
        "source_url",
    ]
    available = [column for column in columns if column in frame.columns]
    return frame[available].copy().rename(
        columns={
            "period": "Fiscal period",
            "metric": "Non-GAAP metric",
            "gaap_display": "GAAP",
            "parsed_adjustment_display": "Parsed line items",
            "expected_adjustment_display": "Non-GAAP minus GAAP",
            "variance_display": "Difference",
            "non_gaap_display": "Non-GAAP",
            "detail_line_count": "Detail lines",
            "category_count": "Categories",
            "tie_out_status": "Tie-out status",
            "tie_out_note": "Review note",
            "source_url": "SEC source",
        }
    )


def coverage_view(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    columns = [
        "period",
        "period_end",
        "periodic_form",
        "periodic_filing_date",
        "earnings_8k_filing_date",
        "source_documents",
        "reconciled_metrics",
        "additional_measures",
        "status",
        "earnings_8k_url",
        "periodic_url",
    ]
    available = [column for column in columns if column in frame.columns]
    view = frame[available].copy()
    for column in ["period_end", "periodic_filing_date", "earnings_8k_filing_date"]:
        if column in view.columns:
            view[column] = view[column].map(format_date)
    return view.rename(
        columns={
            "period": "Fiscal period",
            "period_end": "Period end",
            "periodic_form": "Anchor filing",
            "periodic_filing_date": "10-Q/10-K filed",
            "earnings_8k_filing_date": "Earnings 8-K filed",
            "source_documents": "EX-99 documents checked",
            "reconciled_metrics": "Metrics extracted",
            "additional_measures": "Other measures",
            "status": "Status",
            "earnings_8k_url": "Earnings 8-K",
            "periodic_url": "Anchor 10-Q/10-K",
        }
    )


def source_audit_view(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    columns = [
        "period",
        "document_role",
        "content_type",
        "has_reconciliation",
        "document",
        "description",
        "8k_filing_date",
        "document_url",
        "8k_index_url",
    ]
    available = [column for column in columns if column in frame.columns]
    view = frame[available].copy()
    if "8k_filing_date" in view.columns:
        view["8k_filing_date"] = view["8k_filing_date"].map(format_date)
    if "has_reconciliation" in view.columns:
        view["has_reconciliation"] = view["has_reconciliation"].map(lambda value: "Yes" if bool(value) else "No")
    return view.rename(
        columns={
            "period": "Fiscal period",
            "document_role": "Document role",
            "content_type": "Format",
            "has_reconciliation": "Reconciliation found",
            "document": "Document",
            "description": "Description",
            "8k_filing_date": "8-K filed",
            "document_url": "Exhibit source",
            "8k_index_url": "8-K filing index",
        }
    )


def render_bridge_table(
    frame: pd.DataFrame,
    title: str,
    subtitle: str = "",
    source_url: str = "",
) -> None:
    """Render a presentation-style GAAP -> adjustments -> non-GAAP table."""
    if frame is None or frame.empty:
        st.info("No structured bridge rows were available for this selection.")
        return
    period_columns = [column for column in frame.columns if column not in {"Line item", "Row type"}]
    header_cells = "".join(f"<th>{esc(column)}</th>" for column in period_columns)
    body_rows: list[str] = []
    for _, row in frame.iterrows():
        row_type = clean_text(row.get("Row type")).lower().replace(" ", "-")
        row_class = {
            "gaap": "row-gaap",
            "non-gaap": "row-non-gaap",
            "adjustment": "row-adjustment",
        }.get(row_type, "")
        values = "".join(f"<td>{esc(row.get(column, '—'))}</td>" for column in period_columns)
        body_rows.append(
            f'<tr class="{row_class}"><td>{esc(row.get("Line item", ""))}</td>{values}</tr>'
        )
    subtitle_html = f'<div class="small-note" style="margin:0.2rem 0 0.45rem 0;">{esc(subtitle)}</div>' if subtitle else ""
    link_html = f'<div class="small-note" style="margin-top:0.45rem;"><a href="{esc(source_url)}" target="_blank">Open SEC source exhibit</a></div>' if source_url else ""
    st.markdown(
        f"""
        {subtitle_html}
        <div class="bridge-wrap">
          <table class="bridge-table">
            <thead>
              <tr class="company-head"><th colspan="{len(period_columns) + 1}">{esc(title)}</th></tr>
              <tr class="period-head"><th>Reconciliation line item</th>{header_cells}</tr>
            </thead>
            <tbody>{''.join(body_rows)}</tbody>
          </table>
        </div>
        {link_html}
        """,
        unsafe_allow_html=True,
    )


def _first_ticker(company_record: dict[str, Any]) -> str:
    return clean_text(company_record.get("ticker")).split(",")[0].strip().upper()


def _append_company_columns(frame: pd.DataFrame, company_record: dict[str, Any]) -> pd.DataFrame:
    data = frame.copy() if isinstance(frame, pd.DataFrame) else pd.DataFrame()
    if data.empty:
        return data
    label = _first_ticker(company_record) or clean_text(company_record.get("name"))
    data.insert(0, "company", label)
    data.insert(1, "company_name", clean_text(company_record.get("name")))
    data.insert(2, "cik", company_record.get("cik"))
    return data


def combine_peer_results(results: list[tuple[dict[str, Any], dict[str, pd.DataFrame]]]) -> dict[str, pd.DataFrame]:
    keys = {
        "coverage",
        "reconciliations",
        "adjustments",
        "adjustment_history",
        "adjustment_tieouts",
        "mentions",
        "kpis",
        "sources",
        "evidence",
        "warnings",
    }
    combined: dict[str, pd.DataFrame] = {}
    for key in keys:
        frames = [
            _append_company_columns(analysis.get(key, pd.DataFrame()), company_record)
            for company_record, analysis in results
        ]
        frames = [frame for frame in frames if not frame.empty]
        combined[key] = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    return combined


def resolve_exact_ticker(client: ng.SecClient, ticker: str) -> Optional[dict[str, Any]]:
    ticker = clean_text(ticker).upper()
    if not ticker:
        return None
    matches = ng.search_companies(client, query=ticker, limit=30)
    if matches.empty:
        return None
    exact = matches[matches["ticker"].astype(str).str.upper().eq(ticker)]
    row = exact.iloc[0] if not exact.empty else matches.iloc[0]
    return {"cik": int(row["cik"]), "ticker": clean_text(row.get("ticker")), "name": clean_text(row.get("name"))}


def analyze_peer_company(
    client: ng.SecClient,
    issuer: dict[str, Any],
    progress: Optional[Any] = None,
    max_exhibits: int = 6,
) -> tuple[dict[str, Any], dict[str, pd.DataFrame]]:
    cik = int(issuer["cik"])
    submissions = ng.load_company_submissions(client, cik)
    company_record = ng.company_record(submissions)
    filings = ng.load_all_filings(client, submissions, years_back=5, max_history_files=5)
    anchors = ng.build_period_anchors(
        client,
        cik,
        filings,
        company_record.get("fiscal_year_end", ""),
        max_periodic_filings=20,
        progress=progress,
    )
    if anchors.empty:
        raise ValueError("No recent 10-Q/10-K fiscal-period anchors were found.")
    years = sorted({int(value) for value in anchors["fiscal_year"].dropna().tolist()}, reverse=True)[:2]
    analysis = ng.analyze_company_quarters(
        client,
        cik,
        filings,
        anchors,
        years,
        progress=progress,
        max_exhibits_per_8k=max_exhibits,
    )
    return company_record, analysis


def build_measure_presence_source(peer_analysis: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    reconciliations = peer_analysis.get("reconciliations", pd.DataFrame())
    mentions = peer_analysis.get("mentions", pd.DataFrame())
    if not reconciliations.empty:
        recs = reconciliations.copy()
        if "metric_family" not in recs.columns:
            recs["metric_family"] = recs.apply(
                lambda row: ng.benchmark_metric_family(
                    row.get("metric", ""), row.get("gaap_label", ""), row.get("non_gaap_label", "")
                ),
                axis=1,
            )
        frames.append(recs[["company", "metric_family"]])
    if not mentions.empty:
        mention_data = mentions.copy()
        if "metric_family" not in mention_data.columns:
            mention_data["metric_family"] = mention_data["metric"].map(ng.benchmark_metric_family)
        frames.append(mention_data[["company", "metric_family"]])
    return pd.concat(frames, ignore_index=True).drop_duplicates() if frames else pd.DataFrame(columns=["company", "metric_family"])


def peer_matrix_highlight(matrix: pd.DataFrame, noun: str) -> str:
    if matrix is None or matrix.empty:
        return ""
    company_columns = [column for column in matrix.columns if column not in {"Disclosure", "Total"}]
    common_row = matrix.sort_values("Total", ascending=False).iloc[0]
    counts = {column: int(matrix[column].astype(str).eq("●").sum()) for column in company_columns}
    broadest = max(counts, key=counts.get) if counts else ""
    return (
        f"Most common {noun}: **{common_row['Disclosure']}** ({int(common_row['Total'])} peers). "
        f"Broadest coverage in this selected set: **{broadest}** ({counts.get(broadest, 0)} rows)."
    )


def peer_company_count_frame(matrix: pd.DataFrame, column_label: str) -> pd.DataFrame:
    """Count disclosure rows by peer, matching the bar charts in the benchmark deck."""
    if matrix is None or matrix.empty:
        return pd.DataFrame()
    company_columns = [column for column in matrix.columns if column not in {"Disclosure", "Total"}]
    rows = [
        {"Peer": company, column_label: int(matrix[company].astype(str).eq("●").sum())}
        for company in company_columns
    ]
    return pd.DataFrame(rows).sort_values(column_label, ascending=True).set_index("Peer")


st.markdown(
    f"""
    <div class="app-kicker">Evidence-first SEC filing analysis</div>
    <div class="app-header">
      <div class="app-mark"></div>
      <h1 class="app-title">SEC Non-GAAP <span class="accent">Reconciliation</span> &amp; Peer Benchmarking</h1>
    </div>
    """,
    unsafe_allow_html=True,
)
st.caption(
    "Structured GAAP-to-non-GAAP reconciliations, adjustment bridges, and additional non-GAAP callouts by issuer fiscal quarter."
)
st.markdown(
    """
    <div class="source-rule">
      <strong>Source rule:</strong> Reported metrics are extracted only from the matched earnings Form 8-K and its EX-99 exhibits,
      including HTML press releases, financial supplements, and PDF investor presentations. Forms 10-Q and 10-K are used only
      to establish the issuer fiscal year, fiscal quarter, and the earnings-event matching window.
    </div>
    """,
    unsafe_allow_html=True,
)


with st.sidebar:
    st.header("1. Search issuer")
    default_contact = os.getenv("SEC_CONTACT_EMAIL", "")
    contact_email = st.text_input(
        "SEC contact email",
        value=default_contact,
        placeholder="name@company.com",
        help="The SEC asks automated users to identify themselves in the User-Agent header. The app does not save this email.",
    )

    with st.form("issuer_search_form", clear_on_submit=False):
        query = st.text_input("Company name or ticker", placeholder="LSCC or Lattice Semiconductor")
        sic = st.text_input("SIC code (optional)", placeholder="3674")
        industry = st.text_input("Industry keyword (optional)", placeholder="Semiconductors")
        search_clicked = st.form_submit_button("Find issuers", use_container_width=True)

    if search_clicked:
        if not ng.SecClient.valid_contact(contact_email):
            st.error("Enter a valid contact email before querying SEC data.")
        elif not any(clean_text(value) for value in [query, sic, industry]):
            st.error("Enter a company/ticker, SIC code, or industry keyword.")
        else:
            try:
                with st.spinner("Searching SEC issuer data..."):
                    client = get_client(contact_email.strip())
                    matches = ng.search_companies(
                        client,
                        query=query,
                        sic=sic,
                        industry=industry,
                        limit=100,
                    )
                st.session_state.issuer_matches = matches
                reset_loaded_issuer()
                if matches.empty:
                    st.warning("No issuer matches were found. Try a ticker, a broader company name, or a different industry keyword.")
                else:
                    st.success(f"Found {len(matches):,} issuer match(es).")
            except Exception as exc:
                st.error(f"Issuer search failed: {exc}")

    matches = st.session_state.issuer_matches
    selected_cik: Optional[int] = None
    if isinstance(matches, pd.DataFrame) and not matches.empty:
        st.divider()
        st.header("2. Select issuer")
        option_ciks = [int(value) for value in matches["cik"].tolist()]
        lookup = matches.set_index("cik").to_dict("index")

        def issuer_label(cik_value: int) -> str:
            item = lookup.get(cik_value, {})
            return f"{item.get('name', '')} ({item.get('ticker', '')}) - CIK {cik_value}"

        selected_cik = st.selectbox("Issuer matches", options=option_ciks, format_func=issuer_label)
        load_clicked = st.button("Load fiscal periods", use_container_width=True, type="primary")
        if load_clicked:
            if not ng.SecClient.valid_contact(contact_email):
                st.error("Enter a valid contact email before loading filings.")
            else:
                try:
                    status = st.status("Loading issuer and fiscal-period metadata...", expanded=True)
                    client = get_client(contact_email.strip())
                    submissions = ng.load_company_submissions(client, int(selected_cik))
                    status.update(label="Loading recent filing history...", state="running")
                    filings = ng.load_all_filings(client, submissions, years_back=5, max_history_files=5)

                    def anchor_progress(message: str) -> None:
                        status.update(label=message, state="running")

                    company = ng.company_record(submissions)
                    anchors = ng.build_period_anchors(
                        client,
                        int(selected_cik),
                        filings,
                        company.get("fiscal_year_end", ""),
                        max_periodic_filings=20,
                        progress=anchor_progress,
                    )
                    st.session_state.company = company
                    st.session_state.submissions = submissions
                    st.session_state.filings = filings
                    st.session_state.anchors = anchors
                    st.session_state.analysis = None
                    st.session_state.analysis_years = []
                    st.session_state.loaded_cik = int(selected_cik)
                    st.session_state.loaded_contact = contact_email.strip()
                    status.update(label="Issuer loaded.", state="complete", expanded=False)
                    if anchors.empty:
                        st.warning("No recent 10-Q/10-K fiscal-period anchors were found for this issuer.")
                except Exception as exc:
                    st.error(f"Could not load issuer filings: {exc}")

    if st.session_state.company:
        st.divider()
        st.header("3. Analyze")
        anchors = st.session_state.anchors
        available_years = (
            sorted({int(value) for value in anchors.get("fiscal_year", pd.Series(dtype=int)).dropna().tolist()}, reverse=True)
            if isinstance(anchors, pd.DataFrame) and not anchors.empty
            else []
        )
        default_years = available_years[:2]
        selected_years = st.multiselect(
            "Fiscal years (maximum two)",
            options=available_years,
            default=default_years,
            max_selections=2,
            help="The latest two fiscal years are selected by default. Current fiscal years may contain fewer than four quarters.",
        )
        max_exhibits = st.slider(
            "Maximum EX-99 exhibits per earnings 8-K",
            min_value=3,
            max_value=12,
            value=8,
            help="Press releases, financial supplements, investor presentations, and other EX-99 exhibits are considered.",
        )
        analyze_clicked = st.button("Run 8-K reconciliation analysis", use_container_width=True, type="primary")
        if analyze_clicked:
            if not selected_years:
                st.error("Select at least one fiscal year.")
            elif len(selected_years) > 2:
                st.error("Select no more than two fiscal years.")
            elif contact_email.strip() != st.session_state.loaded_contact:
                st.error("The SEC contact email changed. Reload the issuer so all requests use the same contact identity.")
            else:
                try:
                    status = st.status("Starting 8-K analysis...", expanded=True)
                    client = get_client(contact_email.strip())

                    def analysis_progress(message: str) -> None:
                        status.update(label=message, state="running")

                    analysis = ng.analyze_company_quarters(
                        client,
                        int(st.session_state.loaded_cik),
                        st.session_state.filings,
                        st.session_state.anchors,
                        [int(year) for year in selected_years],
                        progress=analysis_progress,
                        max_exhibits_per_8k=int(max_exhibits),
                    )
                    st.session_state.analysis = analysis
                    st.session_state.analysis_years = [int(year) for year in selected_years]
                    st.session_state.peer_analysis = None
                    st.session_state.peer_companies = pd.DataFrame()
                    st.session_state.peer_errors = []
                    status.update(label="Analysis complete.", state="complete", expanded=False)
                except Exception as exc:
                    st.error(f"The analysis could not be completed: {exc}")

    st.divider()
    st.caption(f"Version {APP_VERSION}. SEC filing data is public. This tool is for research and review, not investment advice.")


company = st.session_state.company
if not company:
    st.subheader("Start with an issuer search")
    st.write(
        "Enter a ticker such as **LSCC**, a company name, a four-digit SIC code, or an industry keyword in the left sidebar. "
        "After selecting an issuer, the app will identify fiscal quarters from its 10-Q/10-K metadata and then analyze only the matched earnings 8-K exhibits."
    )
    st.markdown(
        """
        **The results are organized into:**

        - a two-year quarterly metric matrix;
        - structured GAAP, adjustment, and non-GAAP values;
        - quarter-over-quarter and year-over-year movement;
        - additional non-GAAP measures discussed in the release, supplement, or investor presentation; and
        - an audit trail linking every result to the SEC source document.
        """
    )
    st.stop()


st.subheader("Issuer profile")
profile_columns = st.columns(5)
with profile_columns[0]:
    render_info_card("Company", company.get("name", ""), company.get("exchange", ""))
with profile_columns[1]:
    render_info_card("Ticker", company.get("ticker", ""), f"CIK {company.get('cik', '')}")
with profile_columns[2]:
    render_info_card("SIC", company.get("sic", ""), company.get("sic_description", ""))
with profile_columns[3]:
    render_info_card("Fiscal year end", fye_display(company.get("fiscal_year_end", "")), "Inline XBRL controls period labels when available")
with profile_columns[4]:
    anchor_count = len(st.session_state.anchors) if isinstance(st.session_state.anchors, pd.DataFrame) else 0
    render_info_card("Fiscal periods found", anchor_count, "10-Q/10-K metadata anchors")

anchors = st.session_state.anchors
if isinstance(anchors, pd.DataFrame) and not anchors.empty:
    with st.expander("Review fiscal-period normalization", expanded=False):
        anchor_view = anchors.copy()
        anchor_view["period"] = anchor_view.apply(
            lambda row: f"FY{int(row['fiscal_year'])} {row['fiscal_quarter']}", axis=1
        )
        for column in ["period_end", "periodic_filing_date"]:
            anchor_view[column] = anchor_view[column].map(format_date)
        anchor_view = anchor_view[
            [
                "period",
                "period_end",
                "periodic_form",
                "periodic_filing_date",
                "metadata_source",
                "periodic_url",
            ]
        ].rename(
            columns={
                "period": "Fiscal period",
                "period_end": "Period end",
                "periodic_form": "Anchor form",
                "periodic_filing_date": "Filed",
                "metadata_source": "Normalization source",
                "periodic_url": "SEC anchor filing",
            }
        )
        display_dataframe(
            anchor_view,
            column_config={
                "SEC anchor filing": st.column_config.LinkColumn("SEC anchor filing", display_text="Open filing")
            },
        )
        st.caption("No metric values are taken from these 10-Q/10-K filings; they are used only for fiscal-period normalization and event matching.")


analysis = st.session_state.analysis
if not analysis:
    st.info("Choose up to two fiscal years in the sidebar and run the 8-K reconciliation analysis.")
    st.stop()

coverage = analysis.get("coverage", pd.DataFrame())
reconciliations = analysis.get("reconciliations", pd.DataFrame())
adjustments = analysis.get("adjustments", pd.DataFrame())
adjustment_history = analysis.get("adjustment_history", pd.DataFrame())
if adjustment_history.empty and not adjustments.empty:
    adjustment_history = ng.enrich_adjustments(adjustments, reconciliations)
adjustment_tieouts = analysis.get("adjustment_tieouts", pd.DataFrame())
if adjustment_tieouts.empty and not reconciliations.empty:
    adjustment_tieouts = ng.build_adjustment_tieouts(reconciliations, adjustment_history)
mentions = analysis.get("mentions", pd.DataFrame())
kpis = analysis.get("kpis", pd.DataFrame())
sources = analysis.get("sources", pd.DataFrame())
evidence = analysis.get("evidence", pd.DataFrame())
warnings = analysis.get("warnings", pd.DataFrame())
trends = add_change_columns(reconciliations)
matrix = ng.make_metric_matrix(reconciliations, include_gaap=False)
additional = (
    mentions[mentions["status"].eq("Additional non-GAAP measure")].copy()
    if not mentions.empty and "status" in mentions.columns
    else pd.DataFrame()
)

st.divider()
st.subheader("Analysis summary")
periods_analyzed = len(coverage)
periods_complete = int(coverage["status"].eq("Reconciliation metrics extracted").sum()) if not coverage.empty else 0
unique_metrics = int(reconciliations["metric"].nunique()) if not reconciliations.empty else 0
unique_adjustment_categories = (
    int(adjustment_history["adjustment_category"].nunique())
    if not adjustment_history.empty and "adjustment_category" in adjustment_history.columns
    else 0
)
repeated_adjustment_categories = (
    int(
        ng.adjustment_category_summary(adjustment_history)["periods_observed"].ge(2).sum()
    )
    if not adjustment_history.empty
    else 0
)
presentation_docs = (
    int(sources[sources["document_role"].eq("Investor presentation")]["document_url"].nunique())
    if not sources.empty and {"document_role", "document_url"}.issubset(sources.columns)
    else 0
)
summary_columns = st.columns(4)
with summary_columns[0]:
    render_info_card("Fiscal periods", periods_analyzed, f"{periods_complete} with parsed reconciliations")
with summary_columns[1]:
    render_info_card("Structured metrics", len(reconciliations), f"{unique_metrics} unique metric names")
with summary_columns[2]:
    render_info_card(
        "Adjustment rows",
        len(adjustment_history),
        f"{unique_adjustment_categories} categories; {repeated_adjustment_categories} repeated across periods",
    )
with summary_columns[3]:
    render_info_card("Investor presentations checked", presentation_docs, f"{len(additional)} additional measure callouts")

if reconciliations.empty:
    st.warning(
        "The matched 8-K exhibits were checked, but no GAAP-to-non-GAAP table was parsed into structured values. "
        "Review the Source audit tab for the selected 8-K, PDFs, parsing warnings, and direct SEC links."
    )
else:
    st.markdown(
        '<div class="success-note"><strong>Start with Presentation bridges.</strong> '
        'It shows the comparable GAAP line, every parsed issuer-reported adjustment, and the non-GAAP endpoint in fiscal-period columns. '
        'Quarterly metrics and Adjustment history provide the trend and audit views.</div>',
        unsafe_allow_html=True,
    )

export_payload = dict(analysis)
export_payload["metric_matrix"] = matrix
export_payload["trend_analysis"] = trends
export_payload["adjustment_history"] = adjustment_history
export_payload["adjustment_category_matrix"] = ng.make_adjustment_metric_matrix(adjustment_history)
export_payload["adjustment_category_summary"] = ng.adjustment_category_summary(adjustment_history)
export_payload["adjustment_tieouts"] = adjustment_tieouts
excel_bytes = build_excel_export(company, st.session_state.analysis_years, analysis, matrix, trends)
csv_zip_bytes = ng.build_export_zip(export_payload)

download_columns = st.columns([1, 1, 2])
with download_columns[0]:
    st.download_button(
        "Download formatted Excel",
        data=excel_bytes,
        file_name=f"{clean_text(company.get('ticker')) or company.get('cik')}_non_gaap_8k_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with download_columns[1]:
    st.download_button(
        "Download CSV package",
        data=csv_zip_bytes,
        file_name=f"{clean_text(company.get('ticker')) or company.get('cik')}_non_gaap_8k_csv.zip",
        mime="application/zip",
        use_container_width=True,
    )
with download_columns[2]:
    st.caption("Exports preserve fiscal-period labels, source URLs, parsing evidence, and warnings for review.")


tab_bridge, tab_metrics, tab_details, tab_adjustments, tab_additional, tab_peer, tab_sources = st.tabs(
    [
        "Presentation bridges",
        "Quarterly metrics",
        "Reconciliation detail",
        "Adjustment history",
        "Additional measures",
        "Peer benchmark",
        "Source audit",
    ]
)

with tab_bridge:
    st.subheader("Presentation-style reconciliation bridges")
    st.write(
        "This view reproduces the core structure used in benchmarking presentations: the comparable GAAP result, "
        "the issuer's individual reconciling items in reported order, and the non-GAAP result across fiscal periods."
    )
    if reconciliations.empty:
        st.info("No structured reconciliation pairs were extracted. Use Source audit to inspect the matched exhibit and parsing evidence.")
    else:
        bridge_metric_options = (
            reconciliations.groupby("metric")
            .agg(periods=("period", "nunique"), rows=("pair_id", "size"))
            .sort_values(["periods", "rows"], ascending=False)
            .index.astype(str).tolist()
        )
        preferred_index = next(
            (index for index, value in enumerate(bridge_metric_options) if "gross margin" in value.lower() or "gross profit" in value.lower()),
            0,
        )
        selected_bridge_metric = st.selectbox(
            "Non-GAAP measure",
            options=bridge_metric_options,
            index=preferred_index,
            key="presentation_bridge_metric",
        )
        bridge_metric_rows = reconciliations[
            reconciliations["metric"].astype(str).eq(selected_bridge_metric)
        ].copy()
        bridge_period_options = ng.ordered_fiscal_periods(bridge_metric_rows)
        title_label = clean_text(company.get("name")) or clean_text(company.get("ticker"))
        selected_bridge_periods = st.multiselect(
            "Fiscal-period columns (maximum two)",
            options=bridge_period_options,
            default=bridge_period_options[-2:],
            max_selections=2,
            key="presentation_bridge_periods",
        )
        if not selected_bridge_periods:
            st.info("Select one or two fiscal periods.")
        else:
            bridge_frame = ng.make_reconciliation_bridge_table(
                reconciliations,
                adjustment_history,
                selected_bridge_metric,
                selected_bridge_periods,
            )
            source_rows_for_bridge = bridge_metric_rows[
                bridge_metric_rows["period"].astype(str).isin(selected_bridge_periods)
            ].sort_values(["fiscal_year", "fiscal_quarter"], ascending=[False, False])
            bridge_source = (
                clean_text(source_rows_for_bridge.iloc[0].get("source_url"))
                if not source_rows_for_bridge.empty
                else ""
            )
            render_bridge_table(
                bridge_frame,
                f"{title_label} — {selected_bridge_metric}",
                "Values and line-item labels are preserved from the matched earnings 8-K exhibit; columns follow issuer fiscal periods.",
                bridge_source,
            )
            if bridge_frame["Line item"].astype(str).str.contains("individual items not parsed", case=False).any():
                st.warning(
                    "At least one period has a GAAP-to-non-GAAP endpoint difference but no parsed line-item detail. "
                    "Open the linked exhibit and review Source audit; the app marks the gap rather than hiding it."
                )
            period_links = source_rows_for_bridge[["period", "source_role", "source_url"]].drop_duplicates()
            if not period_links.empty:
                st.markdown("**Source exhibits by period**")
                link_view = period_links.rename(
                    columns={"period": "Fiscal period", "source_role": "Document role", "source_url": "SEC source"}
                )
                display_dataframe(
                    link_view,
                    column_config={"SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")},
                    height=min(280, 90 + 36 * len(link_view)),
                )

        show_latest_bridges = st.toggle(
            "Show every reconciliation for the latest analyzed fiscal period",
            value=False,
            key="show_latest_bridges",
        )
        if show_latest_bridges:
            all_periods = ng.ordered_fiscal_periods(reconciliations)
            latest_period = all_periods[-1] if all_periods else ""
            st.markdown(f"#### All parsed bridges — {latest_period}")
            latest_rows = reconciliations[reconciliations["period"].astype(str).eq(latest_period)].copy()
            for metric_name in sorted(latest_rows["metric"].dropna().astype(str).unique().tolist()):
                metric_bridge = ng.make_reconciliation_bridge_table(
                    reconciliations, adjustment_history, metric_name, [latest_period]
                )
                metric_source_rows = latest_rows[latest_rows["metric"].astype(str).eq(metric_name)]
                metric_source = clean_text(metric_source_rows.iloc[0].get("source_url")) if not metric_source_rows.empty else ""
                render_bridge_table(
                    metric_bridge,
                    f"{title_label} — {metric_name}",
                    latest_period,
                    metric_source,
                )

with tab_metrics:
    st.subheader("Two-year fiscal-quarter metric matrix")
    include_gaap = st.toggle("Show the comparable GAAP value inside each matrix cell", value=False)
    metric_matrix = ng.make_metric_matrix(reconciliations, include_gaap=include_gaap)
    if metric_matrix.empty:
        st.info("No structured reconciliation metrics were extracted.")
    else:
        display_dataframe(metric_matrix, height=min(760, 120 + 35 * len(metric_matrix)))
        st.caption("A blank cell means the measure was not extracted for that fiscal quarter. Values are presented in the issuer's reported units.")

        st.subheader("Quarterly movement")
        detail_view = compact_reconciliation_view(trends)
        display_dataframe(
            detail_view,
            column_config={
                "SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")
            },
            height=min(720, 140 + 35 * len(detail_view)),
        )
        st.caption("For rates, changes are shown in percentage points or basis points. For other measures, changes are percentage changes when a comparable prior value exists.")

        metric_options = sorted(trends["metric"].dropna().astype(str).unique().tolist())
        if metric_options:
            st.subheader("Metric trend")
            selected_metric = st.selectbox("Select a metric", options=metric_options, key="metric_trend_selector")
            trend_chart, chart_unit = chart_frame(trends, selected_metric)
            if not trend_chart.empty:
                st.line_chart(trend_chart, use_container_width=True, color="#86BC25")
                st.caption(f"Chart unit: {chart_unit}. The chart follows issuer fiscal-quarter order, not calendar-quarter order.")

with tab_details:
    st.subheader("GAAP-to-non-GAAP reconciliation detail")
    if reconciliations.empty:
        st.info("No structured reconciliation rows were extracted.")
    else:
        period_options = sorted(reconciliations["period"].dropna().astype(str).unique().tolist(), key=period_sort_key, reverse=True)
        selected_period = st.selectbox("Fiscal period", options=period_options, key="detail_period_selector")
        period_rows = reconciliations[reconciliations["period"].eq(selected_period)].copy()
        metric_options = ["All metrics"] + sorted(period_rows["metric"].dropna().astype(str).unique().tolist())
        selected_detail_metric = st.selectbox("Metric", options=metric_options, key="detail_metric_selector")
        if selected_detail_metric != "All metrics":
            period_rows = period_rows[period_rows["metric"].eq(selected_detail_metric)]

        for _, row in period_rows.sort_values(["metric", "source_role"]).iterrows():
            title = f"{row.get('metric', '')} | {row.get('non_gaap_display', '')}"
            with st.expander(title, expanded=selected_detail_metric != "All metrics"):
                card_columns = st.columns(3)
                with card_columns[0]:
                    render_recon_card("Comparable GAAP", row.get("gaap_display", ""), row.get("gaap_label", ""))
                with card_columns[1]:
                    render_recon_card("Total adjustments", row.get("adjustment_display", ""), "Non-GAAP minus GAAP")
                with card_columns[2]:
                    render_recon_card("Reported non-GAAP", row.get("non_gaap_display", ""), row.get("non_gaap_label", ""))

                pair_id = row.get("pair_id")
                pair_adjustments = (
                    adjustment_history[adjustment_history["pair_id"].eq(pair_id)].copy()
                    if not adjustment_history.empty and "pair_id" in adjustment_history.columns
                    else pd.DataFrame()
                )
                if not pair_adjustments.empty:
                    st.markdown("**Adjustment bridge**")
                    bridge_columns = [
                        "adjustment_category",
                        "adjustment_label",
                        "adjustment_display",
                        "observed_frequency",
                    ]
                    bridge_columns = [column for column in bridge_columns if column in pair_adjustments.columns]
                    bridge = pair_adjustments[bridge_columns].rename(
                        columns={
                            "adjustment_category": "Normalized category",
                            "adjustment_label": "Issuer-reported adjustment",
                            "adjustment_display": "Reported value",
                            "observed_frequency": "Observed across selected periods",
                        }
                    )
                    display_dataframe(bridge)
                else:
                    st.caption("No individual adjustment lines were parsed between the GAAP and non-GAAP rows.")

                source_page = row.get("source_page")
                page_note = f" | PDF page {int(source_page)}" if pd.notna(source_page) else ""
                st.caption(
                    f"{row.get('source_role', '')}{page_note} | Parse confidence: {row.get('confidence', '')} | {row.get('table_title', '')}"
                )
                source_url = clean_text(row.get("source_url"))
                if source_url:
                    st.markdown(f"[Open the SEC source exhibit]({source_url})")

        st.divider()
        st.subheader("Reconciliation evidence")
        if evidence.empty:
            st.info("No table-level evidence records were created.")
        else:
            evidence_period = evidence[evidence["period"].eq(selected_period)].copy()
            for _, row in evidence_period.iterrows():
                page = row.get("source_page")
                page_label = f" | page {int(page)}" if pd.notna(page) else ""
                heading = f"{row.get('source_role', '')} | {row.get('table_title', '')}{page_label}"
                with st.expander(heading):
                    st.write(row.get("text_preview", ""))
                    st.caption(
                        f"Parsed rows: {row.get('parsed_rows', 0)} | Parsed metric pairs: {row.get('parsed_pairs', 0)} | Detection score: {row.get('reconciliation_score', '')}"
                    )
                    if clean_text(row.get("source_url")):
                        st.markdown(f"[Open source]({row.get('source_url')})")

with tab_adjustments:
    st.subheader("Non-GAAP adjustment history")
    st.write(
        "The app retains each issuer-reported adjustment label and also assigns a normalized comparison category so that the same type of adjustment can be followed across fiscal periods."
    )
    st.markdown(
        '<div class="warning-note"><strong>Do not sum adjustment amounts across different non-GAAP metrics.</strong> '
        'The same item can appear in gross margin, operating expense, operating income, net income, and EPS reconciliations. '
        'Use the value matrix for one selected metric; the all-metric matrix shows presence only.</div>',
        unsafe_allow_html=True,
    )

    if adjustment_history.empty:
        st.info("No structured adjustment rows were extracted.")
    else:
        category_summary = ng.adjustment_category_summary(adjustment_history)
        adjustment_periods_ascending = ng.ordered_fiscal_periods(adjustment_history)
        latest_adjustment_period = adjustment_periods_ascending[-1] if adjustment_periods_ascending else ""
        repeated_categories = (
            int(category_summary["periods_observed"].ge(2).sum())
            if not category_summary.empty and "periods_observed" in category_summary.columns
            else 0
        )
        latest_categories = (
            int(
                adjustment_history[adjustment_history["period"].astype(str).eq(latest_adjustment_period)][
                    "adjustment_category"
                ].nunique()
            )
            if latest_adjustment_period
            else 0
        )
        tieout_with_detail = (
            adjustment_tieouts[adjustment_tieouts["tie_out_status"].ne("No line-item detail")]
            if not adjustment_tieouts.empty and "tie_out_status" in adjustment_tieouts.columns
            else pd.DataFrame()
        )
        tieout_passes = (
            int(tieout_with_detail["tie_out_status"].eq("Ties within rounding").sum())
            if not tieout_with_detail.empty
            else 0
        )
        tieout_rate = (
            f"{tieout_passes / len(tieout_with_detail):.0%}"
            if not tieout_with_detail.empty
            else "n/a"
        )

        adjustment_kpis = st.columns(4)
        with adjustment_kpis[0]:
            render_info_card(
                "Normalized categories",
                adjustment_history["adjustment_category"].nunique(),
                "Issuer labels remain available",
            )
        with adjustment_kpis[1]:
            render_info_card(
                "Repeated categories",
                repeated_categories,
                "Observed in at least two selected periods",
            )
        with adjustment_kpis[2]:
            render_info_card(
                "Latest-period categories",
                latest_categories,
                latest_adjustment_period or "No period",
            )
        with adjustment_kpis[3]:
            render_info_card(
                "Bridge tie-out rate",
                tieout_rate,
                f"{tieout_passes} of {len(tieout_with_detail)} bridges with parsed detail" if not tieout_with_detail.empty else "No detailed bridges",
            )

        matrix_tab, changes_tab, records_tab, tieout_tab = st.tabs(
            ["Period matrix", "What changed", "Detailed records", "Tie-out checks"]
        )

        metric_rank = (
            adjustment_history.groupby("metric")
            .agg(periods=("period", "nunique"), rows=("adjustment_label", "size"))
            .sort_values(["periods", "rows"], ascending=False)
        )
        adjustment_metric_options = metric_rank.index.astype(str).tolist()

        with matrix_tab:
            st.markdown("#### Adjustment values by fiscal period")
            selected_adjustment_metric = st.selectbox(
                "Non-GAAP metric",
                options=adjustment_metric_options,
                key="adjustment_matrix_metric",
                help="Amounts are shown within one metric to avoid double counting the same adjustment across several reconciliations.",
            )
            adjustment_matrix = ng.make_adjustment_value_matrix(
                adjustment_history,
                selected_adjustment_metric,
            )
            display_dataframe(
                adjustment_matrix,
                height=min(760, 140 + 36 * len(adjustment_matrix)),
            )
            st.caption(
                "Rows use normalized categories; cells retain the amount and sign reported in the matched 8-K exhibit. A blank cell means that category was not parsed for that metric and period."
            )

            adjustment_chart, adjustment_chart_unit = adjustment_chart_frame(
                adjustment_history,
                selected_adjustment_metric,
            )
            if not adjustment_chart.empty:
                st.markdown("#### Adjustment bridge trend")
                st.bar_chart(
                    adjustment_chart,
                    use_container_width=True,
                    color=chart_series_colors(len(adjustment_chart.columns)),
                )
                st.caption(
                    f"Chart unit: {adjustment_chart_unit}. Up to the eight largest categories by absolute value are shown. Positive and negative bars preserve the issuer's bridge direction."
                )

            with st.expander("Show all-metric category presence"):
                presence_matrix = ng.make_adjustment_presence_matrix(adjustment_history)
                display_dataframe(
                    presence_matrix,
                    height=min(760, 140 + 36 * len(presence_matrix)),
                )
                st.caption(
                    "This matrix counts affected metrics and parsed rows; it deliberately does not add amounts across metrics."
                )

        with changes_tab:
            st.markdown("#### Compare adjustment categories between two fiscal periods")
            comparison_metric = st.selectbox(
                "Non-GAAP metric",
                options=adjustment_metric_options,
                key="adjustment_compare_metric",
            )
            comparison_history = adjustment_history[
                adjustment_history["metric"].astype(str).eq(comparison_metric)
            ]
            comparison_periods = ng.ordered_fiscal_periods(comparison_history)
            if len(comparison_periods) < 2:
                st.info("At least two parsed fiscal periods are needed for a period comparison.")
            else:
                compare_columns = st.columns(2)
                with compare_columns[0]:
                    current_period = st.selectbox(
                        "Current period",
                        options=list(reversed(comparison_periods)),
                        index=0,
                        key="adjustment_current_period",
                    )
                prior_options = [period for period in reversed(comparison_periods) if period != current_period]
                with compare_columns[1]:
                    prior_period = st.selectbox(
                        "Comparison period",
                        options=prior_options,
                        index=0,
                        key="adjustment_prior_period",
                    )
                comparison = ng.compare_adjustment_periods(
                    adjustment_history,
                    comparison_metric,
                    current_period,
                    prior_period,
                )
                if comparison.empty:
                    st.info("No comparable adjustment categories were available.")
                else:
                    change_counts = comparison["status"].value_counts()
                    change_kpis = st.columns(3)
                    with change_kpis[0]:
                        render_info_card(
                            "New",
                            int(change_counts.get("New in current period", 0)),
                            f"Present in {current_period}, absent in {prior_period}",
                        )
                    with change_kpis[1]:
                        render_info_card(
                            "Continued",
                            int(change_counts.get("Continued", 0)),
                            "Present in both selected periods",
                        )
                    with change_kpis[2]:
                        render_info_card(
                            "No longer reported",
                            int(change_counts.get("No longer reported", 0)),
                            f"Present in {prior_period}, absent in {current_period}",
                        )
                    comparison_view = comparison.rename(
                        columns={
                            "status": "Period status",
                            "adjustment_category": "Normalized category",
                            "prior_value": prior_period,
                            "current_value": current_period,
                            "change": "Change",
                            "prior_issuer_labels": f"Issuer labels — {prior_period}",
                            "current_issuer_labels": f"Issuer labels — {current_period}",
                            "observed_periods": "Periods observed",
                            "source_url": "SEC source",
                        }
                    )
                    display_dataframe(
                        comparison_view,
                        column_config={
                            "SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")
                        },
                        height=min(760, 150 + 40 * len(comparison_view)),
                    )
                    st.caption(
                        "'New' and 'no longer reported' describe the two selected disclosures only; they are not conclusions about whether an item is economically recurring or permissible."
                    )

        with records_tab:
            st.markdown("#### Issuer-reported adjustment records")
            filter_columns = st.columns(3)
            with filter_columns[0]:
                record_period = st.selectbox(
                    "Fiscal period",
                    options=["All periods"] + list(reversed(adjustment_periods_ascending)),
                    key="adjustment_record_period",
                )
            with filter_columns[1]:
                record_metric = st.selectbox(
                    "Metric",
                    options=["All metrics"] + adjustment_metric_options,
                    key="adjustment_record_metric",
                )
            category_options = sorted(
                adjustment_history["adjustment_category"].dropna().astype(str).unique().tolist(),
                key=lambda value: (ng.ADJUSTMENT_CATEGORY_ORDER.get(value, 999), value),
            )
            with filter_columns[2]:
                selected_categories = st.multiselect(
                    "Categories",
                    options=category_options,
                    placeholder="All categories",
                    key="adjustment_record_categories",
                )
            record_view = adjustment_history.copy()
            if record_period != "All periods":
                record_view = record_view[record_view["period"].astype(str).eq(record_period)]
            if record_metric != "All metrics":
                record_view = record_view[record_view["metric"].astype(str).eq(record_metric)]
            if selected_categories:
                record_view = record_view[record_view["adjustment_category"].isin(selected_categories)]
            record_view = compact_adjustment_history_view(record_view)
            display_dataframe(
                record_view,
                column_config={
                    "SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")
                },
                height=min(820, 150 + 36 * len(record_view)),
            )

            st.markdown("#### Category dictionary for this issuer")
            summary_view = category_summary.rename(
                columns={
                    "adjustment_category": "Normalized category",
                    "periods_observed": "Periods observed",
                    "first_period": "First selected period",
                    "latest_period": "Latest selected period",
                    "metrics_affected": "Metrics affected",
                    "issuer_labels": "Issuer-reported labels",
                    "recurrence_indicator": "Observed pattern",
                }
            )
            display_dataframe(summary_view, height=min(680, 140 + 38 * len(summary_view)))

        with tieout_tab:
            st.markdown("#### GAAP-to-non-GAAP bridge controls")
            st.write(
                "For each reconciliation, the app compares the sum of parsed line items with the total bridge calculated as non-GAAP minus GAAP. Differences can identify rounding, subtotals, omitted lines, or parsing issues."
            )
            if adjustment_tieouts.empty:
                st.info("No reconciliation tie-out records were created.")
            else:
                tieout_statuses = adjustment_tieouts["tie_out_status"].dropna().astype(str).unique().tolist()
                selected_tieout_status = st.multiselect(
                    "Tie-out status",
                    options=sorted(tieout_statuses),
                    default=sorted(tieout_statuses),
                    key="adjustment_tieout_status",
                )
                filtered_tieouts = adjustment_tieouts[
                    adjustment_tieouts["tie_out_status"].isin(selected_tieout_status)
                ] if selected_tieout_status else adjustment_tieouts.iloc[0:0]
                tieout_display = tieout_view(filtered_tieouts)
                display_dataframe(
                    tieout_display,
                    column_config={
                        "SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")
                    },
                    height=min(820, 150 + 40 * len(tieout_display)),
                )


with tab_additional:
    st.subheader("Other non-GAAP measures discussed in the matched 8-K exhibits")
    st.write(
        "This section calls out measures mentioned in the earnings release, financial supplement, or investor presentation that were not matched to a structured GAAP-to-non-GAAP pair anywhere in the selected earnings 8-K exhibits."
    )
    if additional.empty:
        st.info("No additional non-GAAP measures were identified beyond the structured reconciliation metrics.")
    else:
        callout_view = additional.copy()
        callout_view["Where found"] = callout_view.apply(
            lambda row: "Document containing a reconciliation"
            if bool(row.get("source_has_reconciliation"))
            else ("Investor presentation" if row.get("source_role") == "Investor presentation" else "Other matched EX-99 exhibit"),
            axis=1,
        )
        columns = [
            "period",
            "metric",
            "primary_value",
            "Where found",
            "source_role",
            "source_content_type",
            "source_page",
            "source_url",
        ]
        columns = [column for column in columns if column in callout_view.columns]
        table = callout_view[columns].rename(
            columns={
                "period": "Fiscal period",
                "metric": "Additional measure",
                "primary_value": "Nearby reported value",
                "source_role": "Source type",
                "source_content_type": "Format",
                "source_page": "PDF page",
                "source_url": "SEC source",
            }
        )
        display_dataframe(
            table,
            column_config={
                "SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")
            },
            height=min(650, 140 + 35 * len(table)),
        )

        st.subheader("Source context")
        for _, row in callout_view.sort_values(["period", "metric"], ascending=[False, True]).iterrows():
            value_label = f" | {row.get('primary_value')}" if clean_text(row.get("primary_value")) else ""
            with st.expander(f"{row.get('period', '')} | {row.get('metric', '')}{value_label}"):
                st.write(row.get("context", ""))
                st.caption(
                    f"Source: {row.get('source_role', '')} ({row.get('source_content_type', '')}) | Document: {row.get('source_document', '')}"
                )
                if clean_text(row.get("source_url")):
                    st.markdown(f"[Open source]({row.get('source_url')})")

with tab_peer:
    st.subheader("Peer benchmarking from earnings 8-K exhibit packages")
    st.write(
        "Build presentation-ready disclosure matrices for non-GAAP measures, adjustment types, and operating KPIs, "
        "then open comparable GAAP-to-non-GAAP bridges for each selected company. Each peer is normalized to its own fiscal year end."
    )
    st.markdown(
        '<div class="peer-note"><strong>Scope:</strong> Peer metrics, adjustments, and KPIs come from matched earnings 8-K EX-99 exhibits, '
        'including investor-presentation PDFs. Periodic filings are used only to anchor fiscal periods.</div>',
        unsafe_allow_html=True,
    )

    current_ticker = _first_ticker(company)
    default_peer_text = current_ticker
    peer_text = st.text_area(
        "Peer tickers (comma, space, or one per line; maximum eight)",
        value=default_peer_text,
        placeholder="LSCC, MCHP, AMD, MRVL, QUIK",
        key="peer_ticker_input",
        help="The current issuer can be reused without a second SEC analysis. Add public-company tickers for a live peer set.",
    )
    peer_control_columns = st.columns([1, 1, 2])
    with peer_control_columns[0]:
        peer_max_exhibits = st.number_input(
            "EX-99 exhibits per peer",
            min_value=3,
            max_value=10,
            value=6,
            step=1,
            key="peer_max_exhibits",
        )
    with peer_control_columns[1]:
        minimum_peer_disclosures = st.number_input(
            "Minimum peers per row",
            min_value=1,
            max_value=8,
            value=1,
            step=1,
            key="peer_minimum_disclosures",
        )
    with peer_control_columns[2]:
        st.caption(
            "Peer runs can take several minutes because the app retrieves each issuer's fiscal anchors, matching 8-K, press release, supplement, and presentation."
        )

    run_peer_analysis = st.button(
        "Run peer benchmark",
        type="primary",
        use_container_width=True,
        key="run_peer_benchmark",
    )
    if run_peer_analysis:
        raw_tickers = [
            token.upper()
            for token in re.split(r"[\s,;]+", peer_text or "")
            if clean_text(token)
        ]
        tickers: list[str] = []
        for ticker in raw_tickers:
            if ticker not in tickers:
                tickers.append(ticker)
        if current_ticker and current_ticker not in tickers:
            tickers.insert(0, current_ticker)
        tickers = tickers[:8]

        if not ng.SecClient.valid_contact(contact_email):
            st.error("Enter a valid SEC contact email before running the peer benchmark.")
        elif not tickers:
            st.error("Enter at least one ticker.")
        else:
            peer_results: list[tuple[dict[str, Any], dict[str, pd.DataFrame]]] = []
            peer_company_rows: list[dict[str, Any]] = []
            peer_errors: list[dict[str, str]] = []
            peer_status = st.status("Starting peer benchmark...", expanded=True)
            client = get_client(contact_email.strip())

            for peer_index, ticker in enumerate(tickers, start=1):
                try:
                    peer_status.update(
                        label=f"Peer {peer_index} of {len(tickers)}: resolving {ticker}...",
                        state="running",
                    )
                    if ticker == current_ticker:
                        peer_company = dict(company)
                        peer_result = analysis
                    else:
                        resolved = resolve_exact_ticker(client, ticker)
                        if resolved is None:
                            raise ValueError("Ticker was not found in the SEC company-ticker universe.")

                        def peer_progress(message: str, _ticker: str = ticker) -> None:
                            peer_status.update(label=f"{_ticker}: {message}", state="running")

                        peer_company, peer_result = analyze_peer_company(
                            client,
                            resolved,
                            progress=peer_progress,
                            max_exhibits=int(peer_max_exhibits),
                        )
                    peer_results.append((peer_company, peer_result))
                    peer_company_rows.append(
                        {
                            "Ticker": _first_ticker(peer_company),
                            "Company": clean_text(peer_company.get("name")),
                            "CIK": peer_company.get("cik"),
                            "SIC": peer_company.get("sic"),
                            "Fiscal year end": fye_display(clean_text(peer_company.get("fiscal_year_end"))),
                            "Structured reconciliations": len(peer_result.get("reconciliations", pd.DataFrame())),
                            "Parsed adjustment rows": len(peer_result.get("adjustment_history", pd.DataFrame())),
                        }
                    )
                except Exception as exc:
                    peer_errors.append({"Ticker": ticker, "Error": clean_text(exc)})

            st.session_state.peer_analysis = combine_peer_results(peer_results) if peer_results else None
            st.session_state.peer_companies = pd.DataFrame(peer_company_rows)
            st.session_state.peer_errors = peer_errors
            if peer_results:
                peer_status.update(
                    label=f"Peer benchmark complete for {len(peer_results)} company(ies).",
                    state="complete",
                    expanded=False,
                )
            else:
                peer_status.update(
                    label="No peer analyses completed.",
                    state="error",
                    expanded=True,
                )

    peer_analysis = st.session_state.peer_analysis
    if not peer_analysis:
        st.info("Enter peer tickers and run the benchmark to create the disclosure matrices and company reconciliation tables.")
    else:
        peer_companies = st.session_state.peer_companies
        if isinstance(peer_companies, pd.DataFrame) and not peer_companies.empty:
            st.markdown("#### Companies included")
            display_dataframe(peer_companies, height=min(360, 90 + 36 * len(peer_companies)))

        if st.session_state.peer_errors:
            with st.expander(f"Peer warnings ({len(st.session_state.peer_errors)})", expanded=False):
                display_dataframe(pd.DataFrame(st.session_state.peer_errors))

        peer_download = ng.build_export_zip(peer_analysis)
        st.download_button(
            "Download peer benchmark CSV package",
            data=peer_download,
            file_name="non_gaap_peer_benchmark.csv.zip",
            mime="application/zip",
            key="download_peer_benchmark",
        )

        peer_measure_tab, peer_adjustment_tab, peer_kpi_tab, peer_bridge_tab = st.tabs(
            ["Non-GAAP measures", "Adjustment types", "Operating KPIs", "Detailed peer bridges"]
        )

        with peer_measure_tab:
            measure_source = build_measure_presence_source(peer_analysis)
            measure_matrix_all = ng.make_peer_presence_matrix(
                measure_source,
                row_field="metric_family",
                ordered_rows=ng.BENCHMARK_MEASURE_ORDER,
                minimum_companies=1,
            )
            measure_matrix = ng.make_peer_presence_matrix(
                measure_source,
                row_field="metric_family",
                ordered_rows=ng.BENCHMARK_MEASURE_ORDER,
                minimum_companies=int(minimum_peer_disclosures),
            )
            st.markdown("#### Non-GAAP measures | Peer benchmarking")
            if measure_matrix.empty:
                st.info("No peer measure matrix could be created from the selected analyses.")
            else:
                display_dataframe(measure_matrix, height=min(720, 120 + 36 * len(measure_matrix)))
                highlight = peer_matrix_highlight(measure_matrix, "measure")
                if highlight:
                    st.markdown(highlight)
                measure_counts = peer_company_count_frame(
                    measure_matrix_all, "Number of non-GAAP measures"
                )
                st.markdown("##### Peers by number of non-GAAP measure disclosures")
                st.bar_chart(measure_counts, use_container_width=True, color="#86BC25")

        with peer_adjustment_tab:
            peer_adjustments = peer_analysis.get("adjustment_history", pd.DataFrame())
            st.markdown("#### Non-GAAP adjustment types | Peer benchmarking")
            if peer_adjustments.empty:
                st.info("No structured peer adjustment rows were extracted.")
            else:
                adjustment_order_rows = [
                    name for name, _rank in sorted(ng.ADJUSTMENT_CATEGORY_ORDER.items(), key=lambda item: item[1])
                ]
                adjustment_matrix_peer_all = ng.make_peer_presence_matrix(
                    peer_adjustments,
                    row_field="adjustment_category",
                    ordered_rows=adjustment_order_rows,
                    minimum_companies=1,
                )
                adjustment_matrix_peer = ng.make_peer_presence_matrix(
                    peer_adjustments,
                    row_field="adjustment_category",
                    ordered_rows=adjustment_order_rows,
                    minimum_companies=int(minimum_peer_disclosures),
                )
                display_dataframe(
                    adjustment_matrix_peer,
                    height=min(720, 120 + 36 * len(adjustment_matrix_peer)),
                )
                highlight = peer_matrix_highlight(adjustment_matrix_peer, "adjustment type")
                if highlight:
                    st.markdown(highlight)
                if not adjustment_matrix_peer.empty:
                    adjustment_counts = peer_company_count_frame(
                        adjustment_matrix_peer_all, "Number of adjustment types"
                    )
                    st.markdown("##### Peers by number of non-GAAP adjustment types")
                    st.bar_chart(adjustment_counts, use_container_width=True, color="#86BC25")
                st.caption(
                    "A dot means the adjustment type appeared in at least one parsed reconciliation for that company. "
                    "Counts are disclosure presence, not additive amounts."
                )

        with peer_kpi_tab:
            peer_kpis = peer_analysis.get("kpis", pd.DataFrame())
            st.markdown("#### Key performance indicators | Peer benchmarking")
            if peer_kpis.empty:
                st.info("No KPI mentions were extracted from the selected earnings 8-K exhibit packages.")
            else:
                kpi_matrix_all = ng.make_peer_presence_matrix(
                    peer_kpis,
                    row_field="kpi",
                    ordered_rows=[label for label, _ in ng.KPI_PATTERNS],
                    minimum_companies=1,
                )
                kpi_matrix = ng.make_peer_presence_matrix(
                    peer_kpis,
                    row_field="kpi",
                    ordered_rows=[label for label, _ in ng.KPI_PATTERNS],
                    minimum_companies=int(minimum_peer_disclosures),
                )
                display_dataframe(kpi_matrix, height=min(720, 120 + 36 * len(kpi_matrix)))
                highlight = peer_matrix_highlight(kpi_matrix, "KPI")
                if highlight:
                    st.markdown(highlight)
                if not kpi_matrix.empty:
                    kpi_counts = peer_company_count_frame(kpi_matrix_all, "Number of KPIs")
                    st.markdown("##### Peers by number of KPI disclosures")
                    st.bar_chart(kpi_counts, use_container_width=True, color="#86BC25")
                with st.expander("KPI source context", expanded=False):
                    kpi_columns = [
                        column
                        for column in ["company", "period", "kpi", "primary_value", "source_role", "source_url", "context"]
                        if column in peer_kpis.columns
                    ]
                    kpi_view = peer_kpis[kpi_columns].rename(
                        columns={
                            "company": "Peer",
                            "period": "Fiscal period",
                            "kpi": "KPI",
                            "primary_value": "Nearby value",
                            "source_role": "Source type",
                            "source_url": "SEC source",
                            "context": "Disclosure context",
                        }
                    )
                    display_dataframe(
                        kpi_view,
                        column_config={"SEC source": st.column_config.LinkColumn("SEC source", display_text="Open exhibit")},
                        height=min(720, 140 + 36 * len(kpi_view)),
                    )

        with peer_bridge_tab:
            peer_reconciliations = peer_analysis.get("reconciliations", pd.DataFrame())
            peer_adjustment_history = peer_analysis.get("adjustment_history", pd.DataFrame())
            st.markdown("#### Comparable reconciliation tables by peer")
            if peer_reconciliations.empty:
                st.info("No structured peer reconciliations were available.")
            else:
                if "metric_family" not in peer_reconciliations.columns:
                    peer_reconciliations = peer_reconciliations.copy()
                    peer_reconciliations["metric_family"] = peer_reconciliations.apply(
                        lambda row: ng.benchmark_metric_family(
                            row.get("metric", ""), row.get("gaap_label", ""), row.get("non_gaap_label", "")
                        ),
                        axis=1,
                    )
                family_options = sorted(
                    peer_reconciliations["metric_family"].dropna().astype(str).unique().tolist()
                )
                preferred_family_index = next(
                    (index for index, value in enumerate(family_options) if "gross margin" in value.lower()),
                    0,
                )
                selected_family = st.selectbox(
                    "Comparable non-GAAP measure family",
                    options=family_options,
                    index=preferred_family_index,
                    key="peer_bridge_family",
                )
                family_rows = peer_reconciliations[
                    peer_reconciliations["metric_family"].astype(str).eq(selected_family)
                ].copy()
                companies_in_family = sorted(family_rows["company"].dropna().astype(str).unique().tolist())
                st.caption(
                    "Each table keeps the issuer's exact labels and its two latest analyzed fiscal-period columns. "
                    "The selected family only aligns comparable disclosure types across companies."
                )
                for peer_label in companies_in_family:
                    company_rows = family_rows[family_rows["company"].astype(str).eq(peer_label)].copy()
                    metric_rank = (
                        company_rows.groupby("metric")
                        .agg(periods=("period", "nunique"), rows=("pair_id", "size"))
                        .sort_values(["periods", "rows"], ascending=False)
                    )
                    if metric_rank.empty:
                        continue
                    peer_metric = str(metric_rank.index[0])
                    metric_rows = company_rows[company_rows["metric"].astype(str).eq(peer_metric)].copy()
                    peer_periods = ng.ordered_fiscal_periods(metric_rows)[-2:]
                    peer_adjustments_for_company = (
                        peer_adjustment_history[
                            peer_adjustment_history["company"].astype(str).eq(peer_label)
                        ].copy()
                        if not peer_adjustment_history.empty and "company" in peer_adjustment_history.columns
                        else pd.DataFrame()
                    )
                    bridge = ng.make_reconciliation_bridge_table(
                        metric_rows,
                        peer_adjustments_for_company,
                        peer_metric,
                        peer_periods,
                    )
                    source_url = clean_text(metric_rows.sort_values(["fiscal_year", "fiscal_quarter"]).iloc[-1].get("source_url"))
                    company_name = clean_text(metric_rows.iloc[0].get("company_name")) or peer_label
                    render_bridge_table(
                        bridge,
                        f"{company_name} ({peer_label})",
                        f"{selected_family} | issuer label: {peer_metric}",
                        source_url,
                    )


with tab_sources:
    st.subheader("Fiscal-quarter and 8-K coverage")
    coverage_table = coverage_view(coverage)
    display_dataframe(
        coverage_table,
        column_config={
            "Earnings 8-K": st.column_config.LinkColumn("Earnings 8-K", display_text="Open 8-K"),
            "Anchor 10-Q/10-K": st.column_config.LinkColumn("Anchor 10-Q/10-K", display_text="Open anchor"),
        },
        height=min(620, 150 + 38 * len(coverage_table)) if not coverage_table.empty else None,
    )
    st.caption(
        "The anchor filing determines the issuer fiscal period. The matched Item 2.02 earnings 8-K and its EX-99 exhibits are the only metric sources."
    )

    st.subheader("Documents checked")
    source_table = source_audit_view(sources)
    display_dataframe(
        source_table,
        column_config={
            "Exhibit source": st.column_config.LinkColumn("Exhibit source", display_text="Open exhibit"),
            "8-K filing index": st.column_config.LinkColumn("8-K filing index", display_text="Open filing index"),
        },
        height=min(650, 150 + 36 * len(source_table)) if not source_table.empty else None,
    )

    st.subheader("Parsing warnings")
    if warnings.empty:
        st.success("No parsing warnings were recorded.")
    else:
        warning_view = warnings.copy()
        if "source_url" in warning_view.columns:
            display_dataframe(
                warning_view,
                column_config={
                    "source_url": st.column_config.LinkColumn("Source", display_text="Open source")
                },
            )
        else:
            display_dataframe(warning_view)
        st.caption("Image-only PDFs are flagged because this version does not perform OCR. Review the linked presentation manually when that warning appears.")

    with st.expander("Methodology and limitations", expanded=False):
        st.markdown(
            """
            **Matching logic**

            1. Read recent 10-Q and 10-K filings to obtain `DocumentFiscalYearFocus`, `DocumentFiscalPeriodFocus`, and period-end dates when available.
            2. For each fiscal quarter, score nearby 8-K/8-K/A filings using Item 2.02, Item 9.01, timing, earnings language, period references, and the presence of press-release or presentation exhibits.
            3. Inspect relevant EX-99 HTML, text, and PDF exhibits. Extract a structured pair only where a comparable GAAP row and a non-GAAP/adjusted row can be associated in a reconciliation section.
            4. Record individual reconciling items between those rows, retain the issuer's exact label, and assign a normalized comparison category.
            5. Compare categories across issuer fiscal periods and test whether parsed line items tie to non-GAAP minus GAAP.
            6. Separately call out additional non-GAAP measures discussed in the matched exhibits.

            **Review points**

            - Issuer formats vary, so every result retains a direct SEC source link and parse-confidence label.
            - Adjustment categories and new/continued/no-longer-reported labels are descriptive aids, not conclusions about whether an item is permissible or economically recurring.
            - Do not sum adjustment amounts across different non-GAAP metrics because the same item may be repeated in several reconciliations.
            - The structured value is the first reported numeric column in the matched reconciliation section, which is normally the current fiscal quarter for an earnings release.
            - PDF text extraction depends on an embedded text layer. Image-only slides are identified in the warning log and require manual review.
            - Non-GAAP measures are not standardized across issuers; metric names and definitions should be compared with the source disclosure.
            """
        )
